"""Attendance layout parsers.

Each parser recognises one spreadsheet *layout* (``matches``) and ingests it into the
attendance database (``parse``). Parsers map to layouts, not years: one workbook may
mix layouts across tabs, and the same layout recurs unpredictably across years. The
folder dispatcher (in ``ingest``) tries each parser per sheet and reports any sheet no
parser claimed (those would need a new parser).

All input is read from the pseudonymised ``attendance_outputs`` tree. The layouts supported
so far:

* **Roster** — one row per dancer, a boolean per session column (date in the row above
  the header). Level 1 tabs.
* **Level 2 tally** — a checkbox grid summed by COUNTIF into per-(week, ticket-type)
  headcounts, for two activity groups (Level 2 Classes / Social Only).
* **Level 2 count grid** — the older 2024 'Levels 1-2' layout: a week-per-row table with
  the per-ticket-type headcounts typed straight in (no checkboxes, no COUNTIF).
* **L2 & SO attendance** — the 2026 tabs whose session cells hold a ticket *category*
  ('Level 2 & Social' / 'Social-Only' / 'Absent') rather than a yes/no marker.
* **Social register** — a one-off social's named register (Tea Dances, Sunday Socials, parties):
  one dated social, attendees in one or more side-by-side blocks with a Present?/Attended marker
  and an optional Concession flag; the date comes from the tab name or filename.
* **Booking export** — a dancecloud booking export listing who *bought a ticket*, not who
  turned up; every booking is recorded UNKNOWN unless a status/present note says otherwise.

Every concrete parser subclasses :class:`Parser`, which fixes the ``name``/``matches``/``parse``
contract the dispatcher relies on. The module-level helpers are shared across parsers, and the
canonical event/activity naming helpers (``_event_name`` etc.) are deliberately shared so the
several sources describing one term converge onto a single event.
"""

import abc
import re
from collections import defaultdict
from dataclasses import dataclass, field
from datetime import datetime, timedelta

from openpyxl.utils import column_index_from_string, get_column_letter

from esds_apps.attendance.attendance_db import (
    ActivityType,
    AttendanceDb,
    AttendanceStatus,
    EventType,
    TicketType,
)

_TICKET_LABELS = {
    'members': TicketType.MEMBER,
    'concessions': TicketType.CONCESSION,
    'non-members': TicketType.ORDINARY,
    'non members': TicketType.ORDINARY,
}
_WEEK_RE = re.compile(r'Week\s*(\d+)\s*\(\s*(\d{1,2})\s+([A-Za-z]+)\)', re.IGNORECASE)
# A bare 'Week N' label (no '(DD Mon)' suffix): older sheets carry only the week number.
_WEEK_NO_RE = re.compile(r'\bweek\s*(\d+)\b', re.IGNORECASE)
_COUNTIF_RE = re.compile(r'COUNTIF\(\s*([A-Za-z]+)(\d+)\s*:\s*([A-Za-z]+)(\d+)', re.IGNORECASE)
# ESDS ran Level 2 and Level 3 together; normalise the various spellings to plain 'Level 2'.
_LEVEL_2_3_RE = re.compile(r'level\s*2\s*[/&-]\s*3\b', re.IGNORECASE)


def _is_group_header(text: str) -> bool:
    """True for a tally activity-group header.

    Tolerates the singular/plural 'Level 1 Class(es)' / 'Level 2 Class(es)' and the 'Social Only'
    group. The one-off tallies (e.g. Teacher's Choice) carry all three side by side; the weekly
    Level 2 tallies carry only 'Level 2 Classes' and 'Social Only'.
    """
    t = text.strip().lower()
    return t == 'social only' or t.startswith('level 1 class') or t.startswith('level 2 class')


# ---------------------------------------------------------------------------
# Small value helpers
# ---------------------------------------------------------------------------


# Markers meaning "attended". Rosters vary by year: booleans (True), 'True'/'TRUE'
# strings, the Excel formula '=TRUE()' (read as text because we load formulas, not cached
# values), ticks ('x', 'yes'), and check emoji (☑️/✅/✔/✓). Crosses (❎/❌), 'False',
# '=FALSE()' and 'Refunded' fall through to False — a refunded ticket is a non-attendance.
_TRUE_WORDS = {'true', 'yes', 'y', 'x', '1', '✓', '✔'}
_TRUE_EMOJI = ('☑', '✅', '✔', '✓')  # ☑ ✅ ✔ ✓
# The negative side of the same vocabulary, plus crosses. Used only to tell an attendance
# roster apart from a richer layout whose cells hold categories ('Level 2 & Social').
_FALSE_WORDS = {'false', 'no', 'n', '0', 'refunded', 'absent'}
_MARKER_EMOJI = (*_TRUE_EMOJI, '❎', '❌')


def _is_true(value) -> bool:
    if value is True:
        return True
    if isinstance(value, str):
        # Normalise the '=TRUE()' / '=FALSE()' formula form down to 'true' / 'false'.
        s = value.strip().lower().removeprefix('=').removesuffix('()')
        return s.startswith('true') or s in _TRUE_WORDS or any(e in value for e in _TRUE_EMOJI)
    return False


def _present_status(present: bool) -> AttendanceStatus:
    """Map a present/absent register reading to an AttendanceStatus.

    These layouts are genuine attendance registers — a marked cell means present, a blank
    means the dancer was expected (enrolled or booked) but didn't come — so the only two
    outcomes are ATTENDED and ABSENT. The ticket-bought-but-turnout-unknown case (UNKNOWN)
    comes from sources that record bookings without any attendance mark, handled elsewhere.
    """
    return AttendanceStatus.ATTENDED if present else AttendanceStatus.ABSENT


# Booking exports (dancecloud) list who *bought a ticket*, not who turned up. Where a 'present'
# note column exists it is the only attendance signal: these phrases mean the booker did not come
# (a no-show, or a refund — including a ticket refunded because it was resold on the door)...
_BOOKING_ABSENT_PRESENT = ('no show', 'no-show', 'noshow')
# ...and these mean they were demonstrably there (paid / cash / at the door).
_BOOKING_ATTENDED_PRESENT = ('door', 'cash', 'paid')


def _booking_status(status_value, present_value=None) -> AttendanceStatus:
    """Classify one booking row's turnout from its booking ``Status`` and optional ``present`` note.

    Booking exports record purchases, so the default is UNKNOWN: a ticket was bought but whether
    the dancer attended was never captured. A ``present`` note overrides where present — a no-show
    or refund is ABSENT, a 'paid cash on the door' is ATTENDED — and a Cancelled booking is ABSENT
    (interest registered, then pulled out). Refund is tested before the door words so 'refunded as
    resold on door' reads as the booker's absence, not as an attendance.
    """
    present = present_value.strip().lower() if isinstance(present_value, str) else ''
    if present:
        if present.startswith('refund') or any(w in present for w in _BOOKING_ABSENT_PRESENT):
            return AttendanceStatus.ABSENT
        if _is_true(present_value) or any(w in present for w in _BOOKING_ATTENDED_PRESENT):
            return AttendanceStatus.ATTENDED
    status = status_value.strip().lower() if isinstance(status_value, str) else ''
    if status.startswith('cancel'):
        return AttendanceStatus.ABSENT
    return AttendanceStatus.UNKNOWN


def _is_checked_in(value) -> bool:
    """True if a 'Checked In' cell records a door check-in.

    On the modern dancecloud 'Attendees By Activity' tab this cell is a check-in *timestamp*
    when the dancer was scanned in and blank when they were not; the pivoted 'Check-Ins' tab
    instead uses 'Yes'/'No'/'n/a'. Any timestamp or 'Yes' counts as present; a blank, 'No', or
    'n/a' (the activity was not on their ticket) counts as not-present.
    """
    if isinstance(value, datetime):
        return True
    if isinstance(value, str):
        s = value.strip().lower()
        return bool(s) and s not in ('no', 'n/a', 'na', 'false')
    return False


def _checkin_status(checkin_value, status_value, sheet_has_checkins: bool) -> AttendanceStatus:
    """Attendance status for one modern dancecloud booking row from its check-in and booking Status.

    A 'Cancelled' booking is ABSENT (the dancer pulled out). Otherwise a recorded check-in is
    ATTENDED. A blank check-in is ABSENT when the door scanner ran at all on the sheet (someone,
    somewhere, is checked in) but UNKNOWN when no row carries a check-in — an event whose turnout
    was never scanned, where a blank means 'uncaptured', not 'no-show'. Mirrors the roster's
    column-marked rule.
    """
    status = status_value.strip().lower() if isinstance(status_value, str) else ''
    if status.startswith('cancel'):
        return AttendanceStatus.ABSENT
    if _is_checked_in(checkin_value):
        return AttendanceStatus.ATTENDED
    return AttendanceStatus.ABSENT if sheet_has_checkins else AttendanceStatus.UNKNOWN


def _is_attendance_marker(value) -> bool:
    """True if a cell value is a yes/no attendance marker (not a free-text category)."""
    if isinstance(value, bool):
        return True
    if isinstance(value, str):
        s = value.strip().lower().removeprefix('=').removesuffix('()')
        return (
            s.startswith(('true', 'false'))
            or s in _TRUE_WORDS
            or s in _FALSE_WORDS
            or any(e in value for e in _MARKER_EMOJI)
        )
    return False


def _parse_dt(value) -> datetime | None:
    """Parse a cell value into a datetime, or None.

    Handles real datetimes and the ISO-ish strings the pseudonymiser leaves
    behind ('2025-05-22 00:00:00').
    """
    if isinstance(value, datetime):
        return value
    if isinstance(value, str):
        try:
            return datetime.fromisoformat(value.strip())
        except ValueError:
            return None
    return None


def _month_number(name: str) -> int | None:
    for fmt in ('%b', '%B'):
        try:
            return datetime.strptime(name.strip()[:3] if fmt == '%b' else name.strip(), fmt).month
        except ValueError:
            continue
    return None


# A date written into a one-off event's title or tab name: 'Tea Dance 25 Feb 2024',
# '... 28th April 2024', or '26 March 23'. Day may carry an ordinal suffix; month short or
# full; year 2- or 4-digit. Plus a plain ISO 'YYYY-MM-DD' as it appears in filenames.
_TITLE_DATE_RE = re.compile(r'(\d{1,2})(?:st|nd|rd|th)?\s+([A-Za-z]+)\s+(\d{2,4})', re.IGNORECASE)
_ISO_DATE_RE = re.compile(r'(\d{4})-(\d{2})-(\d{2})')
_CENTURY = 100  # 2-digit years (e.g. '23') are read as 2000 + year

# Booking-export filenames end in the export timestamp or the event date, e.g. '... Term A
# 2023-01-17 2255' or '... Chevaliers 14 Dec 2023' (sometimes with a note like 'members noted').
# Strip that tail so several re-exports of one course collapse into a single event; the real
# per-session dates live on the activities, not in the event name.
_EXPORT_SUFFIX_RE = re.compile(
    r'\s+(?:\d{4}-\d{2}-\d{2}|\d{1,2}(?:st|nd|rd|th)?\s+[A-Za-z]+\s+\d{2,4})\b.*$',
    re.IGNORECASE,
)


def _date_from_title(text: str) -> datetime | None:
    """Parse a 'DD[th] Month YY(YY)' or ISO 'YYYY-MM-DD' date out of a title, or None."""
    m = _TITLE_DATE_RE.search(text)
    if m and (month := _month_number(m.group(2))):
        year = int(m.group(3))
        return datetime(year + 2000 if year < _CENTURY else year, month, int(m.group(1)))
    m = _ISO_DATE_RE.search(text)
    return datetime(int(m.group(1)), int(m.group(2)), int(m.group(3))) if m else None


# A 'DD[th] Month' day+month with no year, as on a one-off tally's tab name ('Teacher's Choice 12th
# Dec'). The year is missing from the tab, so it is supplied separately (the workbook's resolved year).
_DAY_MONTH_RE = re.compile(r'(\d{1,2})(?:st|nd|rd|th)?\s+([A-Za-z]+)')


def _date_from_day_month(text: str, year: int | None) -> datetime | None:
    """Parse a year-less 'DD[th] Month' out of a title and combine it with ``year``, or None."""
    if year is None:
        return None
    m = _DAY_MONTH_RE.search(text)
    if m and (month := _month_number(m.group(2))):
        return datetime(year, month, int(m.group(1)))
    return None


def _activity_type_for(label: str, event_type: EventType) -> ActivityType:
    """An activity's type. A social event holds only socials; otherwise the label decides.

    Session-column headers ('Christmas Party', 'Tea Dance', 'Registered') rarely contain
    the word 'social', so without the event-type override they'd wrongly read as lessons.
    """
    if event_type == EventType.SOCIAL:
        return ActivityType.SOCIAL
    return ActivityType.SOCIAL if 'social' in label.lower() else ActivityType.LESSON


# Free-text activity-label vocabulary for the modern dancecloud export, whose labels are full
# names ('30th Birthday Ball', 'Friday Welcome Party', "Tea dance with ...", 'Track A - Classes',
# 'Collegiate Shag Workshop'). The bare 'social' substring test in _activity_type_for is too weak
# for these: a social is recognised by its party/ball/tea-dance words, a lesson by its
# class/workshop/track words.
_SOCIAL_NAME_RE = re.compile(r'\b(social|party|ball|tea\s*dance|welcome|dance night)\b', re.IGNORECASE)
_LESSON_NAME_RE = re.compile(r'\b(class(?:es)?|lesson|workshop|track|technique|taster|practice)\b', re.IGNORECASE)


def _activity_type_from_name(label: str, event_type: EventType) -> ActivityType:
    """Classify a free-text dancecloud activity label as a social or a lesson.

    Reads the label's own words first — a 'Tea Dance'/'... Party'/'... Ball' is a social, a
    'Workshop'/'Classes'/'Track ...' is a lesson — and falls back to ``_activity_type_for`` (the
    event-type default) when neither vocabulary matches.
    """
    if _SOCIAL_NAME_RE.search(label):
        return ActivityType.SOCIAL
    if _LESSON_NAME_RE.search(label):
        return ActivityType.LESSON
    return _activity_type_for(label, event_type)


def _event_type_for(title: str) -> EventType:
    """Classify an event from keywords in its tab title.

    'level' wins first: the termly "Level 2 & Social Only" tabs contain the word
    "social" but are courses. Workshops next, then socials; anything unmatched
    (Teachers Choice, plain Level N) is a course.
    """
    t = title.lower()
    if 'level' in t:
        return EventType.COURSE
    if 'swingout' in t or 'swing out' in t:
        return EventType.WEEKENDER
    if 'workshop' in t or 'charleston' in t:
        return EventType.WORKSHOP
    if any(k in t for k in ('social', 'party', 'tea dance')):
        return EventType.SOCIAL
    return EventType.COURSE


# Free-text difficulty parsed from a name. Ordered: first pattern to match wins. Lives on
# the activity (not the event) because weekenders/workshops mix levels within one event.
_DIFFICULTY_PATTERNS = [
    (re.compile(r'level\s*1\b|\bl1\b', re.IGNORECASE), 'Level 1'),
    (re.compile(r'level\s*2\b|\bl2\b', re.IGNORECASE), 'Level 2'),
    (re.compile(r'level\s*3\b|\bl3\b', re.IGNORECASE), 'Level 3'),
    (re.compile(r'\bbeg', re.IGNORECASE), 'beginners'),  # beginner(s), 'Begn'
    (re.compile(r'\binterm', re.IGNORECASE), 'intermediate'),
    (re.compile(r'\badv', re.IGNORECASE), 'advanced'),
]


def _difficulty_for(*texts: str | None) -> str | None:
    """Difficulty string ('Level 1', 'beginners', ...) from the first text that matches.

    Texts are tried most-specific first (activity label, then the sheet title), so a
    workshop column headed 'Beginners' beats the tab's overall name. None if unknown.
    """
    for text in texts:
        if not text:
            continue
        for pattern, label in _DIFFICULTY_PATTERNS:
            if pattern.search(text):
                return label
    return None


# ---------------------------------------------------------------------------
# Roster layout helpers
# ---------------------------------------------------------------------------


def _header_columns(header: tuple) -> dict:
    """Map each lower-cased header label to its first column index (shared by the table parsers)."""
    out: dict[str, int] = {}
    for c, v in enumerate(header):
        if isinstance(v, str) and v.strip():
            out.setdefault(v.strip().lower(), c)
    return out


def _roster_header(matrix: list[tuple]) -> tuple[int, int] | None:
    """Return (header_row, dancer_id_col) as 0-based indices, or None."""
    for r, row in enumerate(matrix):
        for c, value in enumerate(row):
            if isinstance(value, str) and value.strip() == 'dancer_id':
                return r, c
    return None


# Tolerant of the 'Consession' misspelling seen in the wild (Jan-Feb 2026) and a trailing plural.
_CONCESSION_HEADER_RE = re.compile(r'con[cs]essions?', re.IGNORECASE)


def _concession_col(header: tuple) -> int | None:
    """Column index of a roster's boolean-ish 'Concession' flag in its header row, or None."""
    for c, value in enumerate(header):
        if isinstance(value, str) and _CONCESSION_HEADER_RE.fullmatch(value.strip()):
            return c
    return None


def _session_columns(
    matrix: list[tuple], header_row: int, year: int | None = None, week_anchor: datetime | None = None
) -> list[tuple]:
    """Session columns: a date either in the header row itself or the row directly above it.

    Older Level 1 tabs put the date in the row above a 'Week N' label; the 2026 tabs put
    the date straight in the header row alongside 'dancer_id'. Two further variants carry no
    real date cell, only a 'Week N' header label: a 'Week N (DD Mon)' label dates itself once a
    ``year`` is known (the early-2022 'Class List' rosters), while a bare 'Week N' is placed at
    ``week_anchor + 7*(N-1)``. A 'Week N' column is recognised as a session even when neither the
    year nor the anchor is available — so layout detection (``matches``) sees it — with its date
    left None; ``parse`` is always given the year/anchor and skips any column still left undated.
    Returns a list of (col_index, activity_name, datetime|None); activity type is decided in parse.
    """
    head = matrix[header_row]
    above = matrix[header_row - 1] if header_row > 0 else ()
    out = []
    for c in range(max(len(head), len(above))):
        label = head[c] if c < len(head) and isinstance(head[c], str) else None
        in_header = _parse_dt(head[c]) if c < len(head) else None
        dt = in_header or (_parse_dt(above[c]) if c < len(above) else None)
        if dt is None:
            if not (label and _WEEK_NO_RE.search(label)):
                continue  # no date and not a 'Week N' column → not a session
            wm = _WEEK_RE.search(label)  # 'Week N (DD Mon)' dates itself given the year
            if wm and (month := _month_number(wm.group(3))) and year is not None:
                dt = datetime(year, month, int(wm.group(2)))
            elif week_anchor is not None:  # bare 'Week N' → anchor + 7*(N-1)
                dt = week_anchor + timedelta(weeks=int(_WEEK_NO_RE.search(label).group(1)) - 1)
            # else: a recognised 'Week N' column we can't date yet — keep it (dt None) for matches.
        # A date in the header cell is its own label; otherwise the header cell names it.
        name = label if (in_header is None and label and label.strip()) else dt.date().isoformat()
        out.append((c, str(name).strip(), dt))
    return out


# ---------------------------------------------------------------------------
# Parser base class
# ---------------------------------------------------------------------------


class Parser(abc.ABC):
    """A spreadsheet-layout parser.

    Each subclass recognises exactly one layout via :meth:`matches` and ingests it via
    :meth:`parse`. Parsers are stateless: the folder dispatcher in ``ingest`` matches each
    worksheet against the registered parsers and calls ``parse`` on the first that claims it.
    """

    #: short identifier recorded in the ingest report (e.g. 'roster').
    name: str

    @abc.abstractmethod
    def matches(self, ws) -> bool:
        """True if this parser recognises the worksheet's layout."""

    @abc.abstractmethod
    def parse(
        self,
        ws,
        db: AttendanceDb,
        term: str,
        year: int | None,
        ingest_id: int | None,
        week_anchor: datetime | None = None,
    ) -> None:
        """Ingest the worksheet into the attendance database."""


# ---------------------------------------------------------------------------
# Roster layout
# ---------------------------------------------------------------------------


class RosterParser(Parser):
    name = 'roster'

    def matches(self, ws) -> bool:
        """True if the sheet has a dancer_id header, dated session columns, and yes/no markers.

        The session cells must be attendance markers (yes/no/true/false/etc.), not free-text
        categories, to rule out sheets like '2026 L2 & SO Attendance' which have dancer_id
        columns but hold category strings like 'Level 2 & Social' instead of attendance bools.
        Requires that at least 50% of non-empty sampled cells are attendance markers.
        """
        matrix = list(ws.iter_rows(values_only=True))
        header = _roster_header(matrix)
        if header is None:
            return False
        sessions = _session_columns(matrix, header[0])
        if not sessions:
            return False
        # Sample cells and check if they're mostly attendance markers (not categories)
        header_row = header[0]
        markers, non_markers = 0, 0
        for col, _, _ in sessions[:3]:  # sample first few sessions
            for r in range(header_row + 1, min(header_row + 5, len(matrix))):  # sample first few rows
                cell = matrix[r][col] if col < len(matrix[r]) else None
                if cell is not None:
                    if _is_attendance_marker(cell):
                        markers += 1
                    else:
                        non_markers += 1
        # Require 50% or more of non-empty cells to be attendance markers
        total = markers + non_markers
        return total > 0 and markers >= total / 2

    def parse(
        self,
        ws,
        db: AttendanceDb,
        term: str,
        year: int | None,
        ingest_id: int | None,
        week_anchor: datetime | None = None,
    ) -> None:
        """Ingest one roster tab: an event, an activity per session column, attendance rows."""
        matrix = list(ws.iter_rows(values_only=True))
        header_row, dancer_col = _roster_header(matrix)
        sessions = _session_columns(matrix, header_row, year, week_anchor)
        ticket_by_did = self._ticket_by_dancer(matrix, header_row, dancer_col, _concession_col(matrix[header_row]))

        title = ws.title
        event_type = _event_type_for(title)
        window = _course_window((dt for _, _, dt in sessions), year)
        event_id = db.upsert_event(_event_name(term, title, year, event_type, window), event_type)

        for col, name, dt in sessions:
            if dt is None:
                continue  # a 'Week N' column we couldn't date (no year/anchor) — nothing to record
            activity_type = _activity_type_for(name, event_type)
            # Resolve the level from the column, then the sheet title, then the term/filename. The
            # term fallback matters for unification: a booking export reads its level from the
            # filename, so a roster whose title omits it (a dated 'Attendees' tab) must reach the
            # same level here, or the two sources name the session differently and never merge.
            difficulty = 'social' if activity_type == ActivityType.SOCIAL else _difficulty_for(name, title, term)
            activity_name = _activity_name(name, activity_type, difficulty, dt.date().isoformat(), event_type)
            activity_id = db.upsert_activity(event_id, activity_name, dt, activity_type, difficulty)
            attended_by: dict[str, list[bool]] = defaultdict(list)
            for r in range(header_row + 1, len(matrix)):
                row = matrix[r]
                did = row[dancer_col] if dancer_col < len(row) else None
                if isinstance(did, str) and did.startswith('DNC-'):
                    attended_by[did].append(_is_true(row[col]) if col < len(row) else False)

            # A wholly-unmarked session column was never registered — nobody filled it in — so a
            # blank cell there means 'turnout unknown', not a class-wide no-show. Only a column
            # someone actually marked is read as a present/absent register; otherwise the enrolled
            # dancers are recorded UNKNOWN (still registered, but attendance uncaptured).
            column_marked = any(any(a) for a in attended_by.values())

            cell = f'{title}!{get_column_letter(col + 1)}'
            anonymous_extra = 0
            for did, attendeds in attended_by.items():
                used = sum(attendeds)
                status = _present_status(used >= 1) if column_marked else AttendanceStatus.UNKNOWN
                db.record_attendance(
                    activity_id,
                    did,
                    status=status,
                    ticket_type=ticket_by_did.get(did),
                    ingest_id=ingest_id,
                    source_cell=cell,
                )
                anonymous_extra += max(used - 1, 0)  # un-renamed duplicate tickets → anonymous attendees
            if anonymous_extra:
                db.record_count(activity_id, None, anonymous_extra, ingest_id=ingest_id, source_cell=cell)

    @staticmethod
    def _ticket_by_dancer(matrix: list[tuple], header_row: int, dancer_col: int, conc_col: int | None) -> dict:
        """Map each dancer_id to a ticket type from its (per-dancer, constant) Concession flag.

        Empty if the roster has no Concession column. A dancer repeated across rows (un-renamed
        extra tickets) takes the first non-blank reading, so a stray blank duplicate row can't
        wipe a known value.
        """
        if conc_col is None:
            return {}
        out: dict[str, TicketType | None] = {}
        for r in range(header_row + 1, len(matrix)):
            row = matrix[r]
            did = row[dancer_col] if dancer_col < len(row) else None
            if not (isinstance(did, str) and did.startswith('DNC-')):
                continue
            ticket = _concession_ticket(row[conc_col]) if conc_col < len(row) else None
            out[did] = out.get(did) or ticket
        return out


# ---------------------------------------------------------------------------
# Level 2 tally layout
# ---------------------------------------------------------------------------


@dataclass
class _Tally:
    """Pre-scanned structure of a tally sheet."""

    matrix: list[tuple]
    groups: list[tuple] = field(default_factory=list)  # (col_1based, label)
    weeks: dict = field(default_factory=dict)  # start_row_1based -> (week_no, day, month_name)

    def cell(self, row1: int, col1: int):
        if 1 <= row1 <= len(self.matrix):
            row = self.matrix[row1 - 1]
            if 1 <= col1 <= len(row):
                return row[col1 - 1]
        return None


def _scan_tally(matrix: list[tuple]) -> _Tally:
    t = _Tally(matrix=matrix)
    for r, row in enumerate(matrix, start=1):
        for c, value in enumerate(row, start=1):
            if not isinstance(value, str):
                continue
            stripped = value.strip()
            if _is_group_header(stripped):
                t.groups.append((c, stripped))
            m = _WEEK_RE.search(stripped)
            if m:
                t.weeks[r] = (int(m.group(1)), int(m.group(2)), m.group(3))
    t.groups.sort()
    return t


def _tally_dates(t: _Tally, year: int | None):
    """The session date of every week label on a tally sheet, for deriving the event's window."""
    for week_no, day, month_name in t.weeks.values():
        month = _month_number(month_name)
        if month is not None and year is not None:
            yield datetime(year, month, day)


def _group_for_col(t: _Tally, col1: int) -> str | None:
    candidates = [(c, label) for c, label in t.groups if c <= col1]
    return max(candidates)[1] if candidates else None


def _ticket_for_totcell(t: _Tally, row1: int, col1: int) -> TicketType | None:
    for c in range(col1 - 1, 0, -1):
        value = t.cell(row1, c)
        if isinstance(value, str) and value.strip().lower() in _TICKET_LABELS:
            return _TICKET_LABELS[value.strip().lower()]
    return None


def _count_true(t: _Tally, start_col: int, start_row: int, end_col: int, end_row: int) -> int:
    return sum(
        1 for r in range(start_row, end_row + 1) for c in range(start_col, end_col + 1) if _is_true(t.cell(r, c))
    )


class Level2TallyParser(Parser):
    name = 'level2_tally'

    def matches(self, ws) -> bool:
        """True if the sheet has COUNTIF-of-TRUE totals, a group header, and a date axis.

        The date axis is usually a column of 'Week N (DD Mon)' labels (the weekly Level 2 tallies).
        A one-off tally (e.g. Teacher's Choice) has no weeks — its single session date sits in the
        tab name instead ('... 12th Dec') — so a day+month there satisfies the requirement too.
        """
        has_countif = has_group = has_week = False
        for row in ws.iter_rows(values_only=True):
            for value in row:
                if not isinstance(value, str):
                    continue
                if 'countif' in value.lower() and 'true' in value.lower():
                    has_countif = True
                if _is_group_header(value):
                    has_group = True
                if _WEEK_RE.search(value):
                    has_week = True
        has_single_date = _DAY_MONTH_RE.search(ws.title) is not None
        return has_countif and has_group and (has_week or has_single_date)

    def parse(
        self,
        ws,
        db: AttendanceDb,
        term: str,
        year: int | None,
        ingest_id: int | None,
        week_anchor: datetime | None = None,
    ) -> None:
        """Ingest one tally tab: a headcount per (week-or-single-date, activity group, ticket type)."""
        matrix = list(ws.iter_rows(values_only=True))
        t = _scan_tally(matrix)
        event_type = _event_type_for(ws.title)
        # A one-off tally carries no week labels; its single date comes from the tab name + the year.
        single_date = None if t.weeks else _date_from_day_month(ws.title, year)
        if single_date is not None:
            # Only one source describes this one-off, so name the event from its own tab rather than
            # the generic term, keeping it distinct from the term's weekly Level 1/2 events.
            event_id = db.upsert_event(f'{term}: {_strip_attendance(ws.title)}', event_type)
        else:
            window = _course_window(_tally_dates(t, year), year)
            event_id = db.upsert_event(_event_name(term, ws.title, year, event_type, window), event_type)

        for r, row in enumerate(matrix, start=1):
            for c, value in enumerate(row, start=1):
                if isinstance(value, str) and (m := _COUNTIF_RE.search(value)):
                    self._ingest_total(db, t, event_id, ws.title, year, ingest_id, r, c, m, single_date)

    def _ingest_total(self, db, t, event_id, title, year, ingest_id, row1, col1, m, single_date=None) -> None:  # noqa: PLR0913
        start_col, start_row = column_index_from_string(m.group(1)), int(m.group(2))
        end_col, end_row = column_index_from_string(m.group(3)), int(m.group(4))

        group = _group_for_col(t, col1)
        if group is None:
            return  # can't place this total without an activity group
        week = t.weeks.get(start_row)
        if week is not None:
            week_no, day, month_name = week
            month = _month_number(month_name)
            if month is None or year is None:
                return
            activity_date = datetime(year, month, day)
            raw_name = f'{group} (Week {week_no})'
        elif single_date is not None:
            activity_date = single_date
            raw_name = group
        else:
            return  # can't place this total without a date

        ticket_type = _ticket_for_totcell(t, row1, col1)
        head_count = _count_true(t, start_col, start_row, end_col, end_row)
        event_type = _event_type_for(title)
        activity_type = _activity_type_for(group, event_type)
        difficulty = 'social' if activity_type == ActivityType.SOCIAL else _difficulty_for(group)
        name = _activity_name(raw_name, activity_type, difficulty, activity_date.date().isoformat(), event_type)
        activity_id = db.upsert_activity(event_id, name, activity_date, activity_type, difficulty)
        cell = f'{title}!{get_column_letter(col1)}{row1}'
        db.record_count(activity_id, ticket_type, head_count, ingest_id=ingest_id, source_cell=cell)


# ---------------------------------------------------------------------------
# Level 2 count-grid layout (2024 'Levels 1-2' workbooks)
# ---------------------------------------------------------------------------


class Level2CountGridParser(Parser):
    """A plain grid: one row per week, with the per-ticket-type headcounts typed straight in.

    Header row carries 'Members' / 'Concessions' / 'Non-Members'; each week row has a
    'Week N (DD Mon)' label in the first column and integer counts under those headers.
    No checkbox tally, no COUNTIF — distinguishes it from the Level 2 tally layout.
    """

    name = 'level2_count_grid'

    def _ticket_columns(self, matrix: list[tuple]) -> tuple[int | None, dict]:
        """Find the header row of ticket labels; return (row_index, {col_index: TicketType})."""
        for r, row in enumerate(matrix):
            cols = {
                c: _TICKET_LABELS[v.strip().lower()]
                for c, v in enumerate(row)
                if isinstance(v, str) and v.strip().lower() in _TICKET_LABELS
            }
            if cols:
                return r, cols
        return None, {}

    def matches(self, ws) -> bool:
        """True if there are ticket-label headers and week labels, but no COUNTIF (that's the tally)."""
        matrix = list(ws.iter_rows(values_only=True))
        flat = [v for row in matrix for v in row if isinstance(v, str)]
        if any('countif' in v.lower() for v in flat):
            return False
        _, cols = self._ticket_columns(matrix)
        return bool(cols) and any(_WEEK_NO_RE.search(v) for v in flat)

    @staticmethod
    def _week_date(label: str, week_no: int, year: int | None, week_anchor: datetime | None) -> datetime | None:
        """Date a week row: from a '(DD Mon)' suffix when present, else week_anchor + 7*(week-1)."""
        m = _WEEK_RE.search(label)
        if m:
            month = _month_number(m.group(3))
            if month is not None and year is not None:
                return datetime(year, month, int(m.group(2)))
        if week_anchor is not None:
            return week_anchor + timedelta(weeks=week_no - 1)
        return None

    @classmethod
    def _row_date(cls, row: tuple, year: int | None, week_anchor: datetime | None) -> datetime | None:
        """The session date of a 'Week N' row, or None if the first cell isn't a week label."""
        label = row[0] if row else None
        m = _WEEK_NO_RE.search(label) if isinstance(label, str) else None
        return cls._week_date(label, int(m.group(1)), year, week_anchor) if m else None

    def parse(
        self,
        ws,
        db: AttendanceDb,
        term: str,
        year: int | None,
        ingest_id: int | None,
        week_anchor: datetime | None = None,
    ) -> None:
        """Ingest one count-grid tab: a headcount per (week, ticket type)."""
        matrix = list(ws.iter_rows(values_only=True))
        _, cols = self._ticket_columns(matrix)
        event_type = _event_type_for(ws.title)
        window = _course_window((self._row_date(row, year, week_anchor) for row in matrix), year)
        event_id = db.upsert_event(_event_name(term, ws.title, year, event_type, window), event_type)

        for r, row in enumerate(matrix):
            label = row[0] if row else None
            m = _WEEK_NO_RE.search(label) if isinstance(label, str) else None
            if not m:
                continue
            week_no = int(m.group(1))
            activity_date = self._week_date(label, week_no, year, week_anchor)
            if activity_date is None:
                continue
            raw_name = f'Level 2 Classes (Week {week_no})'
            activity_type = _activity_type_for(raw_name, event_type)
            difficulty = _difficulty_for(raw_name)
            name = _activity_name(raw_name, activity_type, difficulty, activity_date.date().isoformat(), event_type)
            activity_id = db.upsert_activity(event_id, name, activity_date, activity_type, difficulty)
            for c, ticket_type in cols.items():  # c is a 0-based column index
                value = row[c] if c < len(row) else None
                if isinstance(value, bool) or not isinstance(value, (int, float)):
                    continue  # skip SUM/AVERAGE formula cells and blanks (bool is an int subclass)
                cell = f'{ws.title}!{get_column_letter(c + 1)}{r + 1}'
                db.record_count(activity_id, ticket_type, int(value), ingest_id=ingest_id, source_cell=cell)


class L2SOAttendanceParser(Parser):
    """Parse 2026-style 'L2 & SO Attendance' sheets.

    One row per dancer; dated session columns (the date sits in the header row). Each
    session cell holds a *ticket category*, not a yes/no marker:

    * ``Level 2 & Social`` — at that night's Level 2 lesson (they also stayed for the
      social, but per ESDS we record them as a Level 2 attendee only, not in the social).
    * ``Social-Only`` — at the social only.
    * ``Absent`` — signed up for Level 2 that night but didn't show: a Level 2 roster row
      with ``status = 'absent'``.
    * blank — no record.

    Each night therefore yields up to two activities, a Level 2 lesson and a social, and a
    dancer lands in exactly one of them per session. Activities are created lazily so a
    night with no record of a kind produces no empty activity. Only ``DNC-`` rows are
    read, so the COUNTIF/COUNTA summary formulas below the table are never mistaken for
    categories.
    """

    name = 'l2_so_attendance'

    # Category cell value (lower-cased) -> (activity it counts towards, attendance status).
    _CATEGORIES = {
        'level 2 & social': (ActivityType.LESSON, AttendanceStatus.ATTENDED),
        'social-only': (ActivityType.SOCIAL, AttendanceStatus.ATTENDED),
        'social only': (ActivityType.SOCIAL, AttendanceStatus.ATTENDED),
        'absent': (ActivityType.LESSON, AttendanceStatus.ABSENT),  # signed up for Level 2, didn't show
    }

    def matches(self, ws) -> bool:
        """True if the sheet matches the L2 & SO layout: title contains L2 and SO, dancer_id column, sessions."""
        title_lower = ws.title.lower()
        if 'l2' not in title_lower or 'so' not in title_lower:
            return False
        matrix = list(ws.iter_rows(values_only=True))
        header = _roster_header(matrix)
        if header is None:
            return False
        sessions = _session_columns(matrix, header[0])
        return bool(sessions)

    def parse(
        self,
        ws,
        db: AttendanceDb,
        term: str,
        year: int | None,
        ingest_id: int | None,
        week_anchor: datetime | None = None,
    ) -> None:
        """Ingest one L2 & SO tab: a Level 2 lesson and a social per night, by ticket category."""
        matrix = list(ws.iter_rows(values_only=True))
        header_row, dancer_col = _roster_header(matrix)
        sessions = _session_columns(matrix, header_row, year)

        title = ws.title
        event_type = _event_type_for(title)
        window = _course_window((dt for _, _, dt in sessions), year)
        event_id = db.upsert_event(_event_name(term, title, year, event_type, window), event_type)

        activities: dict[tuple, int] = {}  # (date_iso, ActivityType) -> activity_id

        def activity_for(dt: datetime, kind: ActivityType) -> int:
            key = (dt.date().isoformat(), kind)
            if key not in activities:
                difficulty = 'Level 2' if kind == ActivityType.LESSON else 'social'
                name = _activity_name(key[0], kind, difficulty, key[0], event_type)
                activities[key] = db.upsert_activity(event_id, name, dt, kind, difficulty)
            return activities[key]

        for r in range(header_row + 1, len(matrix)):
            row = matrix[r]
            did = row[dancer_col] if dancer_col < len(row) else None
            if not (isinstance(did, str) and did.startswith('DNC-')):
                continue
            for col, _, dt in sessions:
                if dt is None:
                    continue
                value = row[col] if col < len(row) else None
                entry = self._CATEGORIES.get(value.strip().lower()) if isinstance(value, str) else None
                if entry is None:
                    continue
                kind, status = entry
                cell = f'{title}!{get_column_letter(col + 1)}{r + 1}'
                db.record_attendance(activity_for(dt, kind), did, status=status, ingest_id=ingest_id, source_cell=cell)


# ---------------------------------------------------------------------------
# One-off social register layout (Tea Dances etc.)
# ---------------------------------------------------------------------------


_CONCESSION_YES = {'yes', 'y', 'true', '1'}
_CONCESSION_NO = {'no', 'n', 'false', '0'}


def _concession_ticket(value) -> TicketType | None:
    """Map a boolean-ish 'Concession' cell to a ticket type. Blank means no information (None).

    The column records eligibility for the member/concession rate, not membership as such:
    'Yes' means the attendee is a member or concession (not ordinary) without saying which —
    hence MEMBER_OR_CONCESSION — and 'No' means they paid the ordinary rate. A non-empty value
    that is neither yes nor no is UNKNOWN: the source said *something* about the rate that we
    can't read, which is distinct from a blank cell (no information at all, left as None/NULL).
    """
    if isinstance(value, bool):
        return TicketType.MEMBER_OR_CONCESSION if value else TicketType.ORDINARY
    if not isinstance(value, str) or not (v := value.strip().lower()):
        return None
    if v in _CONCESSION_YES:
        return TicketType.MEMBER_OR_CONCESSION
    if v in _CONCESSION_NO:
        return TicketType.ORDINARY
    return TicketType.UNKNOWN


# Header of the per-attendee attendance marker on a one-off register. Different years use
# different words ('Present?' on the Tea Dances, 'Attended' on the Sunday Socials). Matched
# exactly so 'Weeks attended'/'Times attended' summary columns are never mistaken for it.
_ATTENDANCE_HEADERS = {'present', 'present?', 'attended', 'attended?'}


def _is_attendance_header(label) -> bool:
    """True if a header cell labels the per-attendee attendance marker on a social register."""
    return isinstance(label, str) and label.strip().lower() in _ATTENDANCE_HEADERS


def _register_blocks(header: tuple) -> list[tuple]:
    """Find each (dancer_id, concession_col_or_None, marker_col) block across a register header.

    These registers print the attendee list as one or more halves side by side, each a repeat
    of e.g. '# | dancer_id | redacted | Concession | Present?'. Each dancer_id column owns the
    Concession and attendance-marker columns to its right, up to the next dancer_id column.
    """
    did_cols = [j for j, c in enumerate(header) if isinstance(c, str) and c.strip() == 'dancer_id']
    blocks = []
    for k, dc in enumerate(did_cols):
        end = did_cols[k + 1] if k + 1 < len(did_cols) else len(header)
        marker = concession = None
        for j in range(dc + 1, end):
            label = header[j].strip().lower() if isinstance(header[j], str) else ''
            if marker is None and _is_attendance_header(header[j]):
                marker = j
            if concession is None and label == 'concession':
                concession = j
        if marker is not None:
            blocks.append((dc, concession, marker))
    return blocks


class SocialRegisterParser(Parser):
    """Parse a one-off social's named register (Tea Dances, Sunday Socials, parties).

    One dated event, one social activity. Each row holds one attendee per side-by-side block:
    a ``dancer_id``, optionally a ``Concession`` Yes/No (ticket type), and an attendance marker
    headed ``Present?`` or ``Attended`` (☑️/x attended, ❎/blank not). Distinguished from the
    weekly roster by having that marker column and *no* dated session columns. The event date
    comes from the sheet tab name or the filename, since the cells carry none. A dancer_id
    repeated across rows (un-renamed extra tickets) becomes one named attendee plus anonymous
    extras, as in the roster layout.

    The marker column may be entirely blank — a booked-but-not-yet-run party (the 'End of Term
    Party' allocation lists) where turnout was never recorded. Then every booker is UNKNOWN
    (a held place, turnout uncaptured) rather than ABSENT, and there are no anonymous extras.
    """

    name = 'social_register'

    def matches(self, ws) -> bool:
        """True for a dancer_id register that has a 'Present?'/'Attended' marker column and no dates.

        The marker column must really be one: a booking list can carry a 'present' column that is
        mostly empty with the odd free-text note ('no show', 'paid £10 cash on door'), which must
        not be read as a register — so when the column is marked at all, the marks must mostly be
        real yes/no marks. A wholly blank marker column is allowed (a booked-but-unrun party).
        Dated session columns disqualify it — that is the weekly roster's job.
        """
        matrix = list(ws.iter_rows(values_only=True))
        header = _roster_header(matrix)
        if header is None:
            return False
        header_row = header[0]
        if _session_columns(matrix, header_row):
            return False
        blocks = _register_blocks(matrix[header_row])
        if not blocks:
            return False  # no Present?/Attended marker column → not a register
        filled, recognised = self._marker_fill(matrix, header_row, blocks)
        return filled == 0 or recognised >= filled / 2

    @staticmethod
    def _marker_fill(matrix: list[tuple], header_row: int, blocks: list[tuple]) -> tuple[int, int]:
        """Count (non-empty, recognised-as-marker) marker cells against DNC- rows across blocks."""
        filled = recognised = 0
        for row in matrix[header_row + 1 :]:
            for did_col, _, marker_col in blocks:
                did = row[did_col] if did_col < len(row) else None
                if not (isinstance(did, str) and did.startswith('DNC-')):
                    continue
                cell = row[marker_col] if marker_col < len(row) else None
                if cell is None or (isinstance(cell, str) and not cell.strip()):
                    continue
                filled += 1
                recognised += _is_attendance_marker(cell)
        return filled, recognised

    def parse(
        self,
        ws,
        db: AttendanceDb,
        term: str,
        year: int | None,
        ingest_id: int | None,
        week_anchor: datetime | None = None,
    ) -> None:
        """Ingest one social register: a single dated social activity with named attendance."""
        matrix = list(ws.iter_rows(values_only=True))
        header_row, _ = _roster_header(matrix)
        blocks = _register_blocks(matrix[header_row])
        # The event date is in the tab name (e.g. '26 March 23') or the filename, not the cells.
        event_date = _date_from_title(ws.title) or _date_from_title(term)
        if event_date is None or not blocks:
            return  # without an event date or any attendee block there is nothing to anchor

        event_id = db.upsert_event(_strip_attendance(term), EventType.SOCIAL)
        activity_id = db.upsert_activity(event_id, 'Social', event_date, ActivityType.SOCIAL, 'social')

        # If no marker cell anywhere is a real mark, the register records bookings whose turnout was
        # never captured: every booker is UNKNOWN, and a blank does not mean absent. Mirrors the
        # roster's column-marked rule. A marked register keeps present/absent + anonymous extras.
        _, recognised = self._marker_fill(matrix, header_row, blocks)
        register_marked = recognised > 0

        attendees = self._collect(matrix, header_row, blocks, ws.title)
        anonymous_extra = 0
        for did, (present_count, ticket_type, cell) in attendees.items():
            status = _present_status(present_count >= 1) if register_marked else AttendanceStatus.UNKNOWN
            db.record_attendance(
                activity_id,
                did,
                status=status,
                ticket_type=ticket_type,
                ingest_id=ingest_id,
                source_cell=cell,
            )
            if register_marked:
                anonymous_extra += max(present_count - 1, 0)  # un-renamed duplicate tickets → anonymous attendees
        if anonymous_extra:
            db.record_count(activity_id, None, anonymous_extra, ingest_id=ingest_id, source_cell=ws.title)

    @staticmethod
    def _collect(matrix: list[tuple], header_row: int, blocks: list[tuple], title: str) -> dict:
        """Aggregate per dancer_id across every block and row: (present_count, ticket_type, cell)."""
        out: dict[str, list] = {}
        for r in range(header_row + 1, len(matrix)):
            row = matrix[r]
            for did_col, conc_col, pres_col in blocks:
                did = row[did_col] if did_col < len(row) else None
                if not (isinstance(did, str) and did.startswith('DNC-')):
                    continue
                attended = _is_true(row[pres_col]) if pres_col < len(row) else False
                ticket = _concession_ticket(row[conc_col]) if conc_col is not None and conc_col < len(row) else None
                acc = out.setdefault(did, [0, None, f'{title}!{get_column_letter(did_col + 1)}{r + 1}'])
                acc[0] += int(attended)
                acc[1] = acc[1] or ticket
        return {did: tuple(acc) for did, acc in out.items()}


class BookingExportParser(Parser):
    """Parse a dancecloud booking export — who *bought a ticket*, not who turned up.

    These exports are the only record for several early terms (2022 / early-2023 Level 1) and
    one-off parties, and they carry no attendance: a dancer on the list — and again per session
    on the 'Attendees By Activity' view — only means they held a ticket. So every booking is
    recorded as UNKNOWN attendance (interest, never a counted head) unless a signal says otherwise.
    Two shapes are handled:

    * **dated** ('Attendees By Activity'): a ``Date`` column gives one row per dancer per session.
      Each distinct date becomes an activity; every booking is UNKNOWN (a Cancelled one is ABSENT).
    * **single-event** (the 2023 Christmas Party): no Date column, but a ``present`` column of
      ad-hoc notes ('no show', 'refunded', 'paid cash on door'). The date comes from the title;
      the note refines a row to ABSENT/ATTENDED, otherwise it stays UNKNOWN.

    The plain 'Attendees' summary (a booking Status but no Date and no present column) is not
    claimed: it is the same bookings as 'Attendees By Activity', minus the dates.
    """

    name = 'booking_export'

    @staticmethod
    def _has_dates(matrix: list[tuple], header_row: int, date_col: int) -> bool:
        return any(_parse_dt(row[date_col]) is not None for row in matrix[header_row + 1 :] if date_col < len(row))

    def matches(self, ws) -> bool:
        """True for a dancer_id + booking Status list that is either dated or has a present column.

        A 'Checked In' column means it is the *modern* dancecloud export, which
        :class:`DancecloudActivityParser` (registered ahead of this one) handles richly, so this
        parser stands aside from it.
        """
        matrix = list(ws.iter_rows(values_only=True))
        header = _roster_header(matrix)
        if header is None:
            return False
        cols = _header_columns(matrix[header[0]])
        if 'dancer_id' not in cols or 'status' not in cols or 'checked in' in cols:
            return False
        if 'date' in cols and self._has_dates(matrix, header[0], cols['date']):
            return True
        return 'present' in cols

    def parse(
        self,
        ws,
        db: AttendanceDb,
        term: str,
        year: int | None,
        ingest_id: int | None,
        week_anchor: datetime | None = None,
    ) -> None:
        """Ingest one booking export: an activity per session, every row UNKNOWN unless noted."""
        matrix = list(ws.iter_rows(values_only=True))
        header_row, _ = _roster_header(matrix)
        cols = _header_columns(matrix[header_row])

        event_type = _event_type_for(term)
        fallback_dt = _date_from_title(ws.title) or _date_from_title(term)
        best = self._best_status_per_pair(matrix, header_row, cols, fallback_dt, ws.title)
        if event_type == EventType.COURSE:
            window = _course_window((_parse_dt(date_iso) for date_iso, _ in best), year)
            event_name = _course_event_name(term, year, _difficulty_for(term), window)
        else:
            event_name = _booking_event_name(term, year)
        event_id = db.upsert_event(event_name, event_type)

        activity_type = _activity_type_for(term, event_type)
        difficulty = 'social' if activity_type == ActivityType.SOCIAL else _difficulty_for(term)
        raw_name = 'Social' if activity_type == ActivityType.SOCIAL else 'Lesson'
        activities: dict[str, int] = {}
        for (date_iso, did), (status, cell) in best.items():
            if date_iso not in activities:
                name = _activity_name(raw_name, activity_type, difficulty, date_iso, event_type)
                activities[date_iso] = db.upsert_activity(event_id, name, date_iso, activity_type, difficulty)
            db.record_attendance(activities[date_iso], did, status=status, ingest_id=ingest_id, source_cell=cell)

    @staticmethod
    def _best_status_per_pair(matrix: list[tuple], header_row: int, cols: dict, fallback_dt, title: str) -> dict:
        """One status per (date, dancer), keeping the most informative reading (attended>absent>unknown)."""
        did_col, status_col = cols['dancer_id'], cols['status']
        date_col, present_col = cols.get('date'), cols.get('present')
        rank = {AttendanceStatus.UNKNOWN: 0, AttendanceStatus.ABSENT: 1, AttendanceStatus.ATTENDED: 2}
        best: dict[tuple, tuple] = {}
        for r in range(header_row + 1, len(matrix)):
            row = matrix[r]
            did = row[did_col] if did_col < len(row) else None
            if not (isinstance(did, str) and did.startswith('DNC-')):
                continue
            dt = _parse_dt(row[date_col]) if date_col is not None and date_col < len(row) else None
            dt = dt or fallback_dt
            if dt is None:
                continue
            status = _booking_status(
                row[status_col] if status_col < len(row) else None,
                row[present_col] if present_col is not None and present_col < len(row) else None,
            )
            key = (dt.date().isoformat(), did)
            if key not in best or rank[status] > rank[best[key][0]]:
                best[key] = (status, f'{title}!{get_column_letter(did_col + 1)}{r + 1}')
        return best


# A trailing date phrase on a modern dancecloud export's filename, in either order and with an
# optional day range: 'June 7th 2026', 'March 20th-22nd 2026', '14 Dec 2023'. Stripped so the
# event is named for what it is, not for when one instance of it ran.
_TITLE_DATE_TAIL_RE = re.compile(
    r'\s+(?:'
    r'\d{1,2}(?:st|nd|rd|th)?\s+[A-Za-z]+'  # 14 Dec
    r'|[A-Za-z]+\s+\d{1,2}(?:st|nd|rd|th)?(?:\s*[-–]\s*\d{1,2}(?:st|nd|rd|th)?)?'  # June 7th / March 20th-22nd
    r')\s+\d{2,4}\s*$',
    re.IGNORECASE,
)


def _dancecloud_event_name(term: str, year: int | None) -> str:
    """Event name for a modern dancecloud export: the event title, date tail stripped, year suffixed.

    The 'Attendees By Activity' tab title is the generic export-view name, useless as an event
    name, so the event is named from the filename instead. The trailing date is dropped so a
    recurring fixture (e.g. several "Teachers' Workshop" dates) collapses to one event carrying a
    dated activity per instance; the year keeps same-named events in different years apart.
    """
    base = _TITLE_DATE_TAIL_RE.sub('', _strip_attendance(term)).strip()
    return f'{base} ({year})' if year else base


class DancecloudActivityParser(Parser):
    """Parse a modern dancecloud 'Attendees By Activity' export (2026-on).

    One row per (dancer, activity) booking, carrying the activity's name and datetime in
    ``Activity``/``Date`` and a per-row ``Checked In`` timestamp — dancecloud's attendance signal:
    a timestamp means the dancer was scanned in on the door (ATTENDED), a blank means a confirmed
    booking that never checked in (ABSENT). A 'Cancelled' booking ``Status`` is ABSENT regardless.
    Because dancecloud already expands a multi-activity ticket (a weekend 'Full pass', a combined
    'Workshop and Tea Dance' ticket) into one row per activity, each activity is ingested in its
    own right — splitting e.g. a workshop and its tea dance into two activities, each classified
    from its own name — so the awkward 'one ticket, several activities' case needs no special
    handling here.

    Distinguished from the older booking exports (handled by :class:`BookingExportParser`) purely
    by the ``Checked In`` column, which only the modern export carries; that is also why this
    parser is registered ahead of it. The sibling 'Attendees' rollup and 'Check-Ins' pivot in the
    same workbook are the same data and are skipped as redundant by the dispatcher.

    If no row anywhere on the sheet carries a check-in (an event where the door scanner was never
    used) the column conveys nothing, so every booking is recorded UNKNOWN rather than as a
    sheet-wide no-show — mirroring the roster's column-marked rule.

    A check-in to *any* day of a class implies attendance at every day it ran: door scanning on the
    later day of a weekender is patchy, so a dancer scanned into a multi-day class (the same Activity
    name on more than one date, e.g. a weekender's 'Track A') on one day is taken to have attended it
    on the others too. This lifts only lessons (a social is per-night) and only an otherwise-ABSENT
    reading; it mirrors the all-or-none assumption :class:`StockbridgeSwingoutParser` makes from its
    single Registered flag.
    """

    name = 'dancecloud_activity'

    def matches(self, ws) -> bool:
        """True for a dancer_id table carrying Activity, Date and the modern Checked In columns."""
        matrix = list(ws.iter_rows(values_only=True))
        header = _roster_header(matrix)
        if header is None:
            return False
        cols = _header_columns(matrix[header[0]])
        return all(k in cols for k in ('dancer_id', 'activity', 'date', 'checked in'))

    @staticmethod
    def _sessions(matrix: list[tuple], header_row: int, cols: tuple, title: str) -> dict:
        """Aggregate each (activity label, date) -> {dancer: [best status, attended-ticket count, cell]}.

        The best status is the most informative across that dancer's tickets for the session; the
        attended-ticket surplus becomes anonymous extra heads. If no row anywhere carries a check-in
        (the door scanner never ran) a blank check-in means 'uncaptured' (UNKNOWN), not a no-show.
        """
        did_col, act_col, date_col, checkin_col, status_col = cols
        sheet_has_checkins = any(
            _is_checked_in(row[checkin_col]) for row in matrix[header_row + 1 :] if checkin_col < len(row)
        )
        rank = {AttendanceStatus.UNKNOWN: 0, AttendanceStatus.ABSENT: 1, AttendanceStatus.ATTENDED: 2}
        sessions: dict[tuple, dict] = {}
        for r in range(header_row + 1, len(matrix)):
            row = matrix[r]
            did = row[did_col] if did_col < len(row) else None
            if not (isinstance(did, str) and did.startswith('DNC-')):
                continue
            dt = _parse_dt(row[date_col]) if date_col < len(row) else None
            label = row[act_col] if act_col < len(row) else None
            if dt is None or not (isinstance(label, str) and label.strip()):
                continue
            status = _checkin_status(
                row[checkin_col] if checkin_col < len(row) else None,
                row[status_col] if status_col is not None and status_col < len(row) else None,
                sheet_has_checkins,
            )
            acc = sessions.setdefault((label.strip(), dt), {}).setdefault(did, [AttendanceStatus.UNKNOWN, 0, None])
            if rank[status] >= rank[acc[0]]:
                acc[0], acc[2] = status, f'{title}!{get_column_letter(did_col + 1)}{r + 1}'
            if status == AttendanceStatus.ATTENDED:
                acc[1] += 1
        return sessions

    @staticmethod
    def _attended_lessons(sessions: dict, event_type: EventType) -> dict:
        """Per dancer, the set of lesson labels they were scanned into on some day (for propagation)."""
        attended_lessons: dict[str, set] = defaultdict(set)
        for (label, _), per in sessions.items():
            if _activity_type_from_name(label, event_type) != ActivityType.LESSON:
                continue
            for did, (status, _, _) in per.items():
                if status == AttendanceStatus.ATTENDED:
                    attended_lessons[did].add(label)
        return attended_lessons

    def parse(
        self,
        ws,
        db: AttendanceDb,
        term: str,
        year: int | None,
        ingest_id: int | None,
        week_anchor: datetime | None = None,
    ) -> None:
        """Ingest one modern export: an activity per (label, date), attendance from the check-ins."""
        matrix = list(ws.iter_rows(values_only=True))
        header_row, _ = _roster_header(matrix)
        cols = _header_columns(matrix[header_row])
        did_col, act_col, date_col = cols['dancer_id'], cols['activity'], cols['date']
        checkin_col, status_col = cols['checked in'], cols.get('status')

        event_type = _event_type_for(term)
        event_id = db.upsert_event(_dancecloud_event_name(term, year), event_type)

        sessions = self._sessions(matrix, header_row, (did_col, act_col, date_col, checkin_col, status_col), ws.title)
        attended_lessons = self._attended_lessons(sessions, event_type)

        for (label, dt), per in sessions.items():
            activity_type = _activity_type_from_name(label, event_type)
            difficulty = 'social' if activity_type == ActivityType.SOCIAL else _difficulty_for(label, term)
            activity_id = db.upsert_activity(event_id, label, dt, activity_type, difficulty)
            is_lesson = activity_type == ActivityType.LESSON
            anonymous_extra = 0
            for did, (status, attended, cell) in per.items():
                if is_lesson and status == AttendanceStatus.ABSENT and label in attended_lessons[did]:
                    status = AttendanceStatus.ATTENDED  # scanned into this class on another day
                db.record_attendance(activity_id, did, status=status, ingest_id=ingest_id, source_cell=cell)
                anonymous_extra += max(attended - 1, 0)  # un-renamed duplicate tickets → anonymous heads
            if anonymous_extra:
                db.record_count(activity_id, None, anonymous_extra, ingest_id=ingest_id, source_cell=ws.title)


# ---------------------------------------------------------------------------
# Stockbridge Swingout weekend (one-off, event-specific)
# ---------------------------------------------------------------------------


class StockbridgeSwingoutParser(Parser):
    """Parse the Stockbridge Swingout weekend register — the one frankly event-specific parser.

    There is exactly one Stockbridge Swingout in the tree (Sept-Oct 2025), and its ``Ticket
    Type`` strings are the only per-activity signal, so the ticket -> activity mapping is
    hand-encoded here rather than derived from the layout. Every other parser keys on a layout;
    this one keys on an event, justified because the ticket vocabulary is unique to it.

    Layout: one row per booking with a ``Ticket Type`` string and a boolean ``Registered`` — the
    door attendance marker (True = turned up, False = no-show). A multi-activity ticket grants a
    fixed set of weekend activities (all-or-none), so each booking is expanded into one attendance
    row per granted activity carrying its Registered status. The dated ``Registered`` header sits
    under the Saturday date; the Friday social is the day before.

    The levelled classes (Improvers/Intermediate, Intermediate/Advanced) ran as one track across
    Saturday *and* Sunday, so a Full Pass holder's classes are recorded as two activities, one per
    day (a Saturday Only ticket grants only the Saturday class). This matches how the 30th's
    weekender classes are split per day; and because the single ``Registered`` flag stands in for a
    check-in, a registered Full Pass holder is recorded ATTENDED on both days — the same assumption
    :class:`DancecloudActivityParser` makes for the 30th (a check-in to any day of a class implies
    attendance at every day of it). Counts are ticket-grant counts under the all-or-none assumption,
    deliberately higher than the door headcounts noted in free text on the sheet, which are not
    ingested.
    """

    name = 'stockbridge_swingout'

    def matches(self, ws) -> bool:
        """True for a Stockbridge Swingout tab: 'swingout' in the title with Ticket Type + Registered."""
        if 'swingout' not in ws.title.lower() and 'swing out' not in ws.title.lower():
            return False
        matrix = list(ws.iter_rows(values_only=True))
        header = _roster_header(matrix)
        if header is None:
            return False
        cols = _header_columns(matrix[header[0]])
        return 'ticket type' in cols and 'registered' in cols

    @staticmethod
    def _grants(ticket: str, friday: datetime, saturday: datetime, sunday: datetime) -> list[tuple]:
        """Expand one Ticket Type string into the (name, type, difficulty, date) activities it grants.

        A 'Full Pass' grants the Friday social, its levelled track classes on *both* Saturday and
        Sunday and the Saturday social; a 'Saturday Only' the Saturday class and the Saturday social;
        the two 'Social Only' tickets their one named social. The level after the ' - ' becomes the
        class track's difficulty, and the two days of one track share a name (distinguished by date),
        exactly as the 30th's per-day track activities do.
        """
        base, _, level = ticket.partition(' - ')
        base, level = base.strip().lower(), level.strip()
        friday_social = ('Friday Social', ActivityType.SOCIAL, 'social', friday)
        saturday_social = ('Saturday Social', ActivityType.SOCIAL, 'social', saturday)
        sat_classes = (f'{level} Classes', ActivityType.LESSON, level or None, saturday)
        sun_classes = (f'{level} Classes', ActivityType.LESSON, level or None, sunday)
        if base == 'full pass':
            return [friday_social, sat_classes, sun_classes, saturday_social]
        if base == 'saturday only':
            return [sat_classes, saturday_social]
        if base.startswith('friday social'):
            return [friday_social]
        if base.startswith('saturday social'):
            return [saturday_social]
        return []

    @staticmethod
    def _saturday(matrix: list[tuple], header_row: int, reg_col: int) -> datetime | None:
        """The Saturday date written above the Registered column, else any date cell, else None."""
        if header_row > 0:
            above = matrix[header_row - 1]
            if reg_col < len(above) and (dt := _parse_dt(above[reg_col])):
                return dt
        for row in matrix:
            for value in row:
                if dt := _parse_dt(value):
                    return dt
        return None

    def parse(
        self,
        ws,
        db: AttendanceDb,
        term: str,
        year: int | None,
        ingest_id: int | None,
        week_anchor: datetime | None = None,
    ) -> None:
        """Ingest the Stockbridge weekend: expand each ticket into the activities it granted."""
        matrix = list(ws.iter_rows(values_only=True))
        header_row, dancer_col = _roster_header(matrix)
        cols = _header_columns(matrix[header_row])
        ticket_col, reg_col = cols['ticket type'], cols['registered']

        saturday = self._saturday(matrix, header_row, reg_col)
        if saturday is None:
            return  # no date to anchor the weekend on
        friday, sunday = saturday - timedelta(days=1), saturday + timedelta(days=1)

        title = ws.title
        event_type = _event_type_for(title)  # 'swingout' -> WEEKENDER
        event_id = db.upsert_event(_event_name(term, title, year, event_type), event_type)

        activities: dict[tuple, int] = {}

        def activity_for(name: str, kind: ActivityType, difficulty: str | None, dt: datetime) -> int:
            key = (name, dt.date())  # a track shares a name across its two days; the date keeps them apart
            if key not in activities:
                activities[key] = db.upsert_activity(event_id, name, dt, kind, difficulty)
            return activities[key]

        cols_ix = (dancer_col, ticket_col, reg_col)
        per = self._collect(matrix, header_row, cols_ix, (friday, saturday, sunday), title, activity_for)
        for aid, by_did in per.items():
            anonymous_extra = 0
            for did, (status, attended, cell) in by_did.items():
                db.record_attendance(aid, did, status=status, ingest_id=ingest_id, source_cell=cell)
                anonymous_extra += max(attended - 1, 0)  # un-renamed duplicate tickets -> anonymous heads
            if anonymous_extra:
                db.record_count(aid, None, anonymous_extra, ingest_id=ingest_id, source_cell=title)

    @classmethod
    def _collect(cls, matrix, header_row, cols_ix, days, title, activity_for) -> dict:
        """Aggregate, per (activity, dancer): best Registered status, attended-ticket count, a cell.

        The attended-ticket surplus per (activity, dancer) becomes anonymous extra heads, as in the
        roster layout. ``cols_ix`` is (dancer_col, ticket_col, reg_col); ``days`` is the
        (friday, saturday, sunday) anchor passed to :meth:`_grants`.
        """
        dancer_col, ticket_col, reg_col = cols_ix
        friday, saturday, sunday = days
        rank = {AttendanceStatus.UNKNOWN: 0, AttendanceStatus.ABSENT: 1, AttendanceStatus.ATTENDED: 2}
        per: dict[int, dict[str, list]] = defaultdict(dict)
        for r in range(header_row + 1, len(matrix)):
            row = matrix[r]
            did = row[dancer_col] if dancer_col < len(row) else None
            ticket = row[ticket_col] if ticket_col < len(row) else None
            if not (isinstance(did, str) and did.startswith('DNC-')) or not isinstance(ticket, str):
                continue
            status = _present_status(_is_true(row[reg_col]) if reg_col < len(row) else False)
            cell = f'{title}!{get_column_letter(dancer_col + 1)}{r + 1}'
            for name, kind, difficulty, dt in cls._grants(ticket, friday, saturday, sunday):
                acc = per[activity_for(name, kind, difficulty, dt)].setdefault(did, [AttendanceStatus.UNKNOWN, 0, None])
                if rank[status] >= rank[acc[0]]:
                    acc[0], acc[2] = status, cell
                if status == AttendanceStatus.ATTENDED:
                    acc[1] += 1
        return per


# ---------------------------------------------------------------------------
# Canonical event / activity naming
#
# One term's course is described by several sheets (a weekly roster, a Level 2 tally, the
# dancecloud booking export) under different names. Building every event/activity name from
# the same parts collapses those sources onto one ``(event, name, date)`` key.
# ---------------------------------------------------------------------------


def _strip_attendance(s: str) -> str:
    """Tidy a tab title for use as an event name.

    Drops 'Attendance' (and its 31-char-truncated 'Attendanc') and 'Attendees', which belong
    on a spreadsheet tab but are noise in an event name, and normalises 'Level 2/3' (and '2-3',
    '2 & 3') down to plain 'Level 2' since ESDS ran the two levels as one session.
    """
    s = _LEVEL_2_3_RE.sub('Level 2', s)
    return re.sub(r'\s+', ' ', re.sub(r'\bAttend(?:anc|ee)\w*', '', s, flags=re.IGNORECASE)).strip()


def _booking_event_name(term: str, year: int | None) -> str:
    """A stable event name for a booking export: the course/party with its export tail stripped.

    A year suffix keeps same-named courses in different years apart (Term A 2022 vs 2023) while
    letting several re-exports of one term collapse into one event.
    """
    base = _EXPORT_SUFFIX_RE.sub('', _strip_attendance(term)).strip()
    return f'{base} ({year})' if year else base


# Month names (and the common 'Sept' variant) used to tell whether a term name already carries
# its own time qualifier. The newer termly workbooks do ('May-Jun 2025 Attendance'); the older
# booking exports don't ('Level 1 Fundamentals Term B').
_MONTH_ABBR = ('', 'Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun', 'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec')
_MONTH_NAME_RE = re.compile(r'\b(?:jan|feb|mar|apr|may|jun|jul|aug|sep|sept|oct|nov|dec)', re.IGNORECASE)


def _has_month_name(text: str) -> bool:
    """True if a name already carries a month qualifier, so a derived window would just repeat it."""
    return _MONTH_NAME_RE.search(text) is not None


def _course_window(dates, year: int | None) -> str | None:
    """A 'Mmm YYYY' / 'Mmm-Mmm YYYY' label spanning the months of a course's session dates.

    This is the axis that separates ESDS's reused 'Term A/B' labels within one year: the same
    name recurs three times across 2022 (spring / summer / autumn), but each occupies its own,
    non-overlapping run of session dates. Built from the dates inside the sheets — not the
    export timestamp in the filename, which the export-suffix strip deliberately discards so
    re-exports of one term still collapse — so every source of one term derives the same window
    and converges. None when no dates are known (the caller falls back to the bare year). A term
    that straddles the new year would mis-order its months, but ESDS terms don't, and the
    workbook year is resolved as the dominant one regardless.
    """
    months = sorted({d.month for d in dates if d is not None})
    if not months:
        return None
    span = _MONTH_ABBR[months[0]] if months[0] == months[-1] else f'{_MONTH_ABBR[months[0]]}-{_MONTH_ABBR[months[-1]]}'
    return f'{span} {year}' if year else span


def _course_event_name(term: str, year: int | None, difficulty: str | None, window: str | None = None) -> str:
    """Canonical event name for a *course*, shared by every parser so its sources converge.

    One term's course is described by several sheets — a weekly roster, a Level 2 tally, the
    dancecloud booking export — under different names. Building the event name from the same
    parts collapses them into one event. The export tail and a year suffix are handled as for
    ``_booking_event_name``; on top, the *level* (difficulty) is folded in when the term text
    doesn't already carry it. The level is the axis that separates concurrent courses in one
    term, and it lives in the tab title for a multi-course workbook ('Level 1'/'Level 2' tabs,
    generic filename) but in the filename for a booking export — folding it in either way lets
    those two meet while keeping the two levels of one workbook apart.

    The *time* axis is folded in the same way: when the term name has no month of its own, the
    session-date ``window`` replaces the plain year so ESDS's reused 'Term A/B' labels separate
    within a year. A name that already names its months ('May-Jun 2025') is trusted as-is, both
    to avoid repetition and because all its sources share that name and so stay converged.
    """
    base = _EXPORT_SUFFIX_RE.sub('', _strip_attendance(term)).strip()
    if difficulty and difficulty.lower() not in base.lower():
        base = f'{base} {difficulty}'.strip()
    if window and not _has_month_name(base):
        return f'{base} ({window})'
    return f'{base} ({year})' if year else base


def _course_activity_name(activity_type: ActivityType, difficulty: str | None, date_iso: str) -> str:
    """Canonical name for one course session, shared by every parser so its sources converge.

    A course session recurs across sources under different labels — '2022-06-16' (roster),
    'Level 2 Classes (Week 1)' (tally), 'Lesson' (booking export) — which keeps re-ingests of
    one term in separate activities. Naming it purely from what identifies the session — its
    level (difficulty), or its type when no level applies, plus its date — makes those sources
    land on one ``(event, name, date)`` key. A social's difficulty is 'social', so socials and
    lessons on the same night stay distinct without needing the type spelled out separately.
    """
    label = difficulty or str(activity_type)
    return f'{label} ({date_iso})'


def _event_name(term: str, title: str, year: int | None, event_type: EventType, window: str | None = None) -> str:
    """Event name for a title-bearing sheet: canonical for a course, else term + tidied title.

    The booking export, which has no meaningful tab title, builds its non-course name elsewhere.
    """
    if event_type == EventType.COURSE:
        return _course_event_name(term, year, _difficulty_for(title, term), window)
    return f'{term}: {_strip_attendance(title)}'


def _activity_name(
    raw_name: str, activity_type: ActivityType, difficulty: str | None, date_iso: str, event_type: EventType
) -> str:
    """Activity name: canonical for a course session, else the parser's own raw label."""
    if event_type == EventType.COURSE:
        return _course_activity_name(activity_type, difficulty, date_iso)
    return raw_name


# The registered parsers, in dispatch order. The dispatcher tries each per sheet and uses the
# first whose ``matches`` returns True.
PARSERS: list[Parser] = [
    StockbridgeSwingoutParser(),
    RosterParser(),
    Level2TallyParser(),
    Level2CountGridParser(),
    L2SOAttendanceParser(),
    SocialRegisterParser(),
    DancecloudActivityParser(),
    BookingExportParser(),
]
