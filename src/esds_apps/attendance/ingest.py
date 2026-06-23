"""Attendance ingestion — the folder dispatcher.

Walks the pseudonymised ``attendance_outputs`` tree and feeds each worksheet to the first
layout parser that claims it (the parsers live in ``parsers``). It derives the per-workbook
context the parsers need — the dominant year, the Week-1 anchor and the term label — decides
which sheets/workbooks to skip outright, and reports any sheet no parser claimed (those would
need a new parser) as well as the redundant 'Attendees' rollups.
"""

import re
from collections import Counter
from dataclasses import dataclass, field
from datetime import datetime

import openpyxl

from esds_apps.attendance.attendance_db import AttendanceDb
from esds_apps.attendance.parsers import PARSERS, _parse_dt, _strip_attendance


@dataclass
class IngestReport:
    handled: list[tuple] = field(default_factory=list)  # (file, sheet, parser_name)
    unhandled: list[tuple] = field(default_factory=list)  # (file, sheet)
    redundant: list[tuple] = field(default_factory=list)  # (file, sheet) — captured elsewhere, intentionally skipped

    def summary(self) -> str:
        """Human-readable report of what was ingested and what was skipped."""
        lines = [
            f'{len(self.handled)} sheet(s) ingested, {len(self.unhandled)} need a parser, '
            f'{len(self.redundant)} redundant.'
        ]
        for f, s, p in self.handled:
            lines.append(f'  [{p}] {f} :: {s}')
        if self.redundant:
            lines.append("Skipped (redundant — same bookings as the 'Attendees By Activity' tab in the file):")
            for f, s in self.redundant:
                lines.append(f'  - {f} :: {s}')
        if self.unhandled:
            lines.append('Skipped (no matching parser — would need a new one):')
            for f, s in self.unhandled:
                lines.append(f'  - {f} :: {s}')
        return '\n'.join(lines)


def _term_from(path) -> str:
    return _strip_attendance(path.stem.removesuffix('_pseudonymised'))


# A four-digit 20xx year, used to recover the year from the path when no dated cell carries it.
_YEAR_IN_PATH_RE = re.compile(r'20\d\d')


def _resolve_year(wb, path=None) -> int | None:
    """The workbook's dominant year: from its dated cells, else from the file/folder name.

    Date cells are authoritative and win when present. Some early sheets (the 2022 'Class List'
    rosters) carry no real date — every session date lives inside a 'Week N (DD Mon)' label, with
    the year only in the filename ('... 2022-02-07 ...') and the 'YYYY Attendance Records' folder.
    For those, fall back to the most common 20xx token across the path (the export-time HHMM tail
    like '2029' appears once, the real year twice, so the count breaks the tie correctly).
    """
    years: Counter = Counter()
    for ws in wb.worksheets:
        for row in ws.iter_rows(values_only=True):
            for value in row:
                dt = _parse_dt(value)
                if dt is not None:
                    years[dt.year] += 1
    if years:
        return years.most_common(1)[0][0]
    if path is not None:
        path_years = Counter(int(y) for y in _YEAR_IN_PATH_RE.findall(str(path)))
        if path_years:
            return path_years.most_common(1)[0][0]
    return None


def _week_anchor(wb) -> datetime | None:
    """The earliest date anywhere in the workbook — the Week-1 date for weekly-block tabs.

    Older 'Week N' sheets lost their per-column date formulas in pseudonymisation, and the
    count-grid tabs (e.g. 'Level 2-3') never carried dates at all. A workbook like these is
    one weekly term, so an undated 'Week N' is placed at this anchor + 7*(N-1). Tabs that
    carry their own session dates never consult it.
    """
    earliest: datetime | None = None
    for ws in wb.worksheets:
        for row in ws.iter_rows(values_only=True):
            for value in row:
                dt = _parse_dt(value)
                if dt is not None and (earliest is None or dt < earliest):
                    earliest = dt
    return earliest


# Sheet titles (lower-cased) that never hold attendance: substrings to skip outright.
# 'member'/'mail' are membership and mailing-list dumps; 'exceptions'/'loyalty'/'tickets'
# are bookkeeping tabs that carry no per-session attendance.
_NON_ATTENDANCE_SHEETS = ('readme', 'member', 'mail', 'exceptions', 'loyalty', 'tickets')

# Whole workbooks that are membership purchases / dumps, not class or social attendance. Their
# booking sheets ('Attendees By Activity') look just like a class booking export, so they must be
# skipped by filename: a membership purchase is not turning up to anything we count.
_NON_ATTENDANCE_FILES = ('membership', 'master members')


# A dancecloud export carries a dated 'Attendees By Activity' view (one row per dancer per session)
# alongside two restatements of the same bookings: a plain 'Attendees' rollup with the per-session
# dates dropped, and (in the modern export) a 'Check-Ins' pivot of the same check-in marks. When the
# by-activity view is present, both restatements are wholly contained in it, so no parser claims them
# — but they should be reported as redundant, not as sheets that "need a new parser". They are only
# treated as redundant when their by-activity sibling actually exists: a lone 'Attendees' tab (e.g.
# the Feb-March classes workbook, which has 'Class List' + 'Attendees' and no by-activity view) is
# still flagged, because then it may be the only record of those bookings.
_BY_ACTIVITY_TITLE = 'attendees by activity'
_REDUNDANT_WHEN_BY_ACTIVITY = ('attendees', 'check-ins')


def _redundant_rollup_titles(wb) -> set[str]:
    """Titles of 'Attendees'/'Check-Ins' restatements made redundant by an 'Attendees By Activity' sibling."""
    titles = {ws.title.strip().lower() for ws in wb.worksheets}
    if _BY_ACTIVITY_TITLE not in titles:
        return set()
    return {ws.title for ws in wb.worksheets if ws.title.strip().lower() in _REDUNDANT_WHEN_BY_ACTIVITY}


def ingest_file(
    path,
    db: AttendanceDb,
    parsers: list | None = None,
    rel: str | None = None,
    report: IngestReport | None = None,
) -> IngestReport:
    """Ingest one already-pseudonymised ``.xlsx`` into ``db``; return its IngestReport.

    The dancers it references must already exist in the ``dancer`` table (pseudonymise the file
    first); the foreign keys reject attendance for an unknown dancer. Idempotent: the write API
    upserts, so re-ingesting the same file refreshes its rows rather than duplicating them.
    """
    parsers = parsers if parsers is not None else PARSERS
    report = report if report is not None else IngestReport()
    if path.name.startswith(('~$', '.~lock')):
        return report
    if any(token in path.stem.lower() for token in _NON_ATTENDANCE_FILES):
        return report  # membership / master-members workbooks are not class or social attendance
    rel = rel if rel is not None else path.name
    wb = openpyxl.load_workbook(path, data_only=False)
    term, year, anchor = _term_from(path), _resolve_year(wb, path), _week_anchor(wb)
    redundant_titles = _redundant_rollup_titles(wb)
    for ws in wb.worksheets:
        ws_lower = ws.title.strip().lower()
        if any(pattern in ws_lower for pattern in _NON_ATTENDANCE_SHEETS):
            continue  # skip non-attendance sheets
        parser = next((p for p in parsers if p.matches(ws)), None)
        if parser is None:
            # The plain 'Attendees' rollup of a dancecloud export is the same bookings as its
            # 'Attendees By Activity' sibling minus the dates — captured, not unparsed.
            bucket = report.redundant if ws.title in redundant_titles else report.unhandled
            bucket.append((rel, ws.title))
            continue
        ingest_id = db.start_ingest(rel, ws.title)
        parser.parse(ws, db, term=term, year=year, ingest_id=ingest_id, week_anchor=anchor)
        report.handled.append((rel, ws.title, parser.name))
    return report


def ingest_folder(output_root, db: AttendanceDb, parsers: list | None = None) -> IngestReport:
    """Walk the attendance_outputs tree, dispatch each sheet to a matching parser."""
    parsers = parsers if parsers is not None else PARSERS
    report = IngestReport()
    for path in sorted(output_root.rglob('*.xlsx')):
        ingest_file(path, db, parsers=parsers, rel=path.relative_to(output_root).as_posix(), report=report)
    return report
