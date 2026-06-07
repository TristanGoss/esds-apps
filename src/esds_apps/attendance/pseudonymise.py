"""The xlsx pseudonymisation pipeline.

Reads an attendance workbook, detects which columns hold names/emails, replaces each
person's PII cells with a stable ``DNC-XXXXXXXX`` dancer ID (the first PII column becomes
``dancer_id``, the rest ``redacted``), and writes the redacted copy. The dancer-ID mapping
itself lives in the encrypted store (``pseudonyms_db``); this module is the spreadsheet I/O
and column-detection layer on top of it.

Two quirks of the real sheets shape the code: registers are sometimes printed as several
identical column blocks side by side (each block is a separate person per row — see
``_uniquify`` / ``_RecordBlock``), and aggregate tally sheets carry no PII and must be copied
through cell-for-cell rather than rebuilt from a header-keyed table.
"""

import re
from dataclasses import dataclass
from pathlib import Path

import openpyxl

from esds_apps.attendance.pseudonyms_db import (
    _CANONICAL_NAME_ORDER,
    DbContext,
    get_or_create_dancer_id,
    open_db,
)

EMAIL_RE = re.compile(r'^[a-zA-Z0-9._%+\-]+@[a-zA-Z0-9.\-]+\.[a-zA-Z]{2,}$')
NAME_HEADERS = re.compile(
    r'\b(name|first[_\s-]?name|last[_\s-]?name|surname|forename|full[_\s-]?name)\b',
    re.IGNORECASE,
)
_MEMBER_EXACT = re.compile(r'^member$', re.IGNORECASE)
EMAIL_HEADERS = re.compile(r'\be[_\s-]?mail(\s*(address|addr\.?))?\s*$', re.IGNORECASE)

# Fullmatch variants for header-row detection — prevents matching header words
# that appear as substrings in note cells like 'NB: order A_Z by First Name'.
_NAME_CELL_EXACT = re.compile(
    r'first[_\s-]?name|last[_\s-]?name|surname|forename|full[_\s-]?name|name|member', re.IGNORECASE
)
_EMAIL_CELL_EXACT = re.compile(r'e[_\s-]?mail', re.IGNORECASE)

# Accepts letters (including accented), apostrophes, hyphens, spaces, and initials with periods.
# Rejects anything with digits or most punctuation — i.e. footer notes, not names.
_VALID_NAME_RE = re.compile(r"^[A-Za-zÀ-ž'\-\s\.]+$")

# Summary/aggregate row labels that sit in a name column on the attendance sheets (e.g. the
# 'Totals' / 'All Attendance' / 'Social Only' rows beneath the L2 register). These are pure
# letters, so _VALID_NAME_RE accepts them — without this blocklist each would be minted as a
# spurious dancer. Matched case-insensitively against the whole, stripped cell. No real surname
# collides with these, so rejecting them outright is safe.
_NON_NAME_LABELS = frozenset(
    {
        'total',
        'totals',
        'subtotal',
        'grand total',
        'overall',
        'all attendance',
        'attendance',
        'social only',
        'social-only',
        'level 2',
        'level 1',
        'lesson',
        'social',
        'mean',
        'average',
        'averages',
        'sum',
        'count',
    }
)
_EMAIL_MATCH_THRESHOLD = 0.5


def _canonical_name_key(col: str) -> str:
    col_l = col.lower()
    if re.search(r'first|fore', col_l):
        return 'first_name'
    if re.search(r'last|sur', col_l):
        return 'last_name'
    if re.search(r'name|member', col_l):
        return 'first_name'
    raise ValueError(f'Cannot map column {col!r} to a canonical name field (expected first/forename or last/surname)')


# ---------------------------------------------------------------------------
# XLSX I/O
# ---------------------------------------------------------------------------


# Some sheets repeat the whole column block side by side (e.g. the Tea Dance registers
# print two identical halves). Header-keyed row dicts would collapse those duplicate
# columns onto one key — losing data and clobbering the dancer_id rename — so duplicate
# headers are made unique for internal keying with this separator, then stripped back to
# their original text on output. The unit-separator char never occurs in real headers.
_DUP_SUFFIX = '\x1f'


def _uniquify(fieldnames: list[str]) -> list[str]:
    """Make duplicate header names unique so each column keeps its own dict key."""
    seen: dict[str, int] = {}
    out = []
    for name in fieldnames:
        seen[name] = seen.get(name, 0) + 1
        out.append(name if seen[name] == 1 else f'{name}{_DUP_SUFFIX}{seen[name]}')
    return out


def _original_header(name: str) -> str:
    """Strip the disambiguation suffix added by _uniquify, recovering the original header."""
    return name.split(_DUP_SUFFIX, 1)[0]


def _dup_index(name: str) -> int:
    """Header-occurrence index of a uniquified name (1 for the first/only occurrence).

    Repeated header blocks are how a register printed as several halves side by side
    encodes several independent person-records per row; the occurrence index groups each
    block's columns together.
    """
    name, _, occurrence = name.partition(_DUP_SUFFIX)
    return int(occurrence) if occurrence else 1


def _read_sheet(ws) -> tuple[list[str], list[dict], list[list]]:
    """Read an openpyxl worksheet. Returns (fieldnames, rows, prefix_rows).

    Scans for the first row where any cell exactly matches a known name/email
    header. Rows above that are returned as prefix_rows and written back verbatim.
    Duplicate headers are disambiguated (see _uniquify) so repeated column blocks
    don't collapse onto one dict key.
    """
    all_rows = [[str(c) if c is not None else '' for c in row] for row in ws.iter_rows(values_only=True)]

    header_idx = 0
    for i, row in enumerate(all_rows):
        if any(_NAME_CELL_EXACT.fullmatch(cell.strip()) or _EMAIL_CELL_EXACT.fullmatch(cell.strip()) for cell in row):
            header_idx = i
            break

    fieldnames = _uniquify(all_rows[header_idx]) if all_rows else []
    rows = []
    for raw in all_rows[header_idx + 1 :] if all_rows else []:
        if not any(cell.strip() for cell in raw):
            continue
        padded = (raw + [''] * len(fieldnames))[: len(fieldnames)]
        rows.append(dict(zip(fieldnames, padded)))

    return fieldnames, rows, all_rows[:header_idx]


def _write_sheet(
    ws_out,
    out_fieldnames: list[str],
    rows: list[dict],
    original_fieldnames: list[str],
    prefix_rows: list[list],
) -> None:
    for row in prefix_rows:
        ws_out.append(row)
    ws_out.append(out_fieldnames)
    for row in rows:
        ws_out.append([row[c] for c in original_fieldnames])


def _copy_sheet_verbatim(ws, ws_out) -> None:
    """Copy every cell value (including formulas) at its original coordinate.

    Used for sheets with no detected PII — e.g. aggregate tally grids. The
    header-keyed reconstruction in _read_sheet/_write_sheet collapses any sheet
    that isn't a flat name-per-row table (duplicate or empty header cells map
    multiple columns onto one dict key), so such sheets must be passed through
    untouched rather than rebuilt.
    """
    for row in ws.iter_rows():
        ws_out.append([cell.value for cell in row])


# ---------------------------------------------------------------------------
# Column detection
# ---------------------------------------------------------------------------


def detect_columns(fieldnames: list[str], rows: list[dict]) -> dict[str, list]:
    name_cols, email_cols = [], []
    for col in fieldnames:
        if EMAIL_HEADERS.search(col):
            email_cols.append(col)
        elif '/' not in col and (NAME_HEADERS.search(col) or _MEMBER_EXACT.fullmatch(col.strip())):
            name_cols.append(col)

    # Content sniff over at most 5 rows — short columns are common.
    for col in fieldnames:
        if col in email_cols:
            continue
        sample = [r[col] for r in rows[:5] if r.get(col, '').strip()]
        if sample and sum(bool(EMAIL_RE.match(v.strip())) for v in sample) / len(sample) > _EMAIL_MATCH_THRESHOLD:
            email_cols.append(col)

    return {'name_cols': name_cols, 'email_cols': email_cols}


# ---------------------------------------------------------------------------
# Sheet-level pseudonymisation (internal)
# ---------------------------------------------------------------------------


@dataclass
class _RecordBlock:
    """One person-record's PII columns within a row (one half of a side-by-side register)."""

    id_col: str  # PII column that becomes 'dancer_id'
    redact_cols: list[str]  # remaining PII columns, redacted
    name_cols: list[str]  # name columns in this block
    email_cols: list[str]  # email columns in this block


def _record_fields(row: dict, name_cols: list[str], email_cols: list[str]) -> tuple[dict | None, dict | None]:
    """Extract (name_fields, email_fields) for one person-record from a row, or (None, None)."""
    raw_name = {}
    for c in name_cols:
        val = row.get(c, '').strip()
        if val and _VALID_NAME_RE.match(val) and val.lower() not in _NON_NAME_LABELS:
            raw_name[_canonical_name_key(c)] = val
    name_fields = {k: raw_name[k] for k in _CANONICAL_NAME_ORDER if k in raw_name} or None

    email_val = next((row[c].strip() for c in email_cols if EMAIL_RE.match(row.get(c, '').strip())), None)
    email_fields = {'email': email_val} if email_val else None
    return name_fields, email_fields


def _pseudonymise_sheet(
    ws,
    ws_out,
    ctx: DbContext,
    name_cols: list | None = None,
    email_cols: list | None = None,
) -> tuple[list[dict], int]:
    """Pseudonymise one worksheet and write to ws_out. Returns (rows, new_count)."""
    fieldnames, rows, prefix_rows = _read_sheet(ws)

    detected = detect_columns(fieldnames, rows)
    _name_cols = name_cols if name_cols is not None else detected['name_cols']
    _email_cols = email_cols if email_cols is not None else detected['email_cols']
    print(f'  Name columns:  {_name_cols}')
    print(f'  Email columns: {_email_cols}')

    if not _name_cols and not _email_cols:
        # Nothing to redact (e.g. an aggregate tally sheet). Copy verbatim rather
        # than rebuilding via header-keyed reconstruction, which would destroy any
        # non-tabular layout. No PII is present, so this is safe.
        print('  No name/email columns detected — copying sheet through unchanged.')
        _copy_sheet_verbatim(ws, ws_out)
        return [], 0

    # A register printed as several halves side by side repeats the column block, so each
    # repeat is an independent person-record per row and gets its own dancer_id. Group the
    # PII columns by header-occurrence index into one block per record; within each block the
    # first PII column becomes 'dancer_id' and the rest 'redacted'. A normal single table is
    # just one block, so this preserves the ordinary case.
    pii_cols = sorted(_name_cols + _email_cols, key=lambda c: fieldnames.index(c))
    name_set, email_set = set(_name_cols), set(_email_cols)
    blocks = [
        _RecordBlock(
            id_col=cols[0],
            redact_cols=cols[1:],
            name_cols=[c for c in cols if c in name_set],
            email_cols=[c for c in cols if c in email_set],
        )
        for cols in ([c for c in pii_cols if _dup_index(c) == idx] for idx in sorted({_dup_index(c) for c in pii_cols}))
    ]

    col_rename = {}
    for block in blocks:
        col_rename[block.id_col] = 'dancer_id'
        col_rename.update({c: 'redacted' for c in block.redact_cols})
    out_fieldnames = [col_rename.get(c, _original_header(c)) for c in fieldnames]

    before = ctx.conn.execute('SELECT COUNT(*) FROM pseudonyms').fetchone()[0]
    for row in rows:
        for block in blocks:
            name_fields, email_fields = _record_fields(row, block.name_cols, block.email_cols)
            if not name_fields and not email_fields:
                continue  # no person in this half of this row — leave its cells as they are
            row[block.id_col] = get_or_create_dancer_id(ctx, name_fields, email_fields)
            for col in block.redact_cols:
                row[col] = 'redacted'

    new_count = ctx.conn.execute('SELECT COUNT(*) FROM pseudonyms').fetchone()[0] - before
    _write_sheet(ws_out, out_fieldnames, rows, fieldnames, prefix_rows)
    return rows, new_count


# ---------------------------------------------------------------------------
# Public entry points
# ---------------------------------------------------------------------------


def pseudonymise(  # noqa: PLR0913
    spreadsheet_path: Path,
    db_path: Path,
    passphrase: str,
    output_path: Path | None = None,
    name_cols: list | None = None,
    email_cols: list | None = None,
    suffix: str = '_pseudonymised',
) -> dict[str, list[dict]]:
    """Read an xlsx workbook, replace name/email cells with stable dancer IDs, write the redacted copy.

    Processes all sheets. Returns {sheet_name: rows} for each sheet.
    """
    ctx = open_db(db_path, passphrase)
    wb = openpyxl.load_workbook(spreadsheet_path)
    wb_out = openpyxl.Workbook()
    wb_out.remove(wb_out.active)

    result = {}
    total_new = 0
    for ws in wb.worksheets:
        print(f'\nSheet: {ws.title}')
        ws_out = wb_out.create_sheet(ws.title)
        rows, new_count = _pseudonymise_sheet(ws, ws_out, ctx, name_cols, email_cols)
        result[ws.title] = rows
        total_new += new_count

    ctx.conn.close()

    if output_path is None:
        p = Path(spreadsheet_path)
        output_path = p.parent / f'{p.stem}{suffix}{p.suffix}'

    wb_out.save(output_path)
    print(f'\nNew dancer IDs created: {total_new}')
    print(f'Output written to:      {output_path}')
    return result


def pseudonymise_folder(  # noqa: PLR0913
    input_dir: Path,
    output_dir: Path,
    db_path: Path,
    passphrase: str,
    suffix: str = '_pseudonymised',
    name_cols: list | None = None,
    email_cols: list | None = None,
) -> None:
    """Pseudonymise all xlsx files in input_dir (recursively), writing results to output_dir.

    Preserves subdirectory structure. Output filenames get suffix appended before the extension.
    All files share a single database so dancer IDs are consistent across the whole dataset.
    """
    input_dir = Path(input_dir)
    output_dir = Path(output_dir)
    ctx = open_db(db_path, passphrase)
    total_new = 0

    for src in sorted(input_dir.rglob('*.xlsx')):
        rel = src.relative_to(input_dir)
        dst = output_dir / rel.parent / f'{src.stem}{suffix}{src.suffix}'
        dst.parent.mkdir(parents=True, exist_ok=True)

        print(f'\n{rel}')
        wb = openpyxl.load_workbook(src)
        wb_out = openpyxl.Workbook()
        wb_out.remove(wb_out.active)

        for ws in wb.worksheets:
            print(f'  Sheet: {ws.title}')
            ws_out = wb_out.create_sheet(ws.title)
            _, new_count = _pseudonymise_sheet(ws, ws_out, ctx, name_cols, email_cols)
            total_new += new_count

        wb_out.save(dst)
        print(f'  → {dst}')

    ctx.conn.close()
    print(f'\nTotal new dancer IDs: {total_new}')
