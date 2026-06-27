from datetime import datetime

import openpyxl
import pytest

from esds_apps.attendance import parsers

from .workbooks import (
    _attended_register_ws,
    _blank_party_register_ws,
    _booking_by_activity_ws,
    _booking_list_ws,
    _booking_summary_ws,
    _booking_with_concession_ws,
    _booking_ws_on,
    _checked_in_by_activity_ws,
    _checked_in_multiday_class_ws,
    _checked_in_no_scans_ws,
    _checked_in_with_concession_ws,
    _class_list_roster_ws,
    _count_grid_undated_ws,
    _count_grid_ws,
    _dancer_list_no_marker_ws,
    _l1_attendance_2026_ws,
    _l2_so_absent_heavy_ws,
    _l2_so_ws,
    _roster_concession_ws,
    _roster_weekly_ws,
    _roster_ws,
    _social_register_ws,
    _swingout_ws,
    _tally_ws,
    _teachers_choice_tally_ws,
)

# ---- value helpers ----


@pytest.mark.parametrize('value', [True, 'True', 'TRUE', 'true', '=TRUE()', 'x', 'yes', '☑️', '✅'])
def test_is_true_positives(value):
    assert parsers._is_true(value)


@pytest.mark.parametrize('value', [False, 'False', '=FALSE()', 'Refunded', '❎', '❌', '', None, 0])
def test_is_true_negatives(value):
    assert not parsers._is_true(value)


@pytest.mark.parametrize(
    ('text', 'expected'),
    [
        ('Level 2 Classes', True),  # plural (older sheets)
        ('Level 2 Class', True),  # singular (2025-H2 sheets)
        ('Level 1 Classes', True),  # one-off tallies (Teacher's Choice) carry a Level 1 group too
        ('Level 1 Class', True),
        ('Social Only', True),
        ('Members', False),
        ('Level 1', False),  # the bare level name is not a class-group header
    ],
)
def test_is_group_header(text, expected):
    assert parsers._is_group_header(text) is expected


def test_parse_dt_handles_datetime_and_isoish_string():
    assert parsers._parse_dt(datetime(2025, 5, 22)).year == 2025
    assert parsers._parse_dt('2025-05-22 00:00:00').month == 5
    assert parsers._parse_dt('not a date') is None


def test_strip_attendance_handles_full_and_truncated_word():
    assert parsers._strip_attendance('Level 1 Attendance') == 'Level 1'
    assert parsers._strip_attendance('Level 2 & Social Only Attendanc') == 'Level 2 & Social Only'  # 31-char cut
    assert (
        parsers._strip_attendance('Sunday Social Attendees 2023-03-25') == 'Sunday Social 2023-03-25'
    )  # 'Attendees' too
    assert parsers._strip_attendance('Tea Dance 13th Oct') == 'Tea Dance 13th Oct'  # untouched


def test_strip_attendance_normalises_level_2_3():
    assert parsers._strip_attendance('Level 2-3') == 'Level 2'
    assert parsers._strip_attendance('Level 2/3 Attendance') == 'Level 2'
    assert parsers._strip_attendance('Level 2 & 3') == 'Level 2'
    assert parsers._strip_attendance('Levels 1-3 (Nov-Dec 2023)') == 'Levels 1-3 (Nov-Dec 2023)'  # untouched


@pytest.mark.parametrize(
    ('title', 'expected'),
    [
        ('Level 1 Attendance', 'course'),
        ('Level 2 & Social Only Attendanc', 'course'),  # 'level' beats 'social'
        ('Teachers Choice 12th Dec', 'course'),
        ('Sat 17th Workshop', 'workshop'),
        ('Begn Charleston 24th Feb', 'workshop'),
        ('Christmas Party 19th Dec', 'social'),
        ('Tea Dance 13th Oct', 'social'),
        ('Stockbridge Swingout', 'weekender'),  # swingout is a weekender, not a social
        ('Sun 27th Social', 'social'),
    ],
)
def test_event_type_for_classifies_by_title(title, expected):
    assert parsers._event_type_for(title) == expected


@pytest.mark.parametrize(
    ('texts', 'expected'),
    [
        (('Level 1 Class',), 'Level 1'),
        (('Level 2 Classes (Week 1)',), 'Level 2'),
        (('L3 taster',), 'Level 3'),
        (('Begn Charleston',), 'beginners'),
        (('Intermediate Lindy',), 'intermediate'),
        (('Advanced aerials',), 'advanced'),
        (('Social',), None),
        (('Week 1', 'Level 1 Attendance'), 'Level 1'),  # falls back to the second text
        (('Beginners drop-in', 'Level 2 term'), 'beginners'),  # first text wins
    ],
)
def test_difficulty_for(texts, expected):
    assert parsers._difficulty_for(*texts) == expected


@pytest.mark.parametrize(
    ('value', 'expected'),
    [
        ('Yes', 'member_or_concession'),
        ('yes', 'member_or_concession'),
        (True, 'member_or_concession'),
        ('No', 'ordinary'),
        (False, 'ordinary'),
        ('comp', 'unknown'),  # non-empty but unrecognised -> unknown, not NULL
        ('', None),
        ('   ', None),
        (None, None),
    ],
)
def test_concession_ticket_maps_boolean_ish_and_unknown(value, expected):
    assert parsers._concession_ticket(value) == expected


@pytest.mark.parametrize(
    ('header', 'expected'),
    [
        (('dancer_id', 'Concession', 'Week 1'), 1),
        (('dancer_id', 'Consession', 'Week 1'), 1),  # misspelling seen in Jan-Feb 2026
        (('dancer_id', 'Concessions', 'Week 1'), 1),  # trailing plural
        (('dancer_id', 'Ticket Type', 'Week 1'), None),  # not a boolean-ish flag
        (('dancer_id', 'redacted', 'Week 1'), None),
    ],
)
def test_concession_col_finds_flag_tolerating_misspelling(header, expected):
    assert parsers._concession_col(header) == expected


@pytest.mark.parametrize(
    'status, present, expected',
    [
        ('Confirmed', None, 'unknown'),  # a bought ticket, turnout never captured
        ('Confirmed', '', 'unknown'),
        ('Cancelled', None, 'absent'),  # booked then pulled out
        ('Confirmed', 'no show', 'absent'),
        ('Confirmed', 'refunded', 'absent'),
        ('Confirmed', 'refunded as resold on door', 'absent'),  # refund beats the 'door' word
        ('Confirmed', 'paid £10 cash on door', 'attended'),  # demonstrably there
        ('Confirmed', 'x', 'attended'),
        (None, None, 'unknown'),
    ],
)
def test_booking_status(status, present, expected):
    assert parsers._booking_status(status, present) == expected


@pytest.mark.parametrize(
    ('value', 'expected'),
    [
        (datetime(2026, 5, 10, 14, 3), True),  # a check-in timestamp
        ('Yes', True),  # the pivoted 'Check-Ins' tab's marker
        ('No', False),
        ('n/a', False),  # activity not on the ticket
        ('na', False),
        ('False', False),
        ('', False),
        ('   ', False),
        (None, False),
    ],
)
def test_is_checked_in(value, expected):
    assert parsers._is_checked_in(value) is expected


@pytest.mark.parametrize(
    ('checkin', 'status', 'sheet_has_checkins', 'expected'),
    [
        (datetime(2026, 5, 10), 'Confirmed', True, 'attended'),  # scanned in
        (None, 'Confirmed', True, 'absent'),  # blank but the scanner ran -> no-show
        (None, 'Confirmed', False, 'unknown'),  # scanner never ran -> uncaptured
        (datetime(2026, 5, 10), 'Cancelled', True, 'absent'),  # cancelled beats a stray check-in
        (None, 'Cancelled', False, 'absent'),  # cancelled is absent regardless
    ],
)
def test_checkin_status(checkin, status, sheet_has_checkins, expected):
    assert parsers._checkin_status(checkin, status, sheet_has_checkins) == expected


@pytest.mark.parametrize(
    ('label', 'expected'),
    [
        ('Friday Welcome Party', 'social'),
        ('30th Birthday Ball', 'social'),
        ("Tea dance with Dick Lee's Hot Club Hepcats", 'social'),
        ('Collegiate Shag Workshop', 'lesson'),
        ('Track A - Classes', 'lesson'),
        ('Lindy Technique', 'lesson'),
        ('Something unclassifiable', 'lesson'),  # falls back to the event-type default (workshop -> lesson)
    ],
)
def test_activity_type_from_name(label, expected):
    assert str(parsers._activity_type_from_name(label, parsers.EventType.WORKSHOP)) == expected


def test_dancecloud_event_name_strips_trailing_date_and_suffixes_year():
    """The export's date tail is dropped so recurring fixtures collapse to one year-tagged event."""
    assert parsers._dancecloud_event_name("Teachers' Workshop Attendees Apr 19th 2026", 2026) == (
        "Teachers' Workshop (2026)"
    )
    assert parsers._dancecloud_event_name('30 Years of ESDS March 20th-22nd 2026', 2026) == '30 Years of ESDS (2026)'
    assert parsers._dancecloud_event_name('Stroll Workshop and Tea Dance June 7th 2026', None) == (
        'Stroll Workshop and Tea Dance'
    )
    # The ISO export stamp dancecloud writes ('YYYY-MM-DD HHMM') is dropped, time and all, so
    # several re-exports of one party collapse to a single year-tagged event.
    assert (
        parsers._dancecloud_event_name('End of Term Party with Iain Ewing and the Chevaliers 2026-06-27 1153', 2026)
        == 'End of Term Party with Iain Ewing and the Chevaliers (2026)'
    )


def test_course_window_spans_session_months():
    """A course window is a single month or a Mmm-Mmm span over the session dates, year-suffixed."""
    assert parsers._course_window([datetime(2022, 6, 16), datetime(2022, 7, 21)], 2022) == 'Jun-Jul 2022'
    assert parsers._course_window([datetime(2022, 3, 31)], 2022) == 'Mar 2022'
    assert parsers._course_window([], 2022) is None  # no dates -> caller falls back to the bare year


def test_course_event_name_folds_window_only_when_name_lacks_a_month():
    """The window separates month-less 'Term A/B' names but is suppressed when the name names months."""
    # 'Term B' carries no month, so the window replaces the bare year and splits the reused label
    assert (
        parsers._course_event_name('Level 1 Fundamentals Term B', 2022, 'Level 1', 'Mar-Apr 2022')
        == 'Level 1 Fundamentals Term B (Mar-Apr 2022)'
    )
    # the newer termly name already says 'May-Jun', so repeating it is noise: keep the plain year
    assert parsers._course_event_name('May-Jun 2025', 2025, 'Level 1', 'May 2025') == 'May-Jun 2025 Level 1 (2025)'
    # no window available -> bare year, as before
    assert (
        parsers._course_event_name('Level 1 Fundamentals Term B', 2022, 'Level 1')
        == 'Level 1 Fundamentals Term B (2022)'
    )


# ---- matchers ----


def test_roster_matches_roster_only():
    assert parsers.RosterParser().matches(_roster_ws())
    assert not parsers.RosterParser().matches(_tally_ws())
    assert not parsers.RosterParser().matches(_l2_so_ws())  # rejects category strings


def test_tally_matches_tally_only():
    assert parsers.Level2TallyParser().matches(_tally_ws())
    assert not parsers.Level2TallyParser().matches(_roster_ws())


def test_count_grid_matches_grid_only():
    assert parsers.Level2CountGridParser().matches(_count_grid_ws())
    assert parsers.Level2CountGridParser().matches(_count_grid_undated_ws())  # bare 'Week N' still matches
    assert not parsers.Level2CountGridParser().matches(_tally_ws())  # has COUNTIF
    assert not parsers.Level2CountGridParser().matches(_roster_ws())
    assert not parsers.Level2TallyParser().matches(_count_grid_ws())  # no COUNTIF


def test_tally_matches_one_off_single_date_tally():
    """A tally with COUNTIFs and groups but no week labels matches via a day+month in the tab name."""
    assert parsers.Level2TallyParser().matches(_teachers_choice_tally_ws())


def test_roster_matches_class_list_with_embedded_week_dates():
    """A 'Class List' roster whose only dates are inside 'Week N (DD Mon)' headers is a roster.

    Detection must not need the year (it isn't in the sheet), and the 'Lead / Follow' column is
    not a session.
    """
    assert parsers.RosterParser().matches(_class_list_roster_ws())


def test_l2_so_matches_l2_so_only():
    assert parsers.L2SOAttendanceParser().matches(_l2_so_ws())
    assert not parsers.L2SOAttendanceParser().matches(_roster_ws())


def test_absent_heavy_l2_so_dispatches_to_l2_so_not_roster():
    """An L2 & SO sheet dominated by 'Absent' must route to the L2/SO parser, not the roster.

    'Absent' doubles as a roster 'no' marker, so RosterParser.matches() legitimately fires here;
    the dispatch order (L2/SO before roster) is what keeps the sheet from being mis-read as a plain
    Level 2 roster — which dropped the social and zeroed the Level 2 turnout for May-Jun 2026.
    """
    ws = _l2_so_absent_heavy_ws()
    assert parsers.RosterParser().matches(ws)  # the collision: 'Absent' reads as a marker
    winner = next((p for p in parsers.PARSERS if p.matches(ws)), None)
    assert winner is not None and winner.name == 'l2_so_attendance'


def test_absent_heavy_l2_so_parses_lesson_and_social(db):
    """Parsing the absent-heavy sheet yields a non-empty Level 2 lesson and a social, not all-absent."""
    parsers.L2SOAttendanceParser().parse(_l2_so_absent_heavy_ws(), db, term='May 2026', year=2026, ingest_id=None)
    l2_attended = db.conn.execute(
        'SELECT COUNT(*) FROM attendance at JOIN activity a USING(activity_id) '
        "WHERE a.difficulty = 'Level 2' AND at.status = 'attended'"
    ).fetchone()[0]
    socials = db.conn.execute("SELECT COUNT(*) FROM activity WHERE activity_type = 'social'").fetchone()[0]
    assert l2_attended >= 1  # the 'Level 2 & Social' rows are recorded as Level 2 attendees
    assert socials >= 1  # the 'Social-Only' row creates a social activity


def test_social_register_matches_only_undated_present_sheets():
    assert parsers.SocialRegisterParser().matches(_social_register_ws())
    assert not parsers.SocialRegisterParser().matches(_roster_ws())  # has dated session columns
    assert not parsers.SocialRegisterParser().matches(_tally_ws())  # no dancer_id


def test_social_register_matches_attended_marker_layout():
    """The 'Attended' single-block layout (Sunday Socials) is also a social register."""
    assert parsers.SocialRegisterParser().matches(_attended_register_ws())
    assert not parsers.RosterParser().matches(_attended_register_ws())  # no dated sessions


def test_social_register_rejects_booking_list_with_empty_present_column():
    """A booking list with a mostly-empty 'present' notes column is not an attendance register."""
    assert not parsers.SocialRegisterParser().matches(_booking_list_ws())


def test_social_register_rejects_dancer_list_without_marker_column():
    """A dancer_id list with no Present?/Attended column is not a register (guards over-claiming)."""
    assert not parsers.SocialRegisterParser().matches(_dancer_list_no_marker_ws())


def test_social_register_matches_blank_marker_party_list():
    """An allocation list with an entirely blank Present? column is still a register (turnout unknown)."""
    assert parsers.SocialRegisterParser().matches(_blank_party_register_ws())


def test_booking_export_matches_only_dated_or_present_status_lists():
    p = parsers.BookingExportParser()
    assert p.matches(_booking_by_activity_ws())  # dated booking view
    assert p.matches(_booking_list_ws())  # single-event list with a present-note column
    assert not p.matches(_booking_summary_ws())  # Status but no Date and no present -> not claimed
    assert not p.matches(_roster_ws())  # an attendance roster has no booking Status column
    assert not p.matches(_checked_in_by_activity_ws())  # 'Checked In' -> stands aside for the modern parser


def test_dancecloud_activity_matches_only_the_checked_in_export():
    p = parsers.DancecloudActivityParser()
    assert p.matches(_checked_in_by_activity_ws())  # has dancer_id + Activity + Date + Checked In
    assert p.matches(_checked_in_no_scans_ws())  # an all-blank Checked In column still identifies the layout
    assert not p.matches(_booking_by_activity_ws())  # older export: no Checked In column
    assert not p.matches(_roster_ws())


# ---- roster parsing ----


def test_roster_parse_creates_event_and_activities(db):
    parsers.RosterParser().parse(_roster_ws(), db, term='May-Jun 2025', year=2025, ingest_id=None)
    ev = db.conn.execute('SELECT name, event_type FROM event').fetchone()
    # canonical course event name: 'Attendance' stripped, level folded in, year suffixed
    assert ev == ('May-Jun 2025 Level 1 (2025)', 'course')
    acts = db.conn.execute('SELECT name, date FROM activity ORDER BY date').fetchall()
    # canonical course activity names: level (difficulty) + date, shared across every source
    assert acts == [('Level 1 (2025-05-22)', '2025-05-22'), ('Level 1 (2025-05-29)', '2025-05-29')]
    # difficulty comes from the sheet title ('Level 1 Attendance') since 'Week N' has none
    assert {r[0] for r in db.conn.execute('SELECT DISTINCT difficulty FROM activity')} == {'Level 1'}


def test_roster_2026_format_with_dates_in_header(db):
    """The 2026 L1 format has dates in the header row, not the row above."""
    parsers.RosterParser().parse(_l1_attendance_2026_ws(), db, term='Jan 2026', year=2026, ingest_id=None)
    ev = db.conn.execute('SELECT name, event_type FROM event').fetchone()
    assert ev == ('Jan 2026 Level 1 (2026)', 'course')  # '2026 L1' title yields the 'Level 1' fold
    acts = db.conn.execute('SELECT name, date FROM activity ORDER BY date').fetchall()
    # canonical course activity names from level + date; the header 'Week N' labels are not used
    assert acts == [('Level 1 (2026-01-15)', '2026-01-15'), ('Level 1 (2026-01-22)', '2026-01-22')]
    # difficulty parsed from the title '2026 L1 Attendance' → 'Level 1'
    assert {r[0] for r in db.conn.execute('SELECT DISTINCT difficulty FROM activity')} == {'Level 1'}
    # Verify the attendance was recorded correctly
    rows = db.conn.execute(
        'SELECT dancer_id, COUNT(*) as activity_count FROM attendance GROUP BY dancer_id ORDER BY dancer_id'
    ).fetchall()
    assert rows == [('DNC-1', 2), ('DNC-2', 2)]  # DNC-1 and DNC-2 each have 2 attendance records


def test_roster_parse_reads_mixed_truthy_markers(db):
    parsers.RosterParser().parse(_roster_ws(), db, term='T', year=2025, ingest_id=None)
    # Week 1: DNC-1 (True), DNC-2 (☑️), DNC-3 ('x') all attended
    rows = db.conn.execute(
        'SELECT dancer_id, status FROM attendance JOIN activity USING(activity_id) '
        "WHERE name='Level 1 (2025-05-22)' ORDER BY dancer_id"
    ).fetchall()
    assert rows == [('DNC-1', 'attended'), ('DNC-2', 'attended'), ('DNC-3', 'attended')]


def test_roster_parse_duplicate_ticket_becomes_anonymous_count(db):
    parsers.RosterParser().parse(_roster_ws(), db, term='T', year=2025, ingest_id=None)
    # DNC-1 used two tickets in Week 1 → one named row + one anonymous head.
    anon = db.conn.execute(
        'SELECT ticket_type, head_count FROM attendance_count JOIN activity USING(activity_id) '
        "WHERE name='Level 1 (2025-05-22)'"
    ).fetchall()
    assert anon == [(None, 1)]
    # Week 2: DNC-1's two tickets, only one used → named attended, no anonymous extra.
    assert db.conn.execute(
        "SELECT COUNT(*) FROM attendance_count JOIN activity USING(activity_id) WHERE name='Level 1 (2025-05-29)'"
    ).fetchone() == (0,)


def test_roster_refunded_and_false_are_absent(db):
    parsers.RosterParser().parse(_roster_ws(), db, term='T', year=2025, ingest_id=None)
    w2 = dict(
        db.conn.execute(
            'SELECT dancer_id, status FROM attendance JOIN activity USING(activity_id) '
            "WHERE name='Level 1 (2025-05-29)'"
        ).fetchall()
    )
    assert w2['DNC-2'] == 'absent'  # 'False'
    assert w2['DNC-3'] == 'absent'  # 'Refunded'


def test_roster_blank_column_is_unknown_not_absent(db):
    """A dated session column nobody marked means turnout unrecorded, not a class-wide no-show."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Level 1 Attendance'
    ws['C1'], ws['D1'] = datetime(2022, 6, 16), datetime(2022, 6, 23)  # two dated weeks
    ws['B2'], ws['C2'], ws['D2'] = 'dancer_id', 'Week 1', 'Week 2'
    ws['B3'], ws['C3'] = 'DNC-1', 'x'  # Week 1 marked; Week 2 entirely blank
    ws['B4'] = 'DNC-2'
    parsers.RosterParser().parse(ws, db, term='T', year=2022, ingest_id=None)
    w1 = dict(
        db.conn.execute(
            'SELECT dancer_id, status FROM attendance JOIN activity USING(activity_id) '
            "WHERE name='Level 1 (2022-06-16)'"
        ).fetchall()
    )
    w2 = dict(
        db.conn.execute(
            'SELECT dancer_id, status FROM attendance JOIN activity USING(activity_id) '
            "WHERE name='Level 1 (2022-06-23)'"
        ).fetchall()
    )
    assert w1 == {'DNC-1': 'attended', 'DNC-2': 'absent'}  # a marked column is a real register
    assert w2 == {'DNC-1': 'unknown', 'DNC-2': 'unknown'}  # an unmarked column fabricates no absences


def test_roster_infers_weekly_dates_from_anchor(db):
    """Only Week 1 is dated; Weeks 2-3 are placed at anchor + 7*(N-1)."""
    parsers.RosterParser().parse(
        _roster_weekly_ws(), db, term='Nov-Dec 2023', year=2023, ingest_id=None, week_anchor=datetime(2023, 11, 9)
    )
    acts = db.conn.execute('SELECT name, date FROM activity ORDER BY date').fetchall()
    assert acts == [
        ('Level 1 (2023-11-09)', '2023-11-09'),
        ('Level 1 (2023-11-16)', '2023-11-16'),
        ('Level 1 (2023-11-23)', '2023-11-23'),
    ]


def test_roster_without_anchor_keeps_only_dated_weeks(db):
    """No anchor: undated 'Week N' columns are dropped rather than guessed."""
    parsers.RosterParser().parse(_roster_weekly_ws(), db, term='Nov-Dec 2023', year=2023, ingest_id=None)
    assert db.conn.execute('SELECT name FROM activity').fetchall() == [('Level 1 (2023-11-09)',)]


def test_roster_parses_embedded_week_dates_with_year(db):
    """'Week N (DD Mon)' columns date themselves once the year (from the filename) is supplied."""
    parsers.RosterParser().parse(
        _class_list_roster_ws(), db, term='ESDS Level 2 Technique Feb March', year=2022, ingest_id=None
    )
    acts = db.conn.execute('SELECT name, date FROM activity ORDER BY date').fetchall()
    assert acts == [
        ('Level 2 (2022-02-10)', '2022-02-10'),
        ('Level 2 (2022-02-17)', '2022-02-17'),
        ('Level 2 (2022-02-24)', '2022-02-24'),
    ]
    # DNC-1 attended weeks 1 and 3, absent week 2; the 'Lead / Follow' column produced no activity
    rows = dict(
        db.conn.execute(
            'SELECT a.date, att.status FROM attendance att JOIN activity a USING(activity_id) '
            "WHERE att.dancer_id='DNC-1' ORDER BY a.date"
        ).fetchall()
    )
    assert rows == {'2022-02-10': 'attended', '2022-02-17': 'absent', '2022-02-24': 'attended'}


def test_roster_embedded_week_dates_need_a_year(db):
    """Without a year the embedded-week columns can't be dated, so nothing is recorded (no crash)."""
    parsers.RosterParser().parse(_class_list_roster_ws(), db, term='T', year=None, ingest_id=None)
    assert db.conn.execute('SELECT COUNT(*) FROM activity').fetchone() == (0,)


def test_roster_captures_concession_as_ticket_type(db):
    """A roster's Concession column maps each dancer into ordinary vs member-or-concession."""
    parsers.RosterParser().parse(_roster_concession_ws(), db, term='T', year=2025, ingest_id=None)
    tickets = dict(db.conn.execute('SELECT DISTINCT dancer_id, ticket_type FROM attendance').fetchall())
    assert tickets == {
        'DNC-1': 'ordinary',  # 'No'; the blank duplicate row did not overwrite it
        'DNC-2': 'member_or_concession',  # 'Yes'
        'DNC-4': 'unknown',  # unrecognised value
        'DNC-9': None,  # blank throughout -> no information
    }


def test_roster_without_concession_column_leaves_ticket_type_null(db):
    """The plain roster (no Concession column) records no ticket type."""
    parsers.RosterParser().parse(_roster_ws(), db, term='T', year=2025, ingest_id=None)
    assert {r[0] for r in db.conn.execute('SELECT DISTINCT ticket_type FROM attendance')} == {None}


def test_roster_social_event_forces_social_activities(db):
    # 'Week 1'/'Week 2' headers contain no 'social', but the event classifies as social,
    # so the activities must still be social — not lesson.
    parsers.RosterParser().parse(_roster_ws(title='Christmas Party 19th Dec'), db, term='T', year=2024, ingest_id=None)
    assert db.conn.execute('SELECT event_type FROM event').fetchone() == ('social',)
    assert {r[0] for r in db.conn.execute('SELECT DISTINCT activity_type FROM activity')} == {'social'}
    # social activities get difficulty 'social', not NULL
    assert {r[0] for r in db.conn.execute('SELECT DISTINCT difficulty FROM activity')} == {'social'}


# ---- tally parsing ----


def test_tally_parse_counts_ticks_per_ticket_type(db):
    parsers.Level2TallyParser().parse(_tally_ws(), db, term='May-Jun 2025', year=2025, ingest_id=None)
    rows = db.conn.execute(
        'SELECT a.name, a.date, ac.ticket_type, ac.head_count '
        'FROM attendance_count ac JOIN activity a USING(activity_id) ORDER BY ac.ticket_type'
    ).fetchall()
    assert rows == [
        ('Level 2 (2025-05-22)', '2025-05-22', 'concession', 2),
        ('Level 2 (2025-05-22)', '2025-05-22', 'member', 3),
    ]


def test_tally_parse_sets_event_and_activity_type(db):
    parsers.Level2TallyParser().parse(_tally_ws(), db, term='T', year=2025, ingest_id=None)
    assert db.conn.execute('SELECT event_type FROM event').fetchone() == ('course',)
    assert db.conn.execute('SELECT DISTINCT activity_type FROM activity').fetchone() == ('lesson',)
    # difficulty parsed from the group label 'Level 2 Classes'
    assert db.conn.execute('SELECT DISTINCT difficulty FROM activity').fetchone() == ('Level 2',)


def test_tally_parse_one_off_uses_title_date_and_three_groups(db):
    """A one-off tally dates every group from the tab name and names its event from the tab.

    Level 1 / Level 2 / Social Only each become a 12-Dec activity with their head counts; the event
    is named from the tab so it stays distinct from the term's weekly Level 1/2 events.
    """
    parsers.Level2TallyParser().parse(_teachers_choice_tally_ws(), db, term='Nov-Dec 2024', year=2024, ingest_id=None)
    assert db.conn.execute('SELECT name, event_type FROM event').fetchone() == (
        'Nov-Dec 2024: Teachers Choice 12th Dec',
        'course',
    )
    rows = db.conn.execute(
        'SELECT a.name, a.activity_type, a.difficulty, a.date, ac.head_count '
        'FROM attendance_count ac JOIN activity a USING(activity_id) ORDER BY a.name'
    ).fetchall()
    assert rows == [
        ('Level 1 (2024-12-12)', 'lesson', 'Level 1', '2024-12-12', 3),
        ('Level 2 (2024-12-12)', 'lesson', 'Level 2', '2024-12-12', 2),
        ('social (2024-12-12)', 'social', 'social', '2024-12-12', 1),
    ]


# ---- count-grid parsing ----


def test_count_grid_parse_records_counts_and_skips_mean_row(db):
    parsers.Level2CountGridParser().parse(_count_grid_ws(), db, term='Apr-May 2024', year=2024, ingest_id=None)
    # two week rows -> two activities; the 'Mean' row has no week label and is ignored
    acts = db.conn.execute('SELECT name, date, activity_type, difficulty FROM activity ORDER BY date').fetchall()
    assert acts == [
        ('Level 2 (2024-04-11)', '2024-04-11', 'lesson', 'Level 2'),
        ('Level 2 (2024-04-18)', '2024-04-18', 'lesson', 'Level 2'),
    ]
    week1 = db.conn.execute(
        'SELECT ac.ticket_type, ac.head_count FROM attendance_count ac JOIN activity a USING(activity_id) '
        "WHERE a.name='Level 2 (2024-04-11)' ORDER BY ac.ticket_type"
    ).fetchall()
    assert week1 == [('concession', 3), ('member', 16), ('ordinary', 8)]


def test_count_grid_undated_uses_anchor_and_normalises_level(db):
    """Bare 'Week N' rows fall back to anchor + 7*(N-1); the 'Level 2-3' title becomes 'Level 2'."""
    parsers.Level2CountGridParser().parse(
        _count_grid_undated_ws(), db, term='Nov-Dec 2023', year=2023, ingest_id=None, week_anchor=datetime(2023, 11, 9)
    )
    assert db.conn.execute('SELECT name FROM event').fetchone() == ('Nov-Dec 2023 Level 2 (2023)',)
    acts = db.conn.execute('SELECT name, date, difficulty FROM activity ORDER BY date').fetchall()
    assert acts == [
        ('Level 2 (2023-11-09)', '2023-11-09', 'Level 2'),
        ('Level 2 (2023-11-16)', '2023-11-16', 'Level 2'),
    ]


# ---- L2 & SO attendance parsing ----


def test_l2_so_parse_maps_categories_to_lesson_and_social(db):
    """Each night yields a Level 2 lesson and/or a social; activities are created lazily."""
    parsers.L2SOAttendanceParser().parse(_l2_so_ws(), db, term='Jan 2026', year=2026, ingest_id=None)
    ev = db.conn.execute('SELECT name, event_type FROM event').fetchone()
    assert ev == ('Jan 2026 Level 2 (2026)', 'course')

    # Only the (date, kind) pairs that actually occur are created — no phantom activities,
    # and the '=countif(...)' summary row is never read as a category.
    acts = {(r[0], r[1], r[2]) for r in db.conn.execute('SELECT name, activity_type, difficulty FROM activity')}
    assert acts == {
        ('Level 2 (2026-01-15)', 'lesson', 'Level 2'),  # DNC-1 attended, DNC-2 absent
        ('social (2026-01-22)', 'social', 'social'),  # DNC-1 social-only
        ('Level 2 (2026-01-22)', 'lesson', 'Level 2'),  # DNC-2 attended
    }


def test_l2_so_parse_records_attendance_by_category(db):
    """Level 2 & Social -> Level 2 attendee; Social-Only -> social; Absent -> Level 2 no-show."""
    parsers.L2SOAttendanceParser().parse(_l2_so_ws(), db, term='Jan 2026', year=2026, ingest_id=None)

    # DNC-1: 'Level 2 & Social' on 01-15 -> Level 2 attended; 'Social-Only' on 01-22 -> social.
    dnc1 = db.conn.execute(
        'SELECT a.name, att.status FROM attendance att JOIN activity a USING(activity_id) '
        "WHERE att.dancer_id='DNC-1' ORDER BY a.name"
    ).fetchall()
    assert dnc1 == [('Level 2 (2026-01-15)', 'attended'), ('social (2026-01-22)', 'attended')]

    # DNC-2: 'Absent' on 01-15 -> Level 2 roster row, did not attend; 'Level 2 & Social' on 01-22.
    dnc2 = db.conn.execute(
        'SELECT a.name, att.status FROM attendance att JOIN activity a USING(activity_id) '
        "WHERE att.dancer_id='DNC-2' ORDER BY a.name"
    ).fetchall()
    assert dnc2 == [('Level 2 (2026-01-15)', 'absent'), ('Level 2 (2026-01-22)', 'attended')]


# ---- social register parsing (Tea Dances etc.) ----


def test_social_register_parse_one_dated_social_from_title(db):
    """One social event/activity; the date comes from the title since the sheet has none."""
    parsers.SocialRegisterParser().parse(
        _social_register_ws(), db, term='Tea Dance 25 Feb 2024', year=None, ingest_id=None
    )
    assert db.conn.execute('SELECT name, event_type FROM event').fetchone() == ('Tea Dance 25 Feb 2024', 'social')
    assert db.conn.execute('SELECT name, activity_type, difficulty, date FROM activity').fetchone() == (
        'Social',
        'social',
        'social',
        '2024-02-25',
    )


def test_social_register_parse_reads_both_halves_with_ticket_and_absence(db):
    """Both side-by-side halves are read; Concession maps to ticket type; ❎ is a no-show."""
    parsers.SocialRegisterParser().parse(
        _social_register_ws(), db, term='Tea Dance 25 Feb 2024', year=None, ingest_id=None
    )
    rows = dict(db.conn.execute('SELECT dancer_id, status FROM attendance').fetchall())
    # DNC-3 (right half) read; DNC-5 absent
    assert rows == {'DNC-1': 'attended', 'DNC-2': 'attended', 'DNC-3': 'attended', 'DNC-5': 'absent'}
    # Concession Yes -> member_or_concession (not ordinary); No -> ordinary.
    tickets = dict(db.conn.execute('SELECT dancer_id, ticket_type FROM attendance').fetchall())
    assert tickets == {
        'DNC-1': 'ordinary',
        'DNC-2': 'member_or_concession',
        'DNC-3': 'member_or_concession',
        'DNC-5': 'ordinary',
    }


def test_social_register_parse_duplicate_ticket_becomes_anonymous_extra(db):
    """DNC-1's two present tickets collapse to one named attendee + one anonymous head."""
    parsers.SocialRegisterParser().parse(
        _social_register_ws(), db, term='Tea Dance 25 Feb 2024', year=None, ingest_id=None
    )
    assert db.conn.execute("SELECT status FROM attendance WHERE dancer_id='DNC-1'").fetchall() == [('attended',)]
    assert db.conn.execute('SELECT ticket_type, head_count FROM attendance_count').fetchall() == [(None, 1)]
    # view total: 3 named present (DNC-1/2/3) + 1 anonymous extra = 4
    assert db.conn.execute('SELECT named_total, aggregate_total, total FROM activity_attendance').fetchone() == (
        3,
        1,
        4,
    )


def test_social_register_blank_marker_records_unknown_not_absent(db):
    """When no marker is filled, every booker is UNKNOWN (a held place), not ABSENT; no anon extras."""
    parsers.SocialRegisterParser().parse(
        _blank_party_register_ws(), db, term='End of Term Party 27th June 2024', year=2024, ingest_id=None
    )
    assert db.conn.execute('SELECT name, event_type FROM event').fetchone() == (
        'End of Term Party 27th June 2024',
        'social',
    )
    assert db.conn.execute('SELECT date FROM activity').fetchone() == ('2024-06-27',)
    # DNC-1's duplicate place collapses to one row; both bookers UNKNOWN
    rows = dict(db.conn.execute('SELECT dancer_id, status FROM attendance').fetchall())
    assert rows == {'DNC-1': 'unknown', 'DNC-2': 'unknown'}
    tickets = dict(db.conn.execute('SELECT dancer_id, ticket_type FROM attendance').fetchall())
    assert tickets == {'DNC-1': 'member_or_concession', 'DNC-2': 'ordinary'}
    # an unmarked register makes no anonymous-extra head from the duplicate place
    assert db.conn.execute('SELECT COUNT(*) FROM attendance_count').fetchone() == (0,)


def test_social_register_parse_attended_layout_date_from_tab_name(db):
    """Date comes from the tab name ('26 March 23'); 'x'/blank are present/absent; no ticket type."""
    parsers.SocialRegisterParser().parse(
        _attended_register_ws(), db, term='Sunday Social Attendees 2023-03-25', year=None, ingest_id=None
    )
    # event name has 'Attendees' stripped; activity dated from the tab name, not the filename's 25th
    assert db.conn.execute('SELECT name, event_type FROM event').fetchone() == ('Sunday Social 2023-03-25', 'social')
    assert db.conn.execute('SELECT date FROM activity').fetchone() == ('2023-03-26',)
    rows = dict(db.conn.execute('SELECT dancer_id, status FROM attendance').fetchall())
    assert rows == {'DNC-1': 'attended', 'DNC-2': 'attended', 'DNC-3': 'absent'}  # DNC-3 booked but absent
    # DNC-2's two 'x' tickets -> one named + one anonymous extra; no concession column -> ticket NULL
    assert db.conn.execute('SELECT ticket_type, head_count FROM attendance_count').fetchall() == [(None, 1)]
    assert {r[0] for r in db.conn.execute('SELECT DISTINCT ticket_type FROM attendance')} == {None}


# ---- booking export parsing ----


def test_booking_export_dated_records_unknown_per_session(db):
    """Each dated booking is an UNKNOWN registration; a cancelled one is ABSENT; none attended."""
    parsers.BookingExportParser().parse(
        _booking_by_activity_ws(), db, term='Level 1 Fundamentals Term A 2023-01-17 2255', year=2023, ingest_id=None
    )
    # export tail stripped; the session-month window (these bookings sit in January) disambiguates
    # ESDS's reused 'Term A/B' labels within a year, where the term name carries no month of its own
    assert db.conn.execute('SELECT name, event_type FROM event').fetchone() == (
        'Level 1 Fundamentals Term A (Jan 2023)',
        'course',
    )
    assert sorted(r[0] for r in db.conn.execute('SELECT date FROM activity')) == ['2023-01-19', '2023-01-26']
    rows = db.conn.execute(
        'SELECT a.date, at.dancer_id, at.status FROM attendance at JOIN activity a USING(activity_id) '
        'ORDER BY a.date, at.dancer_id'
    ).fetchall()
    assert rows == [
        ('2023-01-19', 'DNC-1', 'unknown'),
        ('2023-01-19', 'DNC-2', 'unknown'),
        ('2023-01-19', 'DNC-3', 'unknown'),  # the two session-1 tickets collapse to one row
        ('2023-01-26', 'DNC-1', 'unknown'),
        ('2023-01-26', 'DNC-2', 'absent'),  # cancelled booking
    ]
    assert db.conn.execute("SELECT COUNT(*) FROM attendance WHERE status='attended'").fetchone() == (0,)


def test_booking_export_reads_concession_from_sibling_attendees_tab(db):
    """The older booking export also recovers ticket type from the sibling 'Attendees' rollup.

    Its dated 'Attendees By Activity' tab carries no member-rate flag, so the parser reaches across
    to the workbook's 'Attendees' rollup (the shared ``_concession_by_dancer`` path): DNC-1 ('No')
    is ordinary, DNC-2 ('Yes') member/concession, DNC-3 has no rollup row so its ticket type is NULL.
    """
    parsers.BookingExportParser().parse(
        _booking_with_concession_ws(),
        db,
        term='Level 1 Fundamentals Term A 2023-01-17 2255',
        year=2023,
        ingest_id=None,
    )
    tickets = dict(db.conn.execute('SELECT dancer_id, ticket_type FROM attendance').fetchall())
    assert tickets == {'DNC-1': 'ordinary', 'DNC-2': 'member_or_concession', 'DNC-3': None}


def test_booking_export_same_term_name_splits_by_session_months(db):
    """One reused 'Term B' label across non-overlapping months becomes distinct events.

    Two booking exports share a term name but sit in different months; the session-month window
    keeps them apart. A re-export of one of them (same month) must collapse back, not split.
    """
    p = parsers.BookingExportParser()
    p.parse(_booking_ws_on(1, 19), db, term='Level 1 Term B', year=2023, ingest_id=None)  # January
    p.parse(_booking_ws_on(9, 6), db, term='Level 1 Term B', year=2023, ingest_id=None)  # September
    p.parse(_booking_ws_on(9, 13), db, term='Level 1 Term B', year=2023, ingest_id=None)  # re-export -> merges

    names = sorted(r[0] for r in db.conn.execute('SELECT name FROM event'))
    assert names == ['Level 1 Term B (Jan 2023)', 'Level 1 Term B (Sep 2023)']


def test_booking_export_single_event_is_mostly_unknown(db):
    """The 2023 Christmas Party shape: a present-note column, date from the filename."""
    parsers.BookingExportParser().parse(
        _booking_list_ws(), db, term='Christmas Party with Ian Ewing 14 Dec 2023', year=2023, ingest_id=None
    )
    assert db.conn.execute('SELECT name, event_type FROM event').fetchone() == (
        'Christmas Party with Ian Ewing (2023)',
        'social',
    )
    assert db.conn.execute('SELECT date FROM activity').fetchone() == ('2023-12-14',)
    rows = dict(db.conn.execute('SELECT dancer_id, status FROM attendance').fetchall())
    assert rows == {
        'DNC-1': 'unknown',  # confirmed booking, no note
        'DNC-2': 'unknown',
        'DNC-3': 'absent',  # 'no show'
        'DNC-4': 'attended',  # 'paid cash on door'
        'DNC-5': 'absent',  # 'refunded'
    }
    # interest is recorded without inflating attendance: 5 registered, 2 unknown, 1 attended
    assert db.conn.execute(
        'SELECT named_total, named_unknown, named_registered FROM activity_attendance'
    ).fetchone() == (1, 2, 5)


# ---- modern dancecloud (Checked In) export parsing ----


def test_dancecloud_activity_splits_paired_ticket_and_reads_check_ins(db):
    """A combined workshop+tea-dance ticket splits into a lesson and a social; check-ins set turnout."""
    parsers.DancecloudActivityParser().parse(
        _checked_in_by_activity_ws(),
        db,
        term='Workshop and Tea Dance Attendees May 10th 2026',
        year=2026,
        ingest_id=None,
    )
    # the event is named from the filename (date tail stripped, year suffixed), classified a workshop
    assert db.conn.execute('SELECT name, event_type FROM event').fetchone() == (
        'Workshop and Tea Dance (2026)',
        'workshop',
    )
    # one ticket, two activities on the same date: a lesson (the workshop) and a social (the tea dance)
    acts = db.conn.execute('SELECT name, activity_type, difficulty, date FROM activity ORDER BY name').fetchall()
    assert acts == [
        ('Collegiate Shag Workshop', 'lesson', None, '2026-05-10'),
        ('Tea Dance', 'social', 'social', '2026-05-10'),
    ]
    rows = db.conn.execute(
        'SELECT a.name, at.dancer_id, at.status FROM attendance at JOIN activity a USING(activity_id) '
        'ORDER BY a.name, at.dancer_id'
    ).fetchall()
    assert rows == [
        ('Collegiate Shag Workshop', 'DNC-1', 'attended'),  # scanned in
        ('Collegiate Shag Workshop', 'DNC-2', 'absent'),  # confirmed, no scan
        ('Collegiate Shag Workshop', 'DNC-3', 'absent'),  # cancelled
        ('Tea Dance', 'DNC-1', 'absent'),
        ('Tea Dance', 'DNC-2', 'attended'),
    ]
    # DNC-1's two scanned-in workshop tickets -> one named row + one anonymous extra head
    assert db.conn.execute(
        'SELECT a.name, ac.head_count FROM attendance_count ac JOIN activity a USING(activity_id)'
    ).fetchall() == [('Collegiate Shag Workshop', 1)]


def test_dancecloud_activity_unknown_when_scanner_never_ran(db):
    """An export with an entirely blank Checked In column records UNKNOWN, not a sheet-wide no-show."""
    parsers.DancecloudActivityParser().parse(
        _checked_in_no_scans_ws(), db, term="Teachers' Workshop Attendees Jan 17th 2026", year=2026, ingest_id=None
    )
    assert db.conn.execute('SELECT name, event_type FROM event').fetchone() == ("Teachers' Workshop (2026)", 'workshop')
    rows = dict(db.conn.execute('SELECT dancer_id, status FROM attendance').fetchall())
    assert rows == {'DNC-1': 'unknown', 'DNC-2': 'unknown'}  # DNC-1's two tickets collapse to one row
    # no check-in anywhere -> no attended rows, so no anonymous extra heads from the duplicate ticket
    assert db.conn.execute('SELECT COUNT(*) FROM attendance_count').fetchone() == (0,)


def test_dancecloud_activity_propagates_class_checkin_across_its_days(db):
    """A check-in to any day of a multi-day class implies attendance every day; socials stay per-night."""
    parsers.DancecloudActivityParser().parse(
        _checked_in_multiday_class_ws(),
        db,
        term='Big Weekender Attendees March 21st-22nd 2026',
        year=2026,
        ingest_id=None,
    )
    rows = db.conn.execute(
        'SELECT a.name, a.date, at.dancer_id, at.status FROM attendance at JOIN activity a USING(activity_id) '
        'ORDER BY a.name, a.date, at.dancer_id'
    ).fetchall()
    assert rows == [
        ('Evening Social', '2026-03-21', 'DNC-1', 'attended'),  # social scanned
        ('Evening Social', '2026-03-22', 'DNC-1', 'absent'),  # social NOT propagated -> stays absent
        ('Track A - Classes', '2026-03-21', 'DNC-1', 'attended'),  # scanned day 1
        ('Track A - Classes', '2026-03-21', 'DNC-2', 'attended'),  # lifted from day-2 scan
        ('Track A - Classes', '2026-03-21', 'DNC-3', 'absent'),  # never scanned the class
        ('Track A - Classes', '2026-03-22', 'DNC-1', 'attended'),  # lifted from day-1 scan
        ('Track A - Classes', '2026-03-22', 'DNC-2', 'attended'),  # scanned day 2
        ('Track A - Classes', '2026-03-22', 'DNC-3', 'absent'),
    ]


def test_dancecloud_activity_reads_concession_from_sibling_attendees_tab(db):
    """Ticket type is recovered from the sibling 'Attendees' rollup the by-activity tab lacks.

    The 'Attendees By Activity' tab carries no member-rate flag, so the parser reaches across to
    the workbook's 'Attendees' rollup: DNC-1 ('Yes') is member/concession, DNC-2 ('No') ordinary,
    DNC-3 has no rollup row so its ticket type stays unknown (NULL).
    """
    parsers.DancecloudActivityParser().parse(
        _checked_in_with_concession_ws(),
        db,
        term='End of Term Party with Iain Ewing and the Chevaliers 2026-06-27 1153',
        year=2026,
        ingest_id=None,
    )
    tickets = dict(db.conn.execute('SELECT dancer_id, ticket_type FROM attendance').fetchall())
    assert tickets == {'DNC-1': 'member_or_concession', 'DNC-2': 'ordinary', 'DNC-3': None}


# ---- Stockbridge Swingout (event-specific) parsing ----


def test_stockbridge_matches_only_its_own_tab():
    """The Stockbridge parser claims a 'swingout' tab with Ticket Type + Registered, nothing else.

    It must be tried before the roster parser (which also matches the dated Registered column);
    PARSERS order guarantees it wins the dispatch.
    """
    assert parsers.StockbridgeSwingoutParser().matches(_swingout_ws())
    assert not parsers.StockbridgeSwingoutParser().matches(_roster_ws())  # no swingout title / Ticket Type
    order = [type(p) for p in parsers.PARSERS]
    assert order.index(parsers.StockbridgeSwingoutParser) < order.index(parsers.RosterParser)


def test_stockbridge_expands_each_ticket_into_its_weekend_activities(db):
    """Each ticket grants a fixed all-or-none set of activities, with Registered as the status."""
    parsers.StockbridgeSwingoutParser().parse(_swingout_ws(), db, term='Sept-Oct 2025', year=2025, ingest_id=None)
    assert db.conn.execute('SELECT name, event_type FROM event').fetchone() == (
        'Sept-Oct 2025: Stockbridge Swingout',
        'weekender',
    )
    # a Full Pass track is two lessons, one each day; a Saturday Only ticket grants only the Saturday
    # one. The Friday social is the day before; the Sunday class is the day after.
    acts = db.conn.execute('SELECT name, activity_type, difficulty, date FROM activity ORDER BY date, name').fetchall()
    assert acts == [
        ('Friday Social', 'social', 'social', '2025-10-17'),
        ('Improvers / Intermediate Classes', 'lesson', 'Improvers / Intermediate', '2025-10-18'),
        ('Intermediate / Advanced Classes', 'lesson', 'Intermediate / Advanced', '2025-10-18'),
        ('Saturday Social', 'social', 'social', '2025-10-18'),
        ('Improvers / Intermediate Classes', 'lesson', 'Improvers / Intermediate', '2025-10-19'),
    ]
    rows = db.conn.execute(
        'SELECT a.name, a.date, at.dancer_id, at.status FROM attendance at JOIN activity a USING(activity_id) '
        'ORDER BY a.name, a.date, at.dancer_id'
    ).fetchall()
    assert rows == [
        ('Friday Social', '2025-10-17', 'DNC-1', 'attended'),  # via Full Pass
        ('Friday Social', '2025-10-17', 'DNC-4', 'absent'),  # Friday Social Only, no-show
        # DNC-1's Full Pass puts them in the Imp/Int track on both Saturday and Sunday
        ('Improvers / Intermediate Classes', '2025-10-18', 'DNC-1', 'attended'),
        ('Improvers / Intermediate Classes', '2025-10-19', 'DNC-1', 'attended'),
        ('Intermediate / Advanced Classes', '2025-10-18', 'DNC-2', 'attended'),  # Saturday Only, no Sunday row
        ('Saturday Social', '2025-10-18', 'DNC-1', 'attended'),
        ('Saturday Social', '2025-10-18', 'DNC-2', 'attended'),
        ('Saturday Social', '2025-10-18', 'DNC-3', 'attended'),
    ]
    # DNC-1 and DNC-3 each hold a second Saturday-social-granting ticket, both scanned -> two extra heads
    assert db.conn.execute(
        'SELECT a.name, ac.head_count FROM attendance_count ac JOIN activity a USING(activity_id)'
    ).fetchall() == [('Saturday Social', 2)]
