import openpyxl
import pytest

from esds_apps.attendance import pseudonymise, pseudonyms_db

PASSPHRASE = 'test-passphrase-esds'


# ============================================================================
# Fixtures
# ============================================================================


@pytest.fixture
def simple_xlsx(tmp_path):
    """Single-sheet xlsx with standard headers."""
    p = tmp_path / 'simple.xlsx'
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Sheet1'
    ws.append(['First Name', 'Last Name', 'Email', 'Notes'])
    ws.append(['Alice', 'Smith', 'alice@example.com', 'Regular'])
    ws.append(['Bob', 'Jones', 'bob@example.com', 'VIP'])
    wb.save(p)
    return p


@pytest.fixture
def prefixed_xlsx(tmp_path):
    """Xlsx with two metadata rows above the header."""
    p = tmp_path / 'prefixed.xlsx'
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Sheet1'
    ws.append(['Class Register'])
    ws.append(['Level 1'])
    ws.append(['Forename', 'Surname', 'E-mail', 'Attended'])
    ws.append(['Victoria', 'McLean', 'v@example.com', 'Yes'])
    wb.save(p)
    return p


# ============================================================================
# Column canonicalisation
# ============================================================================


def test_canonical_name_key_first_name_variants():
    assert pseudonymise._canonical_name_key('First Name') == 'first_name'
    assert pseudonymise._canonical_name_key('first_name') == 'first_name'
    assert pseudonymise._canonical_name_key('Forename') == 'first_name'


def test_canonical_name_key_last_name_variants():
    assert pseudonymise._canonical_name_key('Last Name') == 'last_name'
    assert pseudonymise._canonical_name_key('Surname') == 'last_name'
    assert pseudonymise._canonical_name_key('LAST NAME') == 'last_name'


def test_canonical_name_key_generic_name_maps_to_first():
    assert pseudonymise._canonical_name_key('Name') == 'first_name'
    assert pseudonymise._canonical_name_key('Full Name') == 'first_name'
    assert pseudonymise._canonical_name_key('Member') == 'first_name'


def test_canonical_name_key_unknown_raises():
    with pytest.raises(ValueError, match='Cannot map column'):
        pseudonymise._canonical_name_key('Notes')


# ============================================================================
# Name validation regex
# ============================================================================


def test_valid_name_accepts_typical_names():
    assert pseudonymise._VALID_NAME_RE.match('Alice')
    assert pseudonymise._VALID_NAME_RE.match('Alice Smith')
    assert pseudonymise._VALID_NAME_RE.match("O'Brien")
    assert pseudonymise._VALID_NAME_RE.match('Mary-Jane')
    assert pseudonymise._VALID_NAME_RE.match('José García')


def test_valid_name_rejects_footer_text():
    assert not pseudonymise._VALID_NAME_RE.match('This list is correct as of 01/01/2024')
    assert not pseudonymise._VALID_NAME_RE.match('Total: 42')
    assert not pseudonymise._VALID_NAME_RE.match('alice@example.com')


def test_valid_name_rejects_digits():
    assert not pseudonymise._VALID_NAME_RE.match('Alice2')


# ============================================================================
# XLSX sheet reading
# ============================================================================


def test_read_sheet_simple(simple_xlsx):
    wb = openpyxl.load_workbook(simple_xlsx)
    fieldnames, rows, prefix = pseudonymise._read_sheet(wb.active)
    assert fieldnames == ['First Name', 'Last Name', 'Email', 'Notes']
    assert len(rows) == 2
    assert rows[0]['First Name'] == 'Alice'
    assert rows[1]['Last Name'] == 'Jones'
    assert prefix == []


def test_read_sheet_detects_prefix(prefixed_xlsx):
    wb = openpyxl.load_workbook(prefixed_xlsx)
    fieldnames, rows, prefix = pseudonymise._read_sheet(wb.active)
    assert fieldnames == ['Forename', 'Surname', 'E-mail', 'Attended']
    assert len(rows) == 1
    assert len(prefix) == 2
    assert prefix[0][0] == 'Class Register'


def test_read_sheet_skips_empty_rows(tmp_path):
    p = tmp_path / 'sparse.xlsx'
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(['First Name', 'Last Name'])
    ws.append(['Alice', 'Smith'])
    ws.append([None, None])
    ws.append(['Bob', 'Jones'])
    wb.save(p)
    wb2 = openpyxl.load_workbook(p)
    _, rows, _ = pseudonymise._read_sheet(wb2.active)
    assert len(rows) == 2


def test_read_sheet_pads_short_rows(tmp_path):
    p = tmp_path / 'short.xlsx'
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(['First Name', 'Last Name', 'Email'])
    ws.append(['Alice', 'Smith'])  # missing email cell
    wb.save(p)
    wb2 = openpyxl.load_workbook(p)
    _, rows, _ = pseudonymise._read_sheet(wb2.active)
    assert rows[0]['Email'] == ''


def test_read_sheet_note_row_not_treated_as_header(tmp_path):
    """A row containing 'First Name' as a substring should not be picked as header."""
    p = tmp_path / 'note.xlsx'
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(['NB: order A_Z by First Name'])
    ws.append(['Forename', 'Surname'])
    ws.append(['Alice', 'Smith'])
    wb.save(p)
    wb2 = openpyxl.load_workbook(p)
    fieldnames, rows, prefix = pseudonymise._read_sheet(wb2.active)
    assert fieldnames == ['Forename', 'Surname']
    assert len(rows) == 1
    assert len(prefix) == 1


# ============================================================================
# XLSX sheet writing
# ============================================================================


def test_write_sheet_round_trip(tmp_path, simple_xlsx):
    wb = openpyxl.load_workbook(simple_xlsx)
    fieldnames, rows, prefix = pseudonymise._read_sheet(wb.active)
    wb_out = openpyxl.Workbook()
    pseudonymise._write_sheet(wb_out.active, fieldnames, rows, fieldnames, prefix)
    out = tmp_path / 'out.xlsx'
    wb_out.save(out)
    wb_check = openpyxl.load_workbook(out)
    data = list(wb_check.active.iter_rows(values_only=True))
    assert list(data[0]) == fieldnames
    assert data[1][0] == 'Alice'


def test_write_sheet_preserves_prefix(tmp_path, prefixed_xlsx):
    wb = openpyxl.load_workbook(prefixed_xlsx)
    fieldnames, rows, prefix = pseudonymise._read_sheet(wb.active)
    wb_out = openpyxl.Workbook()
    pseudonymise._write_sheet(wb_out.active, fieldnames, rows, fieldnames, prefix)
    out = tmp_path / 'out.xlsx'
    wb_out.save(out)
    wb_check = openpyxl.load_workbook(out)
    data = list(wb_check.active.iter_rows(values_only=True))
    assert data[0][0] == 'Class Register'
    assert list(data[2]) == fieldnames


def test_write_sheet_renamed_headers(tmp_path, simple_xlsx):
    wb = openpyxl.load_workbook(simple_xlsx)
    fieldnames, rows, prefix = pseudonymise._read_sheet(wb.active)
    out_fieldnames = ['dancer_id', 'redacted', 'redacted', 'Notes']
    wb_out = openpyxl.Workbook()
    pseudonymise._write_sheet(wb_out.active, out_fieldnames, rows, fieldnames, prefix)
    out = tmp_path / 'out.xlsx'
    wb_out.save(out)
    wb_check = openpyxl.load_workbook(out)
    header = [c.value for c in next(wb_check.active.iter_rows())]
    assert header == out_fieldnames


# ============================================================================
# Column detection
# ============================================================================


def test_detect_columns_by_header(simple_xlsx):
    wb = openpyxl.load_workbook(simple_xlsx)
    fieldnames, rows, _ = pseudonymise._read_sheet(wb.active)
    detected = pseudonymise.detect_columns(fieldnames, rows)
    assert 'First Name' in detected['name_cols']
    assert 'Last Name' in detected['name_cols']
    assert 'Email' in detected['email_cols']
    assert 'Notes' not in detected['name_cols']
    assert 'Notes' not in detected['email_cols']


def test_detect_columns_email_by_content(tmp_path):
    p = tmp_path / 'sniff.xlsx'
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(['First Name', 'Last Name', 'Contact'])
    ws.append(['Alice', 'Smith', 'alice@example.com'])
    ws.append(['Bob', 'Jones', 'bob@example.com'])
    wb.save(p)
    wb2 = openpyxl.load_workbook(p)
    fieldnames, rows, _ = pseudonymise._read_sheet(wb2.active)
    detected = pseudonymise.detect_columns(fieldnames, rows)
    assert 'Contact' in detected['email_cols']


def test_detect_columns_member_exact_only(tmp_path):
    """'Member' alone is a valid name column; adjacent words disqualify it."""
    p = tmp_path / 't.xlsx'
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(['Member', 'Concession / Member?', 'Paid as member', 'Paid as non-member', 'Member 2022'])
    ws.append(['Alice Smith', 'No', 'Yes', 'No', 'Yes'])
    wb.save(p)
    wb2 = openpyxl.load_workbook(p)
    fieldnames, rows, _ = pseudonymise._read_sheet(wb2.active)
    detected = pseudonymise.detect_columns(fieldnames, rows)
    assert detected['name_cols'] == ['Member']


def test_detect_columns_email_false_positives(tmp_path):
    """Columns where 'email' is a modifier word, not the column type."""
    p = tmp_path / 't.xlsx'
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(['Email', 'Email subscriber status', 'email present in Wix Export?'])
    ws.append(['a@b.com', 'subscribed', 'Yes'])
    wb.save(p)
    wb2 = openpyxl.load_workbook(p)
    fieldnames, rows, _ = pseudonymise._read_sheet(wb2.active)
    detected = pseudonymise.detect_columns(fieldnames, rows)
    assert detected['email_cols'] == ['Email']


def test_detect_columns_content_sniff_max_5_rows(tmp_path):
    """Emails appearing only after row 5 should not trigger content detection."""
    p = tmp_path / 'late.xlsx'
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(['First Name', 'Last Name', 'Contact'])
    for _ in range(5):
        ws.append(['Alice', 'Smith', 'not-an-email'])
    ws.append(['Bob', 'Jones', 'bob@example.com'])
    wb.save(p)
    wb2 = openpyxl.load_workbook(p)
    fieldnames, rows, _ = pseudonymise._read_sheet(wb2.active)
    detected = pseudonymise.detect_columns(fieldnames, rows)
    assert 'Contact' not in detected['email_cols']


# ============================================================================
# Full pipeline — single file
# ============================================================================


def test_pseudonymise_replaces_pii(simple_xlsx, tmp_path):
    out = tmp_path / 'out.xlsx'
    result = pseudonymise.pseudonymise(simple_xlsx, tmp_path / 'db.sqlite', PASSPHRASE, output_path=out)
    rows = result['Sheet1']
    assert rows[0]['First Name'].startswith('DNC-')
    assert rows[0]['Last Name'] == 'redacted'
    assert rows[0]['Email'] == 'redacted'
    assert rows[0]['Notes'] == 'Regular'


def test_pseudonymise_side_by_side_record_blocks(tmp_path):
    """A register printed as two halves side by side: each half is a different person per row.

    Regression: duplicate headers first collapsed onto one dict key (clobbering the
    dancer_id), and even once separated only the first half got an id. Each repeated block
    must get its own dancer_id, since the two halves are two halves of one attendee list.
    """
    p = tmp_path / 'teadance.xlsx'
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Sheet1'
    ws.append(['#', 'First Name', 'Surname', 'Concession', 'Present?', '#', 'First Name', 'Surname', 'Concession'])
    ws.append([39, 'Alice', 'Smith', 'Yes', 'x', 60, 'Carol', 'Reid', 'No'])
    ws.append([40, 'Bob', 'Jones', 'No', 'x', 61, 'Dave', 'Watt', 'Yes'])
    wb.save(p)

    out = tmp_path / 'out.xlsx'
    pseudonymise.pseudonymise(p, tmp_path / 'db.sqlite', PASSPHRASE, output_path=out)
    data = list(openpyxl.load_workbook(out).active.iter_rows(values_only=True))

    # Both halves' first-name columns become 'dancer_id'; surnames 'redacted'; non-PII kept.
    assert data[0] == (
        '#',
        'dancer_id',
        'redacted',
        'Concession',
        'Present?',
        '#',
        'dancer_id',
        'redacted',
        'Concession',
    )
    left, right = data[1][1], data[1][6]
    assert left.startswith('DNC-') and right.startswith('DNC-')
    assert left != right  # different people in each half → different ids
    assert data[1][0] == '39' and data[1][5] == '60'  # both halves' data survive (cells stringified)
    assert data[1][4] == 'x'  # Present? marker preserved


def test_pseudonymise_writes_output_file(simple_xlsx, tmp_path):
    out = tmp_path / 'out.xlsx'
    pseudonymise.pseudonymise(simple_xlsx, tmp_path / 'db.sqlite', PASSPHRASE, output_path=out)
    assert out.exists()
    wb = openpyxl.load_workbook(out)
    header = [c.value for c in next(wb.active.iter_rows())]
    assert header[0] == 'dancer_id'


def test_pseudonymise_auto_output_name(simple_xlsx, tmp_path):
    pseudonymise.pseudonymise(simple_xlsx, tmp_path / 'db.sqlite', PASSPHRASE)
    assert (simple_xlsx.parent / 'simple_pseudonymised.xlsx').exists()


def test_pseudonymise_preserves_prefix(prefixed_xlsx, tmp_path):
    out = tmp_path / 'out.xlsx'
    pseudonymise.pseudonymise(prefixed_xlsx, tmp_path / 'db.sqlite', PASSPHRASE, output_path=out)
    wb = openpyxl.load_workbook(out)
    data = list(wb.active.iter_rows(values_only=True))
    assert data[0][0] == 'Class Register'
    assert data[1][0] == 'Level 1'


def test_pseudonymise_stable_ids_across_runs(simple_xlsx, tmp_path):
    db = tmp_path / 'db.sqlite'
    result1 = pseudonymise.pseudonymise(simple_xlsx, db, PASSPHRASE, output_path=tmp_path / 'out1.xlsx')
    result2 = pseudonymise.pseudonymise(simple_xlsx, db, PASSPHRASE, output_path=tmp_path / 'out2.xlsx')
    assert result1['Sheet1'][0]['First Name'] == result2['Sheet1'][0]['First Name']


def test_pseudonymise_skips_footer_text(tmp_path):
    p = tmp_path / 'footer.xlsx'
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Sheet1'
    ws.append(['First Name', 'Last Name', 'Email'])
    ws.append(['Alice', 'Smith', 'alice@example.com'])
    ws.append(['This list is correct as of 01/01/2024', '', ''])
    wb.save(p)
    db = tmp_path / 'db.sqlite'
    result = pseudonymise.pseudonymise(p, db, PASSPHRASE, output_path=tmp_path / 'out.xlsx')
    rows = result['Sheet1']
    assert rows[0]['First Name'].startswith('DNC-')
    assert rows[1]['First Name'] == 'This list is correct as of 01/01/2024'
    c = pseudonyms_db.open_db(db, PASSPHRASE)
    assert len(pseudonyms_db.decrypt_all(c)) == 1
    c.conn.close()


def test_pseudonymise_skips_summary_label_rows(tmp_path):
    """Aggregate rows whose name cell holds a label ('Totals', 'Social Only') are not minted as dancers.

    These labels are pure letters, so _VALID_NAME_RE accepts them; the _NON_NAME_LABELS
    blocklist is what keeps the L2 register's footer rows from becoming spurious dancers.
    """
    p = tmp_path / 'summary.xlsx'
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Sheet1'
    ws.append(['First Name', 'Last Name', 'Email'])
    ws.append(['Alice', 'Smith', 'alice@example.com'])
    ws.append(['', 'Totals', ''])
    ws.append(['', 'All Attendance', ''])
    ws.append(['', 'Social Only', ''])
    wb.save(p)
    db = tmp_path / 'db.sqlite'
    result = pseudonymise.pseudonymise(p, db, PASSPHRASE, output_path=tmp_path / 'out.xlsx')
    rows = result['Sheet1']
    assert rows[0]['First Name'].startswith('DNC-')
    for label_row in rows[1:]:
        assert not label_row['First Name'].startswith('DNC-')
        assert label_row['Last Name'] != 'redacted'
    c = pseudonyms_db.open_db(db, PASSPHRASE)
    assert len(pseudonyms_db.decrypt_all(c)) == 1
    c.conn.close()


def test_pseudonymise_manual_column_override(tmp_path):
    p = tmp_path / 'nonstandard.xlsx'
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Sheet1'
    ws.append(['Forename', 'Surname', 'Contact', 'Notes'])
    ws.append(['Alice', 'Smith', 'alice@example.com', 'Active'])
    wb.save(p)
    result = pseudonymise.pseudonymise(
        p,
        tmp_path / 'db.sqlite',
        PASSPHRASE,
        output_path=tmp_path / 'out.xlsx',
        name_cols=['Forename', 'Surname'],
        email_cols=['Contact'],
    )
    rows = result['Sheet1']
    assert rows[0]['Forename'].startswith('DNC-')
    assert rows[0]['Surname'] == 'redacted'
    assert rows[0]['Contact'] == 'redacted'


def test_pseudonymise_multi_sheet(tmp_path):
    p = tmp_path / 'multi.xlsx'
    wb = openpyxl.Workbook()
    ws1 = wb.active
    ws1.title = 'Jan'
    ws1.append(['First Name', 'Last Name'])
    ws1.append(['Alice', 'Smith'])
    ws2 = wb.create_sheet('Feb')
    ws2.append(['First Name', 'Last Name'])
    ws2.append(['Bob', 'Jones'])
    wb.save(p)
    result = pseudonymise.pseudonymise(p, tmp_path / 'db.sqlite', PASSPHRASE, output_path=tmp_path / 'out.xlsx')
    assert 'Jan' in result
    assert 'Feb' in result
    assert result['Jan'][0]['First Name'].startswith('DNC-')
    assert result['Feb'][0]['First Name'].startswith('DNC-')


def test_pseudonymise_copies_non_pii_sheet_verbatim(tmp_path):
    p = tmp_path / 'tally.xlsx'
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Tally'
    ws.append(['Week 1', 'Members', 'Tot:', 19])
    ws.append([None, True, True, False])
    ws.append([None, False, True, True])
    wb.save(p)

    """A sheet with no name/email columns (e.g. an aggregate tally grid) must be
    copied through cell-for-cell, not rebuilt — rebuilding collapsed such sheets.
    """

    out = tmp_path / 'out.xlsx'
    result = pseudonymise.pseudonymise(p, tmp_path / 'db.sqlite', PASSPHRASE, output_path=out)
    assert result['Tally'] == []  # nothing redacted

    data = list(openpyxl.load_workbook(out)['Tally'].iter_rows(values_only=True))
    assert data[0] == ('Week 1', 'Members', 'Tot:', 19)
    assert data[1] == (None, True, True, False)
    assert data[2] == (None, False, True, True)


def test_pseudonymise_mixed_pii_and_tally_sheets(tmp_path):
    p = tmp_path / 'mixed.xlsx'
    wb = openpyxl.Workbook()
    roster = wb.active
    roster.title = 'Roster'
    roster.append(['First Name', 'Last Name'])
    roster.append(['Alice', 'Smith'])
    tally = wb.create_sheet('Tally')
    tally.append(['Members', 'Tot:', 7])
    tally.append([True, True, False])
    wb.save(p)

    """A roster sheet is pseudonymised while a tally sheet in the same workbook
    survives intact.
    """

    out = tmp_path / 'out.xlsx'
    result = pseudonymise.pseudonymise(p, tmp_path / 'db.sqlite', PASSPHRASE, output_path=out)
    assert result['Roster'][0]['First Name'].startswith('DNC-')
    assert result['Tally'] == []

    wb_out = openpyxl.load_workbook(out)
    assert list(wb_out['Tally'].iter_rows(values_only=True)) == [('Members', 'Tot:', 7), (True, True, False)]


# ============================================================================
# Folder processing
# ============================================================================


def test_pseudonymise_folder_basic(tmp_path, simple_xlsx):
    import shutil

    input_dir = tmp_path / 'in'
    input_dir.mkdir()
    shutil.copy(simple_xlsx, input_dir / 'file.xlsx')
    output_dir = tmp_path / 'out'
    pseudonymise.pseudonymise_folder(input_dir, output_dir, tmp_path / 'db.sqlite', PASSPHRASE)
    assert (output_dir / 'file_pseudonymised.xlsx').exists()


def test_pseudonymise_folder_preserves_subdirs(tmp_path, simple_xlsx):
    import shutil

    input_dir = tmp_path / 'in'
    sub = input_dir / 'sub'
    sub.mkdir(parents=True)
    shutil.copy(simple_xlsx, input_dir / 'top.xlsx')
    shutil.copy(simple_xlsx, sub / 'nested.xlsx')
    output_dir = tmp_path / 'out'
    pseudonymise.pseudonymise_folder(input_dir, output_dir, tmp_path / 'db.sqlite', PASSPHRASE)
    assert (output_dir / 'top_pseudonymised.xlsx').exists()
    assert (output_dir / 'sub' / 'nested_pseudonymised.xlsx').exists()


def test_pseudonymise_folder_shared_db(tmp_path, simple_xlsx):
    """Same person appearing in two files gets the same dancer ID."""
    import shutil

    input_dir = tmp_path / 'in'
    input_dir.mkdir()
    shutil.copy(simple_xlsx, input_dir / 'a.xlsx')
    shutil.copy(simple_xlsx, input_dir / 'b.xlsx')
    db = tmp_path / 'db.sqlite'
    output_dir = tmp_path / 'out'
    pseudonymise.pseudonymise_folder(input_dir, output_dir, db, PASSPHRASE)
    # Only the unique people from simple_xlsx (Alice, Bob) should be in the DB.
    c = pseudonyms_db.open_db(db, PASSPHRASE)
    assert len(pseudonyms_db.decrypt_all(c)) == 2
    c.conn.close()
