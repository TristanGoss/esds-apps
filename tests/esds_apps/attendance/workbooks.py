"""Shared worksheet builders for the ingest / parser tests.

Each ``_*_ws`` returns a single openpyxl worksheet mimicking one real-world layout, so the
parser tests (``test_parsers``) and the folder-dispatch tests (``test_ingest``) can build the
same fixtures without duplicating them.
"""

from datetime import datetime

import openpyxl


def _roster_ws(title='Level 1 Attendance'):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = title
    ws['E1'], ws['F1'] = datetime(2025, 5, 22), datetime(2025, 5, 29)  # dates above the header
    ws['B2'], ws['C2'], ws['D2'], ws['E2'], ws['F2'] = '#', 'dancer_id', 'redacted', 'Week 1', 'Week 2'
    # (dancer, week1, week2) — mixed truthy markers and a duplicate ticket for DNC-1
    data = [
        ('DNC-1', True, False),
        ('DNC-2', '☑️', 'False'),
        ('DNC-3', 'x', 'Refunded'),
        ('DNC-1', True, True),  # second ticket DNC-1 bought but didn't rename
    ]
    for i, (did, w1, w2) in enumerate(data, start=3):
        ws[f'C{i}'], ws[f'E{i}'], ws[f'F{i}'] = did, w1, w2
    return ws


def _tally_ws(title='Level 2 & Social Only Attendanc'):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = title
    ws['C1'] = 'Level 2 Classes'
    ws['C2'], ws['F2'] = 'Members', '=COUNTIF(C3:E4, "TRUE")'
    ws['H2'], ws['K2'] = 'Concessions', '=COUNTIF(H3:I4, "TRUE")'
    ws['B3'] = 'Week 1 (22 May)'
    ws['C3'], ws['D3'], ws['E3'] = True, True, False  # members: 3 ticks
    ws['C4'], ws['D4'], ws['E4'] = True, False, False
    ws['H3'], ws['I3'] = True, False  # concessions: 2 ticks
    ws['H4'], ws['I4'] = False, True
    return ws


def _teachers_choice_tally_ws(title='Teachers Choice 12th Dec'):
    """A one-off tally: Level 1 / Level 2 / Social Only groups, no week labels, date in the tab name.

    Members COUNTIFs only, three groups side by side. Level 1 has 3 ticks, Level 2 has 2, Social
    Only has 1. The single session date comes from the tab name ('12th Dec') plus the workbook year.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = title
    ws['B1'] = "Teacher's Choice Class attendance count (12th Dec 2024)"
    ws['C2'], ws['U2'], ws['AM2'] = 'Level 1 Classes', 'Level 2 Classes', 'Social Only'
    ws['C3'], ws['G3'] = 'Members', '=COUNTIF(C4:G6, "TRUE")'
    ws['U3'], ws['Y3'] = 'Members', '=COUNTIF(U4:Y6, "TRUE")'
    ws['AM3'], ws['AQ3'] = 'Members', '=COUNTIF(AM4:AQ6, "TRUE")'
    for r in (4, 5, 6):
        ws[f'C{r}'] = True  # Level 1: 3 ticks
    ws['U4'], ws['U5'] = True, True  # Level 2: 2 ticks
    ws['AM4'] = True  # Social Only: 1 tick
    return ws


def _count_grid_ws(title='Level 2'):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = title
    ws['B4'], ws['C4'], ws['D4'], ws['E4'] = 'Members', 'Concessions', 'Non-Members', 'Totals'
    ws['A5'], ws['B5'], ws['C5'], ws['D5'], ws['E5'] = 'Week 1\n(11 Apr)', 16, 3, 8, '=SUM(B5:D5)'
    ws['A6'], ws['B6'], ws['C6'], ws['D6'], ws['E6'] = 'Week 2\n(18 Apr)', 23, 1, 7, '=SUM(B6:D6)'
    ws['A8'], ws['B8'] = 'Mean', '=AVERAGE(B5:B6)'  # must be ignored (no week label)
    return ws


def _count_grid_undated_ws(title='Level 2-3'):
    """An older count grid: bare 'Week N' labels with no per-row dates, ticket-type headers.

    Dates must come from the workbook anchor (Week 1 + 7*(N-1)), and the 'Level 2-3' title
    must normalise to plain 'Level 2'.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = title
    ws['B4'], ws['C4'], ws['D4'], ws['E4'] = 'Members', 'Concessions', 'Non-Members', 'Totals'
    ws['A5'], ws['B5'], ws['C5'], ws['D5'], ws['E5'] = 'Week 1', 21, 0, 3, '=SUM(B5:D5)'
    ws['A6'], ws['B6'], ws['C6'], ws['D6'], ws['E6'] = 'Week 2', 27, 3, 6, '=SUM(B6:D6)'
    return ws


def _roster_weekly_ws(title='Level 1'):
    """A roster where only Week 1 carries a date; Weeks 2-3 are bare 'Week N' labels.

    Mimics the older Level 1 tabs whose '=D1+7' week-date formulas were stripped to nothing
    by pseudonymisation, so all but the first week date must be inferred from the anchor.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = title
    ws['D1'] = datetime(2023, 11, 9)  # only Week 1 dated (E1/F1 held '=D1+7', now gone)
    ws['B2'], ws['C2'], ws['D2'], ws['E2'], ws['F2'] = 'dancer_id', 'redacted', 'Week 1', 'Week 2', 'Week 3'
    ws['B3'], ws['D3'], ws['E3'], ws['F3'] = 'DNC-1', '☑️', '☑️', '❎'
    return ws


def _l2_so_ws(title='2026 L2 & SO Attendance'):
    """A 2026-style sheet with dancer_id + dated columns holding ticket categories, not markers.

    The final row is a COUNTIF summary like the real sheets carry: it has no DNC- id, so it
    must never be read as a category (the bug that spawned phantom '=countif(...)' activities).
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = title
    ws['C1'] = datetime(2026, 1, 15)
    ws['D1'] = datetime(2026, 1, 22)
    ws['B2'], ws['C2'], ws['D2'] = 'dancer_id', 'Session 1', 'Session 2'
    ws['B3'], ws['C3'], ws['D3'] = 'DNC-1', 'Level 2 & Social', 'Social-Only'
    ws['B4'], ws['C4'], ws['D4'] = 'DNC-2', 'Absent', 'Level 2 & Social'
    ws['C5'], ws['D5'] = '=countif(C3:C4,"Absent")', '=countif(D3:D4,"Social-Only")'
    return ws


def _l2_so_absent_heavy_ws(title='L2 and SO Attendance'):
    """An L2 & SO sheet whose first rows are mostly 'Absent', as a term of many no-shows looks.

    'Absent' is both a valid L2/SO category and a roster 'no' marker, so the cells RosterParser
    samples here read as attendance markers and tip it over its threshold — the real-world dispatch
    collision. Still genuinely an L2 & SO sheet (some 'Level 2 & Social' / 'Social-Only' below), so a
    correct parse yields a Level 2 lesson and a social with real attendance, not an all-absent roster.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = title
    ws['C1'] = datetime(2026, 5, 21)
    ws['D1'] = datetime(2026, 5, 28)
    ws['B2'], ws['C2'], ws['D2'] = 'dancer_id', 'Session 1', 'Session 2'
    # The rows RosterParser samples (first few) are all 'Absent'.
    ws['B3'], ws['C3'], ws['D3'] = 'DNC-1', 'Absent', 'Absent'
    ws['B4'], ws['C4'], ws['D4'] = 'DNC-2', 'Absent', 'Absent'
    ws['B5'], ws['C5'], ws['D5'] = 'DNC-3', 'Absent', 'Level 2 & Social'
    # Real attendance further down, so a correct parse has a non-empty Level 2 lesson and a social.
    ws['B6'], ws['C6'], ws['D6'] = 'DNC-4', 'Level 2 & Social', 'Social-Only'
    return ws


def _l1_attendance_2026_ws(title='2026 L1 Attendance'):
    """A 2026 L1 Attendance sheet: dates in the header row, session names optional."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = title
    ws['C1'] = datetime(2026, 1, 15)
    ws['D1'] = datetime(2026, 1, 22)
    ws['B2'], ws['C2'], ws['D2'] = 'dancer_id', 'Week 1', 'Week 2'
    ws['B3'], ws['C3'], ws['D3'] = 'DNC-1', True, False
    ws['B4'], ws['C4'], ws['D4'] = 'DNC-2', 'Yes', True
    ws['B5'], ws['C5'], ws['D5'] = 'DNC-1', True, 'x'  # duplicate ticket
    return ws


def _class_list_roster_ws(title='Class List'):
    """An early-2022 'Class List' roster: 'Week N (DD Mon)' headers, no date cells, a leading # column.

    The session dates live only inside the week labels and the year only in the filename, so the
    columns can only be dated once parse is given the year. A 'Lead / Follow' column sits between
    the redacted name and the first week and must not be read as a session.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = title
    ws['B1'] = 'Attendee List: Level 2'
    ws['B3'], ws['C3'], ws['D3'] = 'dancer_id', 'redacted', 'Lead / Follow'
    ws['E3'], ws['F3'], ws['G3'] = 'Week 1 (10 Feb)', 'Week 2 (17 Feb)', 'Week 3 (24 Feb)'
    ws['A4'], ws['B4'], ws['D4'], ws['E4'], ws['F4'], ws['G4'] = 1, 'DNC-1', 'Lead', True, False, True
    ws['A5'], ws['B5'], ws['D5'], ws['E5'], ws['F5'], ws['G5'] = 2, 'DNC-2', 'Follow', True, True, 'False'
    return ws


def _roster_concession_ws(title='Level 1 Attendance'):
    """A roster carrying a per-dancer 'Concession' flag, captured as a ticket type.

    DNC-1 is 'No' in its first row but blank in a duplicate row -> ordinary (the blank must not
    overwrite the known value). DNC-4 has an unrecognised value -> unknown. DNC-9 is blank
    throughout -> NULL (no information).
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = title
    ws['E1'], ws['F1'] = datetime(2025, 5, 22), datetime(2025, 5, 29)  # dates above the header
    ws['B2'], ws['C2'], ws['D2'], ws['E2'], ws['F2'] = 'dancer_id', 'Concession', 'redacted', 'Week 1', 'Week 2'
    data = [
        # (dancer, concession, week1, week2)
        ('DNC-1', 'No', True, False),
        ('DNC-2', 'Yes', True, True),
        ('DNC-4', 'comp', True, False),  # unrecognised -> unknown
        ('DNC-9', None, True, False),  # blank -> NULL
        ('DNC-1', None, False, True),  # duplicate ticket; blank concession must not wipe 'No'
    ]
    for i, (did, conc, w1, w2) in enumerate(data, start=3):
        ws[f'B{i}'], ws[f'E{i}'], ws[f'F{i}'] = did, w1, w2
        if conc is not None:
            ws[f'C{i}'] = conc
    return ws


def _social_register_ws(title='Online bookings'):
    """A one-off social register: the attendee list in two side-by-side halves.

    Each half is '# | dancer_id | redacted | Concession | Present?'; the sheet has no dates
    (the event date lives in the filename) and no dated session columns. DNC-1 appears twice
    (an un-renamed extra ticket); DNC-5 was absent (❎).
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = title
    ws['A1'] = 'Tea Dance 25th Feb 2024'  # a title note above the header
    ws['B2'], ws['C2'], ws['D2'], ws['E2'], ws['F2'] = '#', 'dancer_id', 'redacted', 'Concession', 'Present?'
    ws['H2'], ws['I2'], ws['J2'], ws['K2'], ws['L2'] = '#', 'dancer_id', 'redacted', 'Concession', 'Present?'
    # left half
    ws['B3'], ws['C3'], ws['E3'], ws['F3'] = 1, 'DNC-1', 'No', '☑️'
    ws['B4'], ws['C4'], ws['E4'], ws['F4'] = 2, 'DNC-2', 'Yes', '☑️'
    ws['B5'], ws['C5'], ws['E5'], ws['F5'] = 3, 'DNC-1', 'No', '☑️'  # same person, second ticket
    # right half (different people)
    ws['H3'], ws['I3'], ws['K3'], ws['L3'] = 39, 'DNC-3', 'Yes', '☑️'
    ws['H4'], ws['I4'], ws['K4'], ws['L4'] = 40, 'DNC-5', 'No', '❎'  # booked, absent
    return ws


def _attended_register_ws(title='26 March 23'):
    """A Sunday-Social-style register: a single block, an 'Attended' x marker, no concession.

    The event date is the tab name, not a cell. DNC-2 appears twice (extra ticket); DNC-3
    has a blank Attended cell (booked but absent).
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = title
    ws['A1'], ws['B1'], ws['C1'], ws['D1'], ws['E1'] = 'dancer_id', 'redacted', 'redacted', 'Order', 'Attended'
    ws['A2'], ws['D2'], ws['E2'] = 'DNC-1', 'OAOX', 'x'
    ws['A3'], ws['D3'], ws['E3'] = 'DNC-2', 'O7LY', 'x'
    ws['A4'], ws['D4'], ws['E4'] = 'DNC-2', 'O7LY', 'x'  # un-renamed extra ticket
    ws['A5'], ws['D5'] = 'DNC-3', 'O00J'  # blank Attended -> booked but absent
    return ws


def _blank_party_register_ws(title='Sheet1'):
    """An 'End of Term Party' allocation list: a Present? column that is entirely blank.

    Nobody's turnout was recorded, so every booker is a held place (UNKNOWN), not absent. The
    event date is in a title cell/filename, not in the tab name. DNC-1 holds two places.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = title
    ws['B12'] = 'End of Term Party 27th June 2024'
    ws['A13'], ws['B13'], ws['C13'], ws['D13'], ws['E13'] = '#', 'dancer_id', 'redacted', 'Concession', 'Present?'
    ws['A14'], ws['B14'], ws['D14'] = 1, 'DNC-1', 'Yes'
    ws['A15'], ws['B15'], ws['D15'] = 2, 'DNC-2', 'No'
    ws['A16'], ws['B16'], ws['D16'] = 3, 'DNC-1', 'Yes'  # a second held place, blank marker
    return ws


def _dancer_list_no_marker_ws(title='Class List'):
    """A dancer_id list with a Concession column but no Present?/Attended marker — not a register."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = title
    ws['A1'], ws['B1'], ws['C1'] = 'dancer_id', 'redacted', 'Concession'
    ws['A2'], ws['C2'] = 'DNC-1', 'Yes'
    ws['A3'], ws['C3'] = 'DNC-2', 'No'
    return ws


def _booking_list_ws(title='Attendees By Activity'):
    """A booking list, NOT an attendance register.

    Its 'present' column is mostly empty with the odd free-text note, and must not be read as
    everyone being absent.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = title
    ws['A1'], ws['B1'], ws['C1'], ws['D1'] = 'dancer_id', 'redacted', 'Status', 'present'
    ws['A2'], ws['C2'] = 'DNC-1', 'Confirmed'
    ws['A3'], ws['C3'] = 'DNC-2', 'Confirmed'
    ws['A4'], ws['C4'], ws['D4'] = 'DNC-3', 'Confirmed', 'no show'  # free-text note, not a mark
    ws['A5'], ws['C5'], ws['D5'] = 'DNC-4', 'Confirmed', 'paid £10 cash on door'
    ws['A6'], ws['C6'], ws['D6'] = 'DNC-5', 'Confirmed', 'refunded'  # a recognised non-attendance word
    return ws


def _booking_by_activity_ws(title='Attendees By Activity'):
    """A dancecloud 'Attendees By Activity' booking view: one row per dancer per dated session.

    No attendance is recorded — a row only means a ticket was held for that date. DNC-1 holds a
    ticket for both sessions; DNC-2 cancelled the second; DNC-3 holds two tickets for session 1.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = title
    ws['A1'], ws['B1'], ws['C1'], ws['D1'] = 'dancer_id', 'Activity', 'Date', 'Status'
    rows = [
        ('DNC-1', 'Lesson', datetime(2023, 1, 19, 19, 15), 'Confirmed'),
        ('DNC-1', 'Lesson', datetime(2023, 1, 26, 19, 15), 'Confirmed'),
        ('DNC-2', 'Lesson', datetime(2023, 1, 19, 19, 15), 'Confirmed'),
        ('DNC-2', 'Lesson', datetime(2023, 1, 26, 19, 15), 'Cancelled'),
        ('DNC-3', 'Lesson', datetime(2023, 1, 19, 19, 15), 'Confirmed'),
        ('DNC-3', 'Lesson', datetime(2023, 1, 19, 19, 15), 'Confirmed'),  # un-renamed extra ticket
    ]
    for i, (did, act, dt, status) in enumerate(rows, start=2):
        ws[f'A{i}'], ws[f'B{i}'], ws[f'C{i}'], ws[f'D{i}'] = did, act, dt, status
    return ws


def _booking_summary_ws(title='Attendees'):
    """The plain booking summary: a booking Status but no Date column and no present column.

    Deliberately NOT claimed — it is the same bookings as 'Attendees By Activity' without dates.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = title
    ws['A1'], ws['B1'], ws['C1'], ws['D1'] = 'dancer_id', 'Ticket Type', 'Concession', 'Status'
    ws['A2'], ws['B2'], ws['C2'], ws['D2'] = 'DNC-1', 'Level 1 Term A', 'No', 'Confirmed'
    ws['A3'], ws['B3'], ws['C3'], ws['D3'] = 'DNC-2', 'Level 1 Term A', 'Yes', 'Confirmed'
    return ws


def _booking_with_concession_ws(title='Attendees By Activity'):
    """An older booking export whose workbook also carries the 'Attendees' rollup with Concession.

    The dated 'Attendees By Activity' tab (returned) has no Checked In column — so the older
    :class:`BookingExportParser` claims it — and no member-rate flag. The sibling 'Attendees'
    rollup supplies it: DNC-1 is ordinary ('No'), DNC-2 member/concession ('Yes'), DNC-3 has no
    rollup row so its ticket type stays unknown. The two tabs share a parent so the parser can
    read across.
    """
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    by_act = wb.create_sheet(title)
    by_act['A1'], by_act['B1'], by_act['C1'], by_act['D1'] = 'dancer_id', 'Activity', 'Date', 'Status'
    rows = [
        ('DNC-1', datetime(2023, 1, 19, 19, 15)),
        ('DNC-2', datetime(2023, 1, 19, 19, 15)),
        ('DNC-3', datetime(2023, 1, 19, 19, 15)),
    ]
    for i, (did, dt) in enumerate(rows, start=2):
        by_act[f'A{i}'], by_act[f'B{i}'], by_act[f'C{i}'], by_act[f'D{i}'] = did, 'Lesson', dt, 'Confirmed'

    rollup = wb.create_sheet('Attendees')
    rollup['A1'], rollup['B1'], rollup['C1'], rollup['D1'] = 'dancer_id', 'Ticket Type', 'Concession', 'Status'
    for i, (did, conc) in enumerate([('DNC-1', 'No'), ('DNC-2', 'Yes')], start=2):
        rollup[f'A{i}'], rollup[f'B{i}'], rollup[f'C{i}'], rollup[f'D{i}'] = did, 'Level 1 Term A', conc, 'Confirmed'
    return by_act


def _booking_ws_on(month: int, day: int, title='Attendees By Activity'):
    """A minimal dated booking view: one dancer, one session, on the given 2023 date."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = title
    ws['A1'], ws['B1'], ws['C1'], ws['D1'] = 'dancer_id', 'Activity', 'Date', 'Status'
    ws['A2'], ws['B2'], ws['C2'], ws['D2'] = 'DNC-1', 'Lesson', datetime(2023, month, day, 19, 15), 'Confirmed'
    return ws


def _checked_in_by_activity_ws(title='Attendees By Activity'):
    """A modern (2026-on) dancecloud 'Attendees By Activity' export with a 'Checked In' column.

    One ticket combining a workshop and a tea dance is pre-expanded by dancecloud into one row per
    activity, so the two share a date but must become two activities — a lesson and a social. The
    'Checked In' cell is a door-scan timestamp (attended) or blank (absent); a 'Cancelled' Status
    is absent regardless. DNC-1 holds two workshop tickets, both scanned in (an un-renamed extra).
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = title
    ws['A1'], ws['B1'], ws['C1'], ws['D1'], ws['E1'] = 'dancer_id', 'Activity', 'Date', 'Status', 'Checked In'
    on = datetime(2026, 5, 10, 14, 0)
    scan = datetime(2026, 5, 10, 14, 3)  # a check-in timestamp
    rows = [
        ('DNC-1', 'Collegiate Shag Workshop', on, 'Confirmed', scan),  # attended (lesson)
        ('DNC-1', 'Tea Dance', on, 'Confirmed', None),  # absent (social, scanner ran)
        ('DNC-2', 'Collegiate Shag Workshop', on, 'Confirmed', None),  # absent
        ('DNC-2', 'Tea Dance', on, 'Confirmed', scan),  # attended
        ('DNC-3', 'Collegiate Shag Workshop', on, 'Cancelled', None),  # cancelled -> absent
        ('DNC-1', 'Collegiate Shag Workshop', on, 'Confirmed', scan),  # un-renamed extra ticket, scanned
    ]
    for i, (did, act, dt, status, checkin) in enumerate(rows, start=2):
        ws[f'A{i}'], ws[f'B{i}'], ws[f'C{i}'], ws[f'D{i}'] = did, act, dt, status
        if checkin is not None:
            ws[f'E{i}'] = checkin
    return ws


def _checked_in_multiday_class_ws(title='Attendees By Activity'):
    """A weekender export where a class ('Track A') runs on two days, scanned patchily on day 2.

    Used to check that a check-in to any day of a class implies attendance at every day it ran:
    DNC-1 scanned into Track A on day 1 only, DNC-2 on day 2 only — both should end attended on
    both days. DNC-3 never scanned Track A — absent both days. The nightly social is *not*
    propagated: DNC-1 scanned the day-1 social but not the day-2 one, so day 2 stays absent.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = title
    ws['A1'], ws['B1'], ws['C1'], ws['D1'], ws['E1'] = 'dancer_id', 'Activity', 'Date', 'Status', 'Checked In'
    d1, d2 = datetime(2026, 3, 21, 13, 0), datetime(2026, 3, 22, 13, 0)
    scan1, scan2 = datetime(2026, 3, 21, 13, 5), datetime(2026, 3, 22, 13, 5)
    rows = [
        ('DNC-1', 'Track A - Classes', d1, 'Confirmed', scan1),  # scanned day 1
        ('DNC-1', 'Track A - Classes', d2, 'Confirmed', None),  # blank day 2 -> lifted to attended
        ('DNC-2', 'Track A - Classes', d1, 'Confirmed', None),  # blank day 1 -> lifted to attended
        ('DNC-2', 'Track A - Classes', d2, 'Confirmed', scan2),  # scanned day 2
        ('DNC-3', 'Track A - Classes', d1, 'Confirmed', None),  # never scanned -> absent both days
        ('DNC-3', 'Track A - Classes', d2, 'Confirmed', None),
        ('DNC-1', 'Evening Social', d1, 'Confirmed', scan1),  # social: per-night, not propagated
        ('DNC-1', 'Evening Social', d2, 'Confirmed', None),  # blank -> stays absent
    ]
    for i, (did, act, dt, status, checkin) in enumerate(rows, start=2):
        ws[f'A{i}'], ws[f'B{i}'], ws[f'C{i}'], ws[f'D{i}'] = did, act, dt, status
        if checkin is not None:
            ws[f'E{i}'] = checkin
    return ws


def _swingout_ws(title='Stockbridge Swingout'):
    """The one-off Stockbridge Swingout register: a Ticket Type string and a boolean Registered.

    The Saturday date sits in the row above the 'Registered' header (the Friday social is the day
    before). Each ticket grants a fixed set of weekend activities, all-or-none, with Registered as
    the status. DNC-1 (Full Pass) and DNC-3 (two Saturday-Social-Only tickets) both also grant the
    Saturday social, so the duplicate scanned tickets become anonymous extra heads; DNC-4's Friday
    Social Only is a no-show (Registered False).
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = title
    ws['F2'] = datetime(2025, 10, 18)  # the Saturday date, above the Registered header
    ws['B3'], ws['C3'], ws['D3'], ws['E3'], ws['F3'] = '#', 'dancer_id', 'redacted', 'Ticket Type', 'Registered'
    rows = [
        ('DNC-1', 'Full Pass - Improvers / Intermediate', True),  # Friday social + Imp/Int classes + Sat social
        ('DNC-2', 'Saturday Only - Intermediate / Advanced', True),  # Int/Adv classes + Sat social
        ('DNC-3', 'Saturday Social Only', True),
        ('DNC-3', 'Saturday Social Only', True),  # same person, extra ticket -> anonymous Sat-social head
        ('DNC-4', 'Friday Social Only', False),  # booked the Friday social, no-show
        ('DNC-1', 'Saturday Social Only', True),  # DNC-1 already in Sat social via Full Pass -> extra head
    ]
    for i, (did, ticket, reg) in enumerate(rows, start=4):
        ws[f'C{i}'], ws[f'E{i}'], ws[f'F{i}'] = did, ticket, reg
    return ws


def _checked_in_no_scans_ws(title='Attendees By Activity'):
    """A modern export whose door scanner never ran: a 'Checked In' column that is entirely blank.

    Nobody anywhere is checked in, so a blank conveys nothing — every booking is UNKNOWN, not a
    sheet-wide no-show. DNC-1 holds two tickets for the one session (they collapse to one row).
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = title
    ws['A1'], ws['B1'], ws['C1'], ws['D1'], ws['E1'] = 'dancer_id', 'Activity', 'Date', 'Status', 'Checked In'
    on = datetime(2026, 1, 17, 13, 0)
    ws['A2'], ws['B2'], ws['C2'], ws['D2'] = 'DNC-1', "Teachers' Workshop", on, 'Confirmed'
    ws['A3'], ws['B3'], ws['C3'], ws['D3'] = 'DNC-2', "Teachers' Workshop", on, 'Confirmed'
    ws['A4'], ws['B4'], ws['C4'], ws['D4'] = 'DNC-1', "Teachers' Workshop", on, 'Confirmed'  # extra ticket
    return ws


def _checked_in_with_concession_ws(title='Attendees By Activity'):
    """A modern export whose workbook also carries the 'Attendees' rollup with a Concession column.

    The 'Attendees By Activity' tab (returned) holds no member-rate flag; the sibling 'Attendees'
    rollup in the same workbook does — one row per booking with a Yes/No 'Concession'. DNC-1 is a
    concession ('Yes' -> member/concession rate), DNC-2 an ordinary ('No'), DNC-3 has no rollup row
    (its ticket type stays unknown). The two tabs share a parent so the parser can read across.
    """
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    by_act = wb.create_sheet(title)
    by_act['A1'], by_act['B1'], by_act['C1'], by_act['D1'], by_act['E1'] = (
        'dancer_id',
        'Activity',
        'Date',
        'Status',
        'Checked In',
    )
    on, scan = datetime(2026, 6, 25, 19, 30), datetime(2026, 6, 25, 20, 0)
    rows = [('DNC-1', scan), ('DNC-2', None), ('DNC-3', scan)]
    for i, (did, checkin) in enumerate(rows, start=2):
        by_act[f'A{i}'], by_act[f'B{i}'], by_act[f'C{i}'], by_act[f'D{i}'] = did, 'End of term dance', on, 'Confirmed'
        if checkin is not None:
            by_act[f'E{i}'] = checkin

    rollup = wb.create_sheet('Attendees')
    rollup['A1'], rollup['B1'], rollup['C1'], rollup['D1'] = 'dancer_id', 'Ticket Type', 'Concession', 'Status'
    for i, (did, conc) in enumerate([('DNC-1', 'Yes'), ('DNC-2', 'No')], start=2):
        rollup[f'A{i}'], rollup[f'B{i}'], rollup[f'C{i}'], rollup[f'D{i}'] = did, 'End of term party', conc, 'Confirmed'
    return by_act
