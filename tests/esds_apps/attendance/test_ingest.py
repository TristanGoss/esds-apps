from datetime import datetime
from pathlib import Path

import openpyxl

from esds_apps.attendance import ingest

from .workbooks import _booking_by_activity_ws, _booking_summary_ws, _roster_ws

# ---- workbook-context helpers ----


def test_resolve_year_prefers_date_cells_then_path(tmp_path):
    """The dominant dated cell wins; with no dates, the year is recovered from the file/folder name."""
    wb = openpyxl.Workbook()
    wb.active['A1'] = 'Week 1 (10 Feb)'  # only a week label, no real date cell
    assert ingest._resolve_year(wb) is None
    # '2022' appears in the folder and the export date, '2029' (an export HHMM) only once -> 2022 wins
    p = tmp_path / '2022 Attendance Records' / 'L2 classes Feb March Attendees 2022-02-07 2029.xlsx'
    assert ingest._resolve_year(wb, p) == 2022
    wb.active['A2'] = datetime(2024, 3, 1)  # a real date cell now overrides the path
    assert ingest._resolve_year(wb, p) == 2024


def test_term_from_strips_attendance_and_suffix():
    assert ingest._term_from(Path('May-Jun 2025 Attendance_pseudonymised.xlsx')) == 'May-Jun 2025'
    assert ingest._term_from(Path('Attendance Nov-Dec 2024_pseudonymised.xlsx')) == 'Nov-Dec 2024'


def test_week_anchor_picks_earliest_date():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws['A1'], ws['A2'], ws['A3'] = datetime(2023, 11, 23), datetime(2023, 11, 9), 'not a date'
    assert ingest._week_anchor(wb) == datetime(2023, 11, 9)
    wb2 = openpyxl.Workbook()
    wb2.active['A1'] = 'text only'
    assert ingest._week_anchor(wb2) is None


# ---- folder dispatch ----


def test_ingest_folder_dispatches_and_reports_unhandled(tmp_path, db):
    root = tmp_path / 'outputs'
    (root / 'sub').mkdir(parents=True)
    _roster_ws().parent.save(root / 'sub' / 'May-Jun 2025 Attendance_pseudonymised.xlsx')

    # an unhandled workbook: a flat sheet with neither dancer_id sessions nor a tally
    wb = openpyxl.Workbook()
    wb.active.title = 'Sales'
    wb.active['A1'], wb.active['A2'] = 'total', 5
    wb.save(root / 'misc.xlsx')

    report = ingest.ingest_folder(root, db)
    handled = {(s, p) for _, s, p in report.handled}
    assert ('Level 1 Attendance', 'roster') in handled
    assert any(sheet == 'Sales' for _, sheet in report.unhandled)
    # the roster activities made it into the db via the folder path
    assert db.conn.execute("SELECT COUNT(*) FROM activity WHERE name='Level 1 (2025-05-22)'").fetchone() == (1,)


def test_ingest_folder_skips_readme_silently(tmp_path, db):
    root = tmp_path / 'outputs'
    root.mkdir()
    wb = openpyxl.Workbook()
    wb.active.title = 'README'
    wb.active['A1'] = 'How to use this workbook'
    wb.save(root / 'Jan-Feb 2025 Attendance_pseudonymised.xlsx')

    report = ingest.ingest_folder(root, db)
    assert report.handled == []
    assert all(sheet.lower() != 'readme' for _, sheet in report.unhandled)


def test_ingest_folder_skips_member_and_mail_sheets(tmp_path, db):
    root = tmp_path / 'outputs'
    root.mkdir()
    for title in ['Members', 'Membership', 'Mail List', 'Email']:
        wb = openpyxl.Workbook()
        wb.active.title = title
        wb.active['A1'] = 'data'
        wb.save(root / f'{title}.xlsx')

    report = ingest.ingest_folder(root, db)
    assert report.handled == []
    assert report.unhandled == []  # silently skipped, not in unhandled


def test_ingest_folder_skips_membership_workbooks(tmp_path, db):
    """Membership / master-members workbooks are purchases, not class or social attendance.

    Their 'Attendees By Activity' sheet is shaped exactly like a class booking export, so the
    skip is by filename: the whole workbook is passed over before any parser sees it.
    """
    root = tmp_path / 'outputs'
    root.mkdir()
    ws = _booking_by_activity_ws()  # the booking-export shape a membership file would also have
    ws.parent.save(root / 'ESDS Membership Attendees 2022-11-02_pseudonymised.xlsx')

    report = ingest.ingest_folder(root, db)
    assert report.handled == []
    assert report.unhandled == []  # skipped wholesale by filename, not flagged as needing a parser
    assert db.conn.execute('SELECT COUNT(*) FROM event').fetchone() == (0,)


def test_ingest_folder_reports_attendees_rollup_as_redundant(tmp_path, db):
    """A plain 'Attendees' rollup is reported redundant — not unparsed — when its by-activity twin exists.

    The dancecloud export carries both tabs; the rollup is the same bookings minus the dates, so no
    parser claims it. It belongs in `redundant`, away from the 'needs a new parser' list, while the
    dated 'Attendees By Activity' is ingested normally.
    """
    root = tmp_path / 'outputs'
    root.mkdir()
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    by_act = _booking_by_activity_ws('Attendees By Activity')
    rollup = _booking_summary_ws('Attendees')
    for src in (by_act, rollup):
        dst = wb.create_sheet(src.title)
        for row in src.iter_rows(values_only=True):
            dst.append(row)
    wb.save(root / 'Level 2 Term B Attendees 2022-11-03_pseudonymised.xlsx')

    report = ingest.ingest_folder(root, db)
    assert ('Level 2 Term B Attendees 2022-11-03_pseudonymised.xlsx', 'Attendees') in report.redundant
    assert all(sheet != 'Attendees' for _, sheet in report.unhandled)
    assert any(p == 'booking_export' for _, _, p in report.handled)


def test_ingest_folder_flags_lone_attendees_rollup(tmp_path, db):
    """A lone 'Attendees' rollup (no by-activity twin) is still flagged — it may be the only record."""
    root = tmp_path / 'outputs'
    root.mkdir()
    _booking_summary_ws('Attendees').parent.save(root / 'ESDS Level 2 classes Feb March_pseudonymised.xlsx')

    report = ingest.ingest_folder(root, db)
    assert ('ESDS Level 2 classes Feb March_pseudonymised.xlsx', 'Attendees') in report.unhandled
    assert report.redundant == []


def test_ingest_folder_skips_bookkeeping_sheets(tmp_path, db):
    """'Exceptions', 'Loyalty' and 'Tickets' tabs carry no per-session attendance."""
    root = tmp_path / 'outputs'
    root.mkdir()
    for title in ['Exceptions', 'Loyalty', 'Tickets']:
        wb = openpyxl.Workbook()
        wb.active.title = title
        wb.active['A1'] = 'data'
        wb.save(root / f'{title}.xlsx')

    report = ingest.ingest_folder(root, db)
    assert report.handled == []
    assert report.unhandled == []  # silently skipped, not in unhandled
