import json

import pytest

from esds_apps.attendance import pseudonyms_db

PASSPHRASE = 'test-passphrase-esds'


# ============================================================================
# Key derivation
# ============================================================================


def test_derive_keys_length():
    import secrets

    salt = secrets.token_bytes(32)
    fernet_key, mac_key = pseudonyms_db._derive_keys(PASSPHRASE, salt)
    assert len(mac_key) == 32
    assert len(fernet_key) > 32  # urlsafe-base64 encoded


def test_derive_keys_deterministic():
    import secrets

    salt = secrets.token_bytes(32)
    k1a, k1b = pseudonyms_db._derive_keys(PASSPHRASE, salt)
    k2a, k2b = pseudonyms_db._derive_keys(PASSPHRASE, salt)
    assert k1a == k2a
    assert k1b == k2b


def test_derive_keys_differs_with_different_salt():
    import secrets

    _, mac1 = pseudonyms_db._derive_keys(PASSPHRASE, secrets.token_bytes(32))
    _, mac2 = pseudonyms_db._derive_keys(PASSPHRASE, secrets.token_bytes(32))
    assert mac1 != mac2


def test_derive_id_key_deterministic():
    assert pseudonyms_db.derive_id_key(PASSPHRASE) == pseudonyms_db.derive_id_key(PASSPHRASE)


def test_derive_id_key_differs_with_different_passphrase():
    assert pseudonyms_db.derive_id_key(PASSPHRASE) != pseudonyms_db.derive_id_key('other-passphrase')


# ============================================================================
# Database setup and passphrase validation
# ============================================================================


def test_setup_db_creates_tables(tmp_db):
    conn = pseudonyms_db._setup_db(tmp_db)
    tables = {r[0] for r in conn.execute("SELECT name FROM sqlite_master WHERE type='table'")}
    assert {'meta', 'pseudonyms'} <= tables
    conn.close()


def test_open_db_creates_file(tmp_db):
    assert not tmp_db.exists()
    c = pseudonyms_db.open_db(tmp_db, PASSPHRASE)
    c.conn.close()
    assert tmp_db.exists()


def test_open_db_returns_dbcontext(tmp_db):
    c = pseudonyms_db.open_db(tmp_db, PASSPHRASE)
    assert isinstance(c, pseudonyms_db.DbContext)
    c.conn.close()


def test_open_db_stores_salt(tmp_db):
    c = pseudonyms_db.open_db(tmp_db, PASSPHRASE)
    salt_row = c.conn.execute('SELECT value FROM meta WHERE key="salt"').fetchone()
    c.conn.close()
    assert salt_row is not None


def test_open_db_reuses_salt_across_opens(tmp_db):
    c1 = pseudonyms_db.open_db(tmp_db, PASSPHRASE)
    mac1 = c1.mac_key
    c1.conn.close()
    c2 = pseudonyms_db.open_db(tmp_db, PASSPHRASE)
    mac2 = c2.mac_key
    c2.conn.close()
    assert mac1 == mac2


def test_open_db_wrong_passphrase_raises(tmp_db):
    c = pseudonyms_db.open_db(tmp_db, PASSPHRASE)
    c.conn.close()
    with pytest.raises(ValueError, match='Wrong passphrase'):
        pseudonyms_db.open_db(tmp_db, 'wrong-passphrase')


# ============================================================================
# Value hashing
# ============================================================================


def test_value_hash_consistent(ctx):
    assert pseudonyms_db._value_hash('Alice Smith', ctx.mac_key) == pseudonyms_db._value_hash(
        'Alice Smith', ctx.mac_key
    )


def test_value_hash_case_insensitive(ctx):
    assert pseudonyms_db._value_hash('alice smith', ctx.mac_key) == pseudonyms_db._value_hash(
        'ALICE SMITH', ctx.mac_key
    )


def test_value_hash_strips_whitespace(ctx):
    assert pseudonyms_db._value_hash('Alice', ctx.mac_key) == pseudonyms_db._value_hash('  Alice  ', ctx.mac_key)


def test_value_hash_differs_for_different_values(ctx):
    assert pseudonyms_db._value_hash('Alice', ctx.mac_key) != pseudonyms_db._value_hash('Bob', ctx.mac_key)


# ============================================================================
# Dancer ID management
# ============================================================================


def test_get_or_create_dancer_id_format(ctx):
    did = pseudonyms_db.get_or_create_dancer_id(
        ctx,
        {'first_name': 'Alice', 'last_name': 'Smith'},
        {'email': 'alice@example.com'},
    )
    assert did.startswith('DNC-')
    assert len(did) == 12  # 'DNC-' + 8 hex chars


def test_get_or_create_dancer_id_deduplicates_by_email(ctx):
    alice = {'first_name': 'Alice', 'last_name': 'Smith'}
    alicia = {'first_name': 'Alicia', 'last_name': 'Schmidt'}
    email = {'email': 'alice@example.com'}
    id1 = pseudonyms_db.get_or_create_dancer_id(ctx, alice, email)
    id2 = pseudonyms_db.get_or_create_dancer_id(ctx, alicia, email)
    assert id1 == id2


def test_get_or_create_dancer_id_deduplicates_by_name(ctx):
    bob = {'first_name': 'Bob', 'last_name': 'Jones'}
    id1 = pseudonyms_db.get_or_create_dancer_id(ctx, bob, {'email': 'bob1@example.com'})
    id2 = pseudonyms_db.get_or_create_dancer_id(ctx, bob, {'email': 'bob2@example.com'})
    assert id1 == id2


def test_get_or_create_dancer_id_distinct_people(ctx):
    alice = {'first_name': 'Alice', 'last_name': 'Smith'}
    bob = {'first_name': 'Bob', 'last_name': 'Jones'}
    id1 = pseudonyms_db.get_or_create_dancer_id(ctx, alice, {'email': 'alice@example.com'})
    id2 = pseudonyms_db.get_or_create_dancer_id(ctx, bob, {'email': 'bob@example.com'})
    assert id1 != id2


_SELECT_ENC = 'SELECT enc_name, enc_email FROM pseudonyms WHERE dancer_id=?'


def test_get_or_create_dancer_id_name_only(ctx):
    alice = {'first_name': 'Alice', 'last_name': 'Smith'}
    did = pseudonyms_db.get_or_create_dancer_id(ctx, alice, None)
    assert did.startswith('DNC-')
    enc_name, enc_email = ctx.conn.execute(_SELECT_ENC, (did,)).fetchone()
    assert enc_name is not None
    assert enc_email is None


def test_get_or_create_dancer_id_email_only(ctx):
    did = pseudonyms_db.get_or_create_dancer_id(ctx, None, {'email': 'anon@example.com'})
    assert did.startswith('DNC-')
    enc_name, enc_email = ctx.conn.execute(_SELECT_ENC, (did,)).fetchone()
    assert enc_name is None
    assert enc_email is not None


def test_get_or_create_dancer_id_stores_json(ctx):
    name_fields = {'first_name': 'Alice', 'last_name': 'Smith'}
    email_fields = {'email': 'alice@example.com'}
    did = pseudonyms_db.get_or_create_dancer_id(ctx, name_fields, email_fields)
    enc_name, enc_email = ctx.conn.execute(_SELECT_ENC, (did,)).fetchone()
    assert json.loads(ctx.fernet.decrypt(enc_name.encode())) == name_fields
    assert json.loads(ctx.fernet.decrypt(enc_email.encode())) == email_fields


def test_get_or_create_updates_missing_name(ctx):
    """If dancer was found by email but had no name, name should be filled in."""
    did = pseudonyms_db.get_or_create_dancer_id(ctx, None, {'email': 'alice@example.com'})
    pseudonyms_db.get_or_create_dancer_id(
        ctx, {'first_name': 'Alice', 'last_name': 'Smith'}, {'email': 'alice@example.com'}
    )
    result = pseudonyms_db.decrypt_dancer(ctx, did)
    assert result['name'] == {'first_name': 'Alice', 'last_name': 'Smith'}


def test_get_or_create_updates_missing_email(ctx):
    """If dancer was found by name but had no email, email should be filled in."""
    did = pseudonyms_db.get_or_create_dancer_id(ctx, {'first_name': 'Alice', 'last_name': 'Smith'}, None)
    pseudonyms_db.get_or_create_dancer_id(
        ctx, {'first_name': 'Alice', 'last_name': 'Smith'}, {'email': 'alice@example.com'}
    )
    result = pseudonyms_db.decrypt_dancer(ctx, did)
    assert result['email'] == {'email': 'alice@example.com'}


def test_get_or_create_stores_alt_first_name(ctx):
    """Same email, different first name → alt_first_name stored on existing record."""
    did = pseudonyms_db.get_or_create_dancer_id(
        ctx, {'first_name': 'Alice', 'last_name': 'Smith'}, {'email': 'alice@example.com'}
    )
    pseudonyms_db.get_or_create_dancer_id(
        ctx, {'first_name': 'Xiaoling', 'last_name': 'Smith'}, {'email': 'alice@example.com'}
    )
    result = pseudonyms_db.decrypt_dancer(ctx, did)
    assert result['name']['first_name'] == 'Alice'
    assert result['name'].get('alt_first_name') == 'Xiaoling'


def test_get_or_create_alt_first_name_not_overwritten(ctx):
    """alt_first_name is only written once; a third name doesn't overwrite it."""
    did = pseudonyms_db.get_or_create_dancer_id(
        ctx, {'first_name': 'Alice', 'last_name': 'Smith'}, {'email': 'alice@example.com'}
    )
    pseudonyms_db.get_or_create_dancer_id(
        ctx, {'first_name': 'Xiaoling', 'last_name': 'Smith'}, {'email': 'alice@example.com'}
    )
    pseudonyms_db.get_or_create_dancer_id(
        ctx, {'first_name': 'Xiao', 'last_name': 'Smith'}, {'email': 'alice@example.com'}
    )
    result = pseudonyms_db.decrypt_dancer(ctx, did)
    assert result['name']['alt_first_name'] == 'Xiaoling'  # not overwritten by 'Xiao'


def test_get_or_create_stores_alt_last_name(ctx):
    """Same email, different last name (e.g. married) → alt_last_name stored on existing record."""
    did = pseudonyms_db.get_or_create_dancer_id(
        ctx, {'first_name': 'Alice', 'last_name': 'Smith'}, {'email': 'alice@example.com'}
    )
    pseudonyms_db.get_or_create_dancer_id(
        ctx, {'first_name': 'Alice', 'last_name': 'Jones'}, {'email': 'alice@example.com'}
    )
    result = pseudonyms_db.decrypt_dancer(ctx, did)
    assert result['name']['last_name'] == 'Smith'
    assert result['name'].get('alt_last_name') == 'Jones'


def test_get_or_create_alt_last_name_not_overwritten(ctx):
    """alt_last_name is only written once; a third last name doesn't overwrite it."""
    did = pseudonyms_db.get_or_create_dancer_id(
        ctx, {'first_name': 'Alice', 'last_name': 'Smith'}, {'email': 'alice@example.com'}
    )
    pseudonyms_db.get_or_create_dancer_id(
        ctx, {'first_name': 'Alice', 'last_name': 'Jones'}, {'email': 'alice@example.com'}
    )
    pseudonyms_db.get_or_create_dancer_id(
        ctx, {'first_name': 'Alice', 'last_name': 'Brown'}, {'email': 'alice@example.com'}
    )
    result = pseudonyms_db.decrypt_dancer(ctx, did)
    assert result['name']['alt_last_name'] == 'Jones'  # not overwritten by 'Brown'


def test_get_or_create_stores_alt_first_and_last_name_together(ctx):
    """A differing first and last name on the same encounter are both captured."""
    did = pseudonyms_db.get_or_create_dancer_id(
        ctx, {'first_name': 'Alice', 'last_name': 'Smith'}, {'email': 'alice@example.com'}
    )
    pseudonyms_db.get_or_create_dancer_id(
        ctx, {'first_name': 'Ali', 'last_name': 'Jones'}, {'email': 'alice@example.com'}
    )
    result = pseudonyms_db.decrypt_dancer(ctx, did)
    assert result['name']['alt_first_name'] == 'Ali'
    assert result['name']['alt_last_name'] == 'Jones'


def test_get_or_create_stores_alt_email(ctx):
    """Same name, different second email → alt_email stored on existing record."""
    did = pseudonyms_db.get_or_create_dancer_id(
        ctx, {'first_name': 'Alice', 'last_name': 'Smith'}, {'email': 'alice@work.com'}
    )
    pseudonyms_db.get_or_create_dancer_id(
        ctx, {'first_name': 'Alice', 'last_name': 'Smith'}, {'email': 'alice@home.com'}
    )
    result = pseudonyms_db.decrypt_dancer(ctx, did)
    assert result['email']['email'] == 'alice@work.com'
    assert result['email'].get('alt_email') == 'alice@home.com'


def test_get_or_create_alt_email_not_overwritten(ctx):
    """alt_email is only written once; a third email doesn't overwrite it."""
    did = pseudonyms_db.get_or_create_dancer_id(
        ctx, {'first_name': 'Alice', 'last_name': 'Smith'}, {'email': 'alice@work.com'}
    )
    pseudonyms_db.get_or_create_dancer_id(
        ctx, {'first_name': 'Alice', 'last_name': 'Smith'}, {'email': 'alice@home.com'}
    )
    pseudonyms_db.get_or_create_dancer_id(
        ctx, {'first_name': 'Alice', 'last_name': 'Smith'}, {'email': 'alice@other.com'}
    )
    result = pseudonyms_db.decrypt_dancer(ctx, did)
    assert result['email']['alt_email'] == 'alice@home.com'  # not overwritten by 'alice@other.com'


def test_dancer_id_stable_across_fresh_dbs(tmp_path):
    """Same passphrase and input produces the same dancer ID in a freshly created DB."""
    name = {'first_name': 'Alice', 'last_name': 'Smith'}
    email = {'email': 'alice@example.com'}

    ctx1 = pseudonyms_db.open_db(tmp_path / 'db1.sqlite', PASSPHRASE)
    id1 = pseudonyms_db.get_or_create_dancer_id(ctx1, name, email)
    ctx1.conn.close()

    ctx2 = pseudonyms_db.open_db(tmp_path / 'db2.sqlite', PASSPHRASE)
    id2 = pseudonyms_db.get_or_create_dancer_id(ctx2, name, email)
    ctx2.conn.close()

    assert id1 == id2


def test_dancer_id_differs_with_different_passphrase(tmp_path):
    """Different passphrases produce different dancer IDs for the same input."""
    name = {'first_name': 'Alice', 'last_name': 'Smith'}
    email = {'email': 'alice@example.com'}

    ctx1 = pseudonyms_db.open_db(tmp_path / 'db1.sqlite', PASSPHRASE)
    id1 = pseudonyms_db.get_or_create_dancer_id(ctx1, name, email)
    ctx1.conn.close()

    ctx2 = pseudonyms_db.open_db(tmp_path / 'db2.sqlite', 'different-passphrase')
    id2 = pseudonyms_db.get_or_create_dancer_id(ctx2, name, email)
    ctx2.conn.close()

    assert id1 != id2


# ============================================================================
# Decryption
# ============================================================================


def test_decrypt_all_returns_canonical_dicts(ctx):
    pseudonyms_db.get_or_create_dancer_id(
        ctx, {'first_name': 'Alice', 'last_name': 'Smith'}, {'email': 'alice@example.com'}
    )
    pseudonyms_db.get_or_create_dancer_id(
        ctx, {'first_name': 'Bob', 'last_name': 'Jones'}, {'email': 'bob@example.com'}
    )
    results = pseudonyms_db.decrypt_all(ctx)
    assert len(results) == 2
    names = {r['name']['first_name'] for r in results}
    assert names == {'Alice', 'Bob'}
    for r in results:
        assert 'dancer_id' in r
        assert 'last_name' in r['name']
        assert 'email' in r['email']


def test_decrypt_dancer_found(ctx):
    did = pseudonyms_db.get_or_create_dancer_id(
        ctx, {'first_name': 'Alice', 'last_name': 'Smith'}, {'email': 'alice@example.com'}
    )
    result = pseudonyms_db.decrypt_dancer(ctx, did)
    assert result['dancer_id'] == did
    assert result['name'] == {'first_name': 'Alice', 'last_name': 'Smith'}
    assert result['email'] == {'email': 'alice@example.com'}


def test_decrypt_dancer_not_found(ctx):
    assert pseudonyms_db.decrypt_dancer(ctx, 'DNC-NONEXIST') is None


def test_decrypt_dancer_null_email(ctx):
    did = pseudonyms_db.get_or_create_dancer_id(ctx, {'first_name': 'Alice', 'last_name': 'Smith'}, None)
    result = pseudonyms_db.decrypt_dancer(ctx, did)
    assert result['email'] is None
    assert result['name']['first_name'] == 'Alice'


# ============================================================================
# Dancer ID substitution
# ============================================================================


def test_substitute_removes_old_id(ctx):
    id1 = pseudonyms_db.get_or_create_dancer_id(ctx, {'first_name': 'Alice', 'last_name': 'Smith'}, None)
    id2 = pseudonyms_db.get_or_create_dancer_id(
        ctx, {'first_name': 'Alicia', 'last_name': 'Smith'}, {'email': 'alice@example.com'}
    )
    pseudonyms_db.substitute_dancer_id(ctx, id1, id2)
    assert pseudonyms_db.decrypt_dancer(ctx, id1) is None


def test_substitute_merges_name_into_new(ctx):
    """old_id has name, new_id has email only → name moves to new_id."""
    id1 = pseudonyms_db.get_or_create_dancer_id(ctx, {'first_name': 'Alice', 'last_name': 'Smith'}, None)
    id2 = pseudonyms_db.get_or_create_dancer_id(ctx, None, {'email': 'alice@example.com'})
    pseudonyms_db.substitute_dancer_id(ctx, id1, id2)
    result = pseudonyms_db.decrypt_dancer(ctx, id2)
    assert result['name'] == {'first_name': 'Alice', 'last_name': 'Smith'}
    assert result['email'] == {'email': 'alice@example.com'}


def test_substitute_does_not_overwrite_existing_fields(ctx):
    """new_id already has a name — old_id's name should not overwrite it."""
    id1 = pseudonyms_db.get_or_create_dancer_id(ctx, {'first_name': 'OldName', 'last_name': 'X'}, None)
    id2 = pseudonyms_db.get_or_create_dancer_id(
        ctx, {'first_name': 'NewName', 'last_name': 'Y'}, {'email': 'x@example.com'}
    )
    pseudonyms_db.substitute_dancer_id(ctx, id1, id2)
    result = pseudonyms_db.decrypt_dancer(ctx, id2)
    assert result['name']['first_name'] == 'NewName'


def test_substitute_stores_alt_first_name_when_explicitly_passed(ctx):
    """conflict_first_name is stored as alt_first_name on new_id when provided."""
    id1 = pseudonyms_db.get_or_create_dancer_id(
        ctx, {'first_name': 'Mei', 'last_name': 'Chen'}, {'email': 'mei.chen@example.com'}
    )
    id2 = pseudonyms_db.get_or_create_dancer_id(ctx, {'first_name': 'Mary', 'last_name': 'Chen'}, None)
    pseudonyms_db.substitute_dancer_id(ctx, id1, id2, conflict_first_name='Mei')
    result = pseudonyms_db.decrypt_dancer(ctx, id2)
    assert result['name']['first_name'] == 'Mary'
    assert result['name']['alt_first_name'] == 'Mei'


def test_substitute_discards_conflict_first_name_by_default(ctx):
    """Without conflict_first_name, differing first names are silently dropped."""
    id1 = pseudonyms_db.get_or_create_dancer_id(ctx, {'first_name': 'Mei', 'last_name': 'Chen'}, None)
    id2 = pseudonyms_db.get_or_create_dancer_id(ctx, {'first_name': 'Mary', 'last_name': 'Chen'}, None)
    pseudonyms_db.substitute_dancer_id(ctx, id1, id2)
    result = pseudonyms_db.decrypt_dancer(ctx, id2)
    assert result['name']['first_name'] == 'Mary'
    assert 'alt_first_name' not in result['name']


def test_substitute_stores_alt_last_name_when_explicitly_passed(ctx):
    """conflict_last_name is stored as alt_last_name on new_id when provided."""
    id1 = pseudonyms_db.get_or_create_dancer_id(
        ctx, {'first_name': 'Mei', 'last_name': 'Smith'}, {'email': 'mei.smith@example.com'}
    )
    id2 = pseudonyms_db.get_or_create_dancer_id(ctx, {'first_name': 'Mei', 'last_name': 'Jones'}, None)
    pseudonyms_db.substitute_dancer_id(ctx, id1, id2, conflict_last_name='Smith')
    result = pseudonyms_db.decrypt_dancer(ctx, id2)
    assert result['name']['last_name'] == 'Jones'
    assert result['name']['alt_last_name'] == 'Smith'


def test_substitute_discards_conflict_last_name_by_default(ctx):
    """Without conflict_last_name, differing last names are silently dropped."""
    id1 = pseudonyms_db.get_or_create_dancer_id(ctx, {'first_name': 'Mei', 'last_name': 'Smith'}, None)
    id2 = pseudonyms_db.get_or_create_dancer_id(ctx, {'first_name': 'Mei', 'last_name': 'Jones'}, None)
    pseudonyms_db.substitute_dancer_id(ctx, id1, id2)
    result = pseudonyms_db.decrypt_dancer(ctx, id2)
    assert result['name']['last_name'] == 'Jones'
    assert 'alt_last_name' not in result['name']


def test_substitute_stores_alt_email_when_explicitly_passed(ctx):
    """conflict_email is stored as alt_email on new_id when provided."""
    id1 = pseudonyms_db.get_or_create_dancer_id(
        ctx, {'first_name': 'Alice', 'last_name': 'Smith'}, {'email': 'alice@work.com'}
    )
    id2 = pseudonyms_db.get_or_create_dancer_id(ctx, {'first_name': 'Alicia', 'last_name': 'Smith'}, None)
    pseudonyms_db.substitute_dancer_id(ctx, id2, id1, conflict_email='alice@personal.com')
    result = pseudonyms_db.decrypt_dancer(ctx, id1)
    assert result['email']['email'] == 'alice@work.com'
    assert result['email']['alt_email'] == 'alice@personal.com'


def test_substitute_discards_conflict_email_by_default(ctx):
    """Without conflict_email, no alt_email is written to the surviving record."""
    id1 = pseudonyms_db.get_or_create_dancer_id(
        ctx, {'first_name': 'Alice', 'last_name': 'Smith'}, {'email': 'alice@work.com'}
    )
    id2 = pseudonyms_db.get_or_create_dancer_id(ctx, {'first_name': 'Alicia', 'last_name': 'Smith'}, None)
    pseudonyms_db.substitute_dancer_id(ctx, id2, id1)
    result = pseudonyms_db.decrypt_dancer(ctx, id1)
    assert 'alt_email' not in result['email']


def test_substitute_raises_for_missing_id(ctx):
    id1 = pseudonyms_db.get_or_create_dancer_id(ctx, {'first_name': 'Alice', 'last_name': 'Smith'}, None)
    with pytest.raises(ValueError):
        pseudonyms_db.substitute_dancer_id(ctx, id1, 'DNC-NOTEXIST')


def test_substitute_rewrites_xlsx_files(tmp_path, ctx):
    import openpyxl

    id1 = pseudonyms_db.get_or_create_dancer_id(ctx, {'first_name': 'Alice', 'last_name': 'Smith'}, None)
    id2 = pseudonyms_db.get_or_create_dancer_id(ctx, None, {'email': 'alice@example.com'})

    # Write an xlsx that contains id1.
    p = tmp_path / 'out.xlsx'
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(['dancer_id', 'Notes'])
    ws.append([id1, 'something'])
    wb.save(p)

    pseudonyms_db.substitute_dancer_id(ctx, id1, id2, output_dir=tmp_path)

    wb2 = openpyxl.load_workbook(p)
    values = [row[0].value for row in wb2.active.iter_rows()]
    assert id2 in values
    assert id1 not in values
