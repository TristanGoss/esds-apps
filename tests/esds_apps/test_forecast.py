"""Tests for the /forecast what-if backend and its routes.

The behavioural context is normally fitted from the attendance database; here we build a small synthetic
``ForecastContext`` so the ledger, the unscented transform and the response shape can be tested without a
database. A complete defaults dict stands in for the private spreadsheet.
"""

import io
from datetime import date
from http import HTTPStatus

import numpy as np
import openpyxl
import pytest
from fastapi.testclient import TestClient

from esds_apps import forecast
from esds_apps.main import app

# A full set of financial defaults, standing in for the private spreadsheet.
DEFAULTS = {
    'forecast_ay': 2026,
    'n_tea_dances': 2,
    'have_weekender': True,
    'loyalty_enabled': True,
    'loyalty_every': 8,
    'class_size_cap': None,
    'l1_per_night': False,
    'class_card_size': 1,
    'confidence': 0.95,
    'price_class_disc': 7,
    'price_class_ord': 8,
    'price_social_disc': 9,
    'price_social_ord': 10,
    'price_social_only_disc': 3.5,
    'price_social_only_ord': 4.5,
    'price_wk_full_disc': 100,
    'price_wk_full_ord': 120,
    'price_wk_day_disc': 40,
    'price_wk_day_ord': 45,
    'price_wk_social': 40,
    'membership_fee': 12,
    'room_per_hour': 35,
    'room_per_hour_b': 20,
    'class_room_hour_a': 2.75,
    'class_room_hour_b': 2.0,
    'teacher_rate': 15,
    'teacher_hours_a': 2,
    'teacher_hours_b': 2,
    'teacher_hours_c': 1,
    'teacher_hours_d': 1,
    'social_snacks': 65,
    'social_room_hours': 4,
    'social_room_per_hour': 35,
    'band_cost_mean': 717.0,
    'band_cost_std': 97.0,
    'weekender_bands': 3,
    'weekender_room_hours': 40,
    'weekender_room_per_hour': 40,
    'weekender_teachers': 2,
    'weekender_teacher_rate': 60,
    'weekender_teacher_hours': 12,
    'weekender_flight': 250,
    'weekender_board_per_night': 90,
    'weekender_nights': 2,
    'n_committee': 9,
    'n_safer_spaces': 4,
    'safer_spaces_accounts': True,
    'n_extra_volunteers': 4,
    'n_legacy_accounts': 2,
    'n_shared_accounts': 1,
    'gsuite_seat_monthly': 5.9,
    'wix_reseller_annual': 157.44,
    'volunteer_social_per_head': 35,
    'n_volunteer_socials': 2,
    'oh_website': 129.6,
    'oh_email_marketing': 237.6,
    'oh_insurance': 109.6,
    'oh_storage_container': 516.66,
    'oh_spotify': 127.3,
    'oh_survey_monkey': 148.5,
    'oh_pat_testing': 110.4,
    'oh_equipment_recap': 150,
    'oh_society_phone': 96,
    'oh_posters': 120,
    'oh_stationery': 120,
    'oh_canva': 110,
    'n_members': 72,
    'membership_std': 11.6,
    'current_balance': 19498,
    'dc_rate': 0.025,
    'dc_fixed': 0.5,
    'dc_cap': 5.0,
    'stripe_rate': 0.015,
    'stripe_fixed': 0.2,
}


def make_context():
    """A small but internally consistent context (positive-definite covariance, valid survival fit)."""
    return forecast.ForecastContext(
        beta=np.array([30.0, -7.0, 5.0, 20.0, 0.0, -8.0, -18.0]),
        s2=25.0,
        xtxi=np.eye(7) * 0.01,
        dof=10,
        l1_attend_curve=[0.74, 0.66, 0.62, 0.57, 0.50, 0.44, 0.39],
        l1_history=[{'ay': 2024, 'pos': 1, 'interest': 50}, {'ay': 2025, 'pos': 1, 'interest': 45}],
        l2_vals=[25.0, 27.0, 24.0, 26.0],
        soc_vals=[68.0, 70.0, 66.0, 72.0, 64.0, 69.0, 71.0],
        xmas_vals=[68.0, 136.0],
        social_only_vals=[5.0, 4.0, 6.0, 5.0],
        wk_mix={
            'reference': 'Ref',
            'attendees_ref': 195,
            'retained': 87,
            'full': 31,
            'partial': 17,
            'social': 24,
            'single': 15,
        },
        fractions={'l1': 0.34, 'l2': 0.68, 'soc': 0.41, 'wk': 0.5},
        frac_n={'l1': 1000, 'l2': 800, 'soc': 300, 'wk': 195},
        loyalty_family='weibull',
        loyalty_popt=[3.0, 1.0],
        loyalty_n_dancers=110,
        loyalty_obs_thu=23,
        survival_l2={
            'k': [1, 2, 3],
            'empirical': [1.0, 0.5, 0.3],
            'fit': [1.0, 0.5, 0.3],
            'r2': 0.99,
            'family': 'weibull',
        },
        defaults=dict(DEFAULTS),
    )


@pytest.fixture
def ctx(monkeypatch):
    c = make_context()
    monkeypatch.setattr(forecast, '_CONTEXT', c)
    monkeypatch.setattr(forecast, 'get_context', lambda force_rebuild=False: c)
    return c


# ---------------------------------------------------------------- calendar / socials
def test_build_calendar_six_terms():
    cal = forecast.build_calendar(2026)
    assert len(cal['terms']) == 6
    assert cal['first_class'].month == 9 and cal['first_class'].weekday() == forecast.THU
    assert cal['eoy_party'].month == 6
    # terms run back to back within each region and cover 5-7 weeks each
    assert all(5 <= t['weeks'] <= 7 for t in cal['terms'])


def test_place_socials_orders_weekender_in_spring():
    cal = forecast.build_calendar(2026)
    socials = forecast.place_socials(cal, n_tea=2, have_weekender=True)
    labels = [s[0] for s in socials]
    assert labels.count('Christmas party') == 1
    assert labels.count('End-of-year party') == 1
    weekender = [d for lbl, d in socials if lbl == 'Weekender'][0]
    assert weekender.weekday() == forecast.FRI
    assert date(2027, 3, 1) <= weekender <= date(2027, 6, 1)


def test_place_socials_without_weekender():
    cal = forecast.build_calendar(2026)
    socials = forecast.place_socials(cal, n_tea=2, have_weekender=False)
    assert 'Weekender' not in [s[0] for s in socials]
    assert len(socials) == 4  # 2 tea dances + 2 parties


# ---------------------------------------------------------------- predict / ledger
def test_predict_shape_and_reconciliation(ctx):
    r = forecast.predict()
    # budget point estimate reconciles with the headline net
    rev = sum(b['amount'] for b in r['budget'] if b['kind'] == 'revenue')
    cost = sum(b['amount'] for b in r['budget'] if b['kind'] == 'cost')
    assert abs((rev - cost) - r['net']['mean']) <= 1
    # confidence interval brackets the mean and widens with the confidence level
    assert r['net']['lo'] < r['net']['mean'] < r['net']['hi']
    assert 0.0 <= r['p_green'] <= 1.0
    assert r['balance']['dates'] and len(r['balance']['mean']) == len(r['balance']['dates'])
    assert all(lo <= m <= hi for lo, m, hi in zip(r['balance']['lo'], r['balance']['mean'], r['balance']['hi']))
    assert r['year_end_balance']['mean'] == r['balance']['mean'][-1]
    assert r['detail']['l1_signups'] and len(r['detail']['inputs']) == 18


def test_confidence_widens_interval(ctx):
    narrow = forecast.predict(confidence=0.80)['net']
    wide = forecast.predict(confidence=0.99)['net']
    assert (wide['hi'] - wide['lo']) > (narrow['hi'] - narrow['lo'])


def test_override_changes_result(ctx):
    base = forecast.predict()['net']['mean']
    dearer = forecast.predict({'price_class_ord': 12})['net']['mean']
    assert dearer > base  # charging more for classes improves the year


def test_disabling_loyalty_is_stable(ctx):
    """A toggle that zeroes an input's variance must not break the covariance Cholesky."""
    r = forecast.predict({'loyalty_enabled': False})
    assert 'Loyalty refunds' not in [b['category'] for b in r['budget']]
    assert r['net']['lo'] < r['net']['hi']


def test_per_night_l1_uses_attendance_curve(ctx):
    """Per-night selling brings in the drop-off, so Level 1 income falls versus a term block."""
    block = forecast.predict({'l1_per_night': False})
    nightly = forecast.predict({'l1_per_night': True})
    block_l1 = [b['amount'] for b in block['budget'] if b['category'] == 'Level 1 classes'][0]
    nightly_l1 = [b['amount'] for b in nightly['budget'] if b['category'] == 'Level 1 classes'][0]
    assert nightly_l1 < block_l1


def test_merge_params_coerces_types(ctx):
    p = forecast._merge_params(
        ctx, {'n_committee': '15', 'have_weekender': 'false', 'class_size_cap': '', 'teacher_rate': '22.5'}
    )
    assert p['n_committee'] == 15 and isinstance(p['n_committee'], int)
    assert p['have_weekender'] is False
    assert p['class_size_cap'] is None
    assert p['teacher_rate'] == 22.5


def test_class_size_cap_limits_signups(ctx):
    base = forecast.predict()['net']['mean']
    capped = forecast.predict({'class_size_cap': 5})['net']['mean']
    assert capped < base  # capping paying sign-ups reduces class income


def test_weekender_social_pass_priced(ctx):
    """The social-pass share of the weekender audience earns revenue at its flat price."""
    base = forecast.predict()['net']['mean']
    dearer = forecast.predict({'price_wk_social': 200})['net']['mean']
    assert dearer > base  # more social-pass income lifts the year
    off = forecast.predict({'have_weekender': False})
    assert 'Weekender tickets' not in [b['category'] for b in off['budget']]


def test_volunteer_socials_count_scales_cost(ctx):
    """The number of volunteer socials is a control; doubling it doubles that cost line."""
    two = forecast.predict({'n_volunteer_socials': 2})
    four = forecast.predict({'n_volunteer_socials': 4})
    c2 = [b['amount'] for b in two['budget'] if b['category'] == 'Volunteer socials'][0]
    c4 = [b['amount'] for b in four['budget'] if b['category'] == 'Volunteer socials'][0]
    assert c4 == 2 * c2
    none = forecast.predict({'n_volunteer_socials': 0})
    assert 'Volunteer socials' not in [b['category'] for b in none['budget']]


def test_calendar_dates_exposed(ctx):
    cal = forecast.predict()['calendar']
    assert cal['first_class'] and len(cal['terms']) == 6
    assert {'Christmas party', 'End-of-year party'} <= {s['label'] for s in cal['socials']}


def test_safer_spaces_accounts_toggle(ctx):
    """Turning off safer-spaces Workspace accounts drops those seats from the overheads bill."""
    on = forecast.predict({'safer_spaces_accounts': True})
    off = forecast.predict({'safer_spaces_accounts': False})
    oh_on = [b['amount'] for b in on['budget'] if b['category'] == 'Overheads'][0]
    oh_off = [b['amount'] for b in off['budget'] if b['category'] == 'Overheads'][0]
    # 4 safer-spaces seats * £5.90/mo * 12 = ~£283/yr less
    assert oh_off < oh_on
    assert round(oh_on - oh_off) == round(4 * 5.9 * 12)


def test_separate_room_rates(ctx):
    """The second Thursday room, socials and weekender each have their own hourly rate."""
    base = forecast.predict()['net']['mean']
    # a dearer second Thursday room only raises class-venue cost
    dearer_second = forecast.predict({'room_per_hour_b': 40})['net']['mean']
    assert dearer_second < base
    # the social and weekender rates are independent of the Thursday main rate
    dearer_social = forecast.predict({'social_room_per_hour': 100})['net']['mean']
    dearer_wk = forecast.predict({'weekender_room_per_hour': 100})['net']['mean']
    assert dearer_social < base and dearer_wk < base


def test_class_card_reduces_dancecloud_fees(ctx):
    """Bigger class-card packs cut Dancecloud fees (fixed fee amortised over the pack) without changing revenue."""
    single = forecast.predict({'class_card_size': 1})
    pack = forecast.predict({'class_card_size': 10})
    fee_single = [b['amount'] for b in single['budget'] if b['category'] == 'Dancecloud fees'][0]
    fee_pack = [b['amount'] for b in pack['budget'] if b['category'] == 'Dancecloud fees'][0]
    assert fee_pack < fee_single
    # Level 2 revenue is unchanged by how the tickets are packaged
    rev_single = [b['amount'] for b in single['budget'] if b['category'] == 'Level 2 classes'][0]
    rev_pack = [b['amount'] for b in pack['budget'] if b['category'] == 'Level 2 classes'][0]
    assert rev_single == rev_pack


# ---------------------------------------------------------------- defaults CSV loading
def test_coerce_infers_types():
    assert forecast._coerce('') is None
    assert forecast._coerce('True') is True and forecast._coerce('False') is False
    assert forecast._coerce('2026') == 2026 and isinstance(forecast._coerce('2026'), int)
    assert forecast._coerce('0.95') == 0.95 and isinstance(forecast._coerce('0.95'), float)
    assert forecast._coerce('5.0') == 5.0 and isinstance(forecast._coerce('5.0'), float)


def test_load_defaults_reads_csv(tmp_path, monkeypatch):
    path = tmp_path / 'forecast_defaults.csv'
    path.write_text('key,value\nforecast_ay,2026\nhave_weekender,True\nclass_size_cap,\ndc_cap,5.0\n', encoding='utf-8')
    monkeypatch.setattr('esds_apps.forecast.config.FORECAST_DEFAULTS_PATH', path)
    d = forecast._load_defaults()
    assert d == {'forecast_ay': 2026, 'have_weekender': True, 'class_size_cap': None, 'dc_cap': 5.0}


def test_load_defaults_missing_raises(tmp_path, monkeypatch):
    monkeypatch.setattr('esds_apps.forecast.config.FORECAST_DEFAULTS_PATH', tmp_path / 'nope.csv')
    with pytest.raises(FileNotFoundError):
        forecast._load_defaults()


# ---------------------------------------------------------------- workbook export
def test_to_workbook_is_valid_xlsx(ctx):
    content = forecast.to_workbook({'teacher_rate': 20}, 0.9)
    wb = openpyxl.load_workbook(io.BytesIO(content))
    assert {'summary', 'budget', 'parameters', 'balance'} <= set(wb.sheetnames)
    # The parameters sheet lists a readable label, the value, then the raw key (row = [label, value, key]).
    params = {row[2]: row[1] for row in wb['parameters'].iter_rows(min_row=2, values_only=True)}
    labels = {row[0] for row in wb['parameters'].iter_rows(min_row=2, values_only=True)}
    assert params['teacher_rate'] == 20
    assert 'Teacher pay per hour' in labels


# ---------------------------------------------------------------- routes
@pytest.fixture
def client():
    return TestClient(app, follow_redirects=False)


@pytest.fixture
def auth_client(monkeypatch):
    monkeypatch.setattr('esds_apps.auth._get_authenticated_email', lambda req: 'user@example.com')
    return TestClient(app, follow_redirects=False)


def test_forecast_page_renders(auth_client, monkeypatch):
    monkeypatch.setattr('esds_apps.main.forecast.get_context', lambda: make_context())
    monkeypatch.setattr('esds_apps.main.forecast.predict', lambda: {'net': {'mean': -2000}})
    monkeypatch.setattr('esds_apps.main.forecast.default_params', lambda c: {'teacher_rate': 15})
    resp = auth_client.get('/forecast')
    assert resp.status_code == HTTPStatus.OK
    assert 'forecast.js' in resp.text
    assert 'confidence interval' in resp.text.lower()


def test_forecast_page_unavailable_when_data_missing(auth_client, monkeypatch):
    def _raise():
        raise FileNotFoundError('no defaults')

    monkeypatch.setattr('esds_apps.main.forecast.get_context', _raise)
    resp = auth_client.get('/forecast')
    assert resp.status_code == HTTPStatus.OK
    assert "isn't set up" in resp.text


def test_forecast_predict_json(auth_client, monkeypatch):
    captured = {}

    def fake_predict(params, confidence):
        captured['args'] = (params, confidence)
        return {'net': {'mean': -2081, 'lo': -5500, 'hi': 1400}}

    monkeypatch.setattr('esds_apps.main.forecast.predict', fake_predict)
    resp = auth_client.post('/forecast/predict.json', json={'params': {'teacher_rate': 25}, 'confidence': 0.9})
    assert resp.status_code == HTTPStatus.OK
    assert resp.json()['net']['mean'] == -2081
    assert captured['args'] == ({'teacher_rate': 25}, 0.9)


def test_forecast_predict_requires_auth(client):
    resp = client.post('/forecast/predict.json', json={'params': {}})
    assert resp.status_code == HTTPStatus.UNAUTHORIZED


def test_forecast_predict_db_missing(auth_client, monkeypatch):
    def _raise(params, confidence):
        raise FileNotFoundError('no data')

    monkeypatch.setattr('esds_apps.main.forecast.predict', _raise)
    resp = auth_client.post('/forecast/predict.json', json={'params': {}})
    assert resp.status_code == HTTPStatus.SERVICE_UNAVAILABLE
    assert 'error' in resp.json()


def test_forecast_download_xlsx(auth_client, monkeypatch):
    monkeypatch.setattr('esds_apps.main.forecast.to_workbook', lambda params, confidence: b'PK\x03\x04fakexlsx')
    resp = auth_client.post('/forecast/download.xlsx', json={'params': {}, 'confidence': 0.95})
    assert resp.status_code == HTTPStatus.OK
    assert resp.content == b'PK\x03\x04fakexlsx'
    assert resp.headers['content-disposition'].startswith('attachment; filename=esds_forecast_')


def test_forecast_download_requires_auth(client):
    resp = client.post('/forecast/download.xlsx', json={'params': {}})
    assert resp.status_code == HTTPStatus.UNAUTHORIZED
